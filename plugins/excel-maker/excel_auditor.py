"""Workbook quality auditor."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from excel_formula import explain_formula

REQUIRED_SHEETS = ["README", "Raw_Data", "Clean_Data", "Summary", "Charts", "Formula_Check", "Audit_Log"]


class WorkbookAuditor:
    def audit(self, workbook_path: str | Path, output_path: str | Path | None = None) -> dict[str, Any]:
        try:
            import openpyxl  # type: ignore
        except ImportError as exc:
            raise RuntimeError("openpyxl is required to audit .xlsx files.") from exc

        path = Path(workbook_path)
        wb = openpyxl.load_workbook(path, data_only=False)
        items: list[dict[str, Any]] = []
        for sheet in REQUIRED_SHEETS:
            if sheet not in wb.sheetnames:
                items.append(
                    {
                        "severity": "error",
                        "category": "structure",
                        "message": f"Required sheet is missing: {sheet}",
                        "location": sheet,
                        "suggestion": "Regenerate the workbook with the default report plan.",
                    }
                )
        for ws in wb.worksheets:
            if ws.max_row <= 1 and ws.max_column <= 1:
                items.append(
                    {
                        "severity": "warning",
                        "category": "content",
                        "message": f"Sheet appears empty: {ws.title}",
                        "location": ws.title,
                        "suggestion": "Confirm whether this sheet is expected to be empty.",
                    }
                )
            for row in ws.iter_rows():
                for cell in row:
                    value = cell.value
                    if isinstance(value, str) and value.startswith("="):
                        if value.strip() == "=":
                            items.append(
                                {
                                    "severity": "error",
                                    "category": "formula",
                                    "message": "Empty formula detected",
                                    "location": f"{ws.title}!{cell.coordinate}",
                                    "suggestion": "Replace with a valid Excel formula.",
                                }
                            )
                        elif "#REF!" in value.upper():
                            items.append(
                                {
                                    "severity": "error",
                                    "category": "formula",
                                    "message": "Formula contains #REF reference",
                                    "location": f"{ws.title}!{cell.coordinate}",
                                    "suggestion": explain_formula(value),
                                }
                            )
        if not items:
            items.append(
                {
                    "severity": "info",
                    "category": "quality",
                    "message": "Workbook passed basic structure and formula audit.",
                    "location": str(path.name),
                    "suggestion": "Open the workbook and review business口径 before delivery.",
                }
            )
        result = {"workbook_path": str(path), "items": items, "ok": not any(i["severity"] == "error" for i in items)}
        if output_path:
            Path(output_path).write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
        return result

