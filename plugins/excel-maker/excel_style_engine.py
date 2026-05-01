"""Workbook style helpers implemented with openpyxl."""

from __future__ import annotations

from typing import Any


def _openpyxl_styles():
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    return Alignment, Border, Font, PatternFill, Side, get_column_letter


def estimate_width(value: Any) -> int:
    text = str(value or "")
    width = 0
    for char in text:
        width += 2 if ord(char) > 127 else 1
    return min(max(width + 2, 10), 48)


def apply_table_style(ws, *, header_row: int = 1, brand_color: str = "2563EB") -> None:
    Alignment, Border, Font, PatternFill, Side, get_column_letter = _openpyxl_styles()
    color = brand_color.replace("#", "").upper()
    header_fill = PatternFill("solid", fgColor=color)
    zebra_fill = PatternFill("solid", fgColor="EFF6FF")
    border = Border(bottom=Side(style="thin", color="CBD5E1"))
    for cell in ws[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1), start=header_row + 1):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if row_idx % 2 == 0:
                cell.fill = zebra_fill
    ws.freeze_panes = f"A{header_row + 1}"
    if ws.max_row >= header_row and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions
    autofit_columns(ws)


def autofit_columns(ws) -> None:
    _, _, _, _, _, get_column_letter = _openpyxl_styles()
    for col_idx in range(1, ws.max_column + 1):
        max_width = 10
        for row_idx in range(1, min(ws.max_row, 200) + 1):
            max_width = max(max_width, estimate_width(ws.cell(row=row_idx, column=col_idx).value))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_width


def apply_title_style(ws) -> None:
    Alignment, _, Font, PatternFill, _, _ = _openpyxl_styles()
    ws["A1"].font = Font(bold=True, size=16, color="0F172A")
    ws["A1"].fill = PatternFill("solid", fgColor="DBEAFE")
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 28
    autofit_columns(ws)

