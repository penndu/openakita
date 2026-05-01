"""Template diagnostics for uploaded .xlsx workbooks."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any


class TemplateDiagnosticError(RuntimeError):
    pass


class TemplateManager:
    def diagnose(self, template_path: str | Path, output_path: str | Path | None = None) -> dict[str, Any]:
        try:
            import openpyxl  # type: ignore
        except ImportError as exc:
            raise TemplateDiagnosticError("openpyxl is required to diagnose templates.") from exc

        path = Path(template_path)
        if not path.is_file():
            raise TemplateDiagnosticError(f"Template not found: {path}")
        wb = openpyxl.load_workbook(path, data_only=False)
        result = {
            "path": str(path),
            "sheets": [
                {
                    "name": ws.title,
                    "row_count": ws.max_row,
                    "column_count": ws.max_column,
                    "freeze_panes": str(ws.freeze_panes or ""),
                    "merged_ranges": [str(rng) for rng in ws.merged_cells.ranges],
                    "formula_count": sum(
                        1
                        for row in ws.iter_rows()
                        for cell in row
                        if isinstance(cell.value, str) and cell.value.startswith("=")
                    ),
                }
                for ws in wb.worksheets
            ],
            "defined_names": [defined_name.name for defined_name in wb.defined_names.values()],
        }
        if output_path:
            Path(output_path).write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
        return result

