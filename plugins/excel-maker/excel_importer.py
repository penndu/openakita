"""Workbook and CSV importer for excel-maker."""

from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

SUPPORTED_EXTENSIONS = {".xlsx", ".csv", ".tsv"}
MAX_PREVIEW_ROWS = 100
MAX_PREVIEW_COLS = 50


class WorkbookImportError(RuntimeError):
    pass


@dataclass
class ImportedWorkbook:
    source_path: Path
    imported_path: Path
    profile_path: Path
    sheets: list[dict[str, Any]]
    preview: dict[str, Any]
    warnings: list[str]


def _clean_cell(value: Any) -> Any:
    if isinstance(value, str):
        return "".join(ch for ch in value if ch == "\t" or ch == "\n" or ord(ch) >= 32)
    return value


def _column_letter(index: int) -> str:
    index = max(index, 1)
    letters = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def _try_openpyxl():
    try:
        import openpyxl  # type: ignore

        return openpyxl
    except ImportError as exc:
        raise WorkbookImportError(
            "openpyxl is required to import .xlsx files. Install dependency group table_core."
        ) from exc


def detect_delimiter(sample: str, suffix: str) -> str:
    if suffix == ".tsv":
        return "\t"
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except csv.Error:
        return ","


def detect_header_row(rows: list[list[Any]], max_scan: int = 10) -> int | None:
    best_index: int | None = None
    best_score = -1
    for idx, row in enumerate(rows[:max_scan]):
        non_empty = [cell for cell in row if str(cell or "").strip()]
        if not non_empty:
            continue
        stringish = sum(1 for cell in non_empty if isinstance(cell, str))
        unique = len({str(cell).strip() for cell in non_empty})
        score = len(non_empty) + stringish + unique
        if score > best_score:
            best_score = score
            best_index = idx + 1
    return best_index


class WorkbookImporter:
    def __init__(self, data_root: str | Path, *, workbooks_root: str | Path | None = None) -> None:
        self._data_root = Path(data_root)
        self._workbooks_root = Path(workbooks_root) if workbooks_root else self._data_root / "workbooks"

    def import_file(self, source_path: str | Path, workbook_id: str) -> ImportedWorkbook:
        source = Path(source_path).expanduser().resolve()
        if not source.is_file():
            raise WorkbookImportError(f"Workbook file not found: {source}")
        suffix = source.suffix.lower()
        if suffix not in SUPPORTED_EXTENSIONS:
            raise WorkbookImportError(f"Unsupported workbook type: {suffix}")

        target_dir = self._workbooks_root / workbook_id
        target_dir.mkdir(parents=True, exist_ok=True)
        imported_path = target_dir / source.name
        if imported_path != source:
            imported_path.write_bytes(source.read_bytes())

        if suffix == ".xlsx":
            sheets, preview, warnings = self._inspect_xlsx(imported_path)
        else:
            sheets, preview, warnings = self._inspect_csv(imported_path)

        profile = {
            "workbook_id": workbook_id,
            "source_path": str(source),
            "imported_path": str(imported_path),
            "sheets": sheets,
            "preview": preview,
            "warnings": warnings,
        }
        profile_path = target_dir / "import_profile.json"
        profile_path.write_text(json.dumps(profile, ensure_ascii=False, indent=2), encoding="utf-8")
        return ImportedWorkbook(source, imported_path, profile_path, sheets, preview, warnings)

    def preview(self, profile_path: str | Path, sheet_name: str | None = None) -> dict[str, Any]:
        profile = json.loads(Path(profile_path).read_text(encoding="utf-8"))
        preview = profile.get("preview", {})
        if sheet_name:
            return {sheet_name: preview.get(sheet_name, {"headers": [], "rows": []})}
        return preview

    def _inspect_xlsx(self, path: Path) -> tuple[list[dict[str, Any]], dict[str, Any], list[str]]:
        openpyxl = _try_openpyxl()
        wb = openpyxl.load_workbook(path, data_only=False, read_only=False)
        sheets: list[dict[str, Any]] = []
        preview: dict[str, Any] = {}
        warnings: list[str] = []
        for index, ws in enumerate(wb.worksheets):
            rows = [
                [_clean_cell(cell.value) for cell in row[:MAX_PREVIEW_COLS]]
                for row in ws.iter_rows(max_row=MAX_PREVIEW_ROWS)
            ]
            header_row = detect_header_row(rows)
            headers = rows[header_row - 1] if header_row and header_row <= len(rows) else []
            data_rows = rows[header_row:] if header_row else rows
            formula_count = 0
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_count += 1
            hidden_rows = sum(1 for dim in ws.row_dimensions.values() if dim.hidden)
            hidden_cols = sum(1 for dim in ws.column_dimensions.values() if dim.hidden)
            sheet = {
                "name": ws.title,
                "index": index,
                "row_count": ws.max_row or 0,
                "column_count": ws.max_column or 0,
                "header_row": header_row,
                "data_range": ws.calculate_dimension(),
                "formula_count": formula_count,
                "merged_range_count": len(ws.merged_cells.ranges),
                "hidden_row_count": hidden_rows,
                "hidden_column_count": hidden_cols,
                "metadata": {"has_auto_filter": bool(ws.auto_filter and ws.auto_filter.ref)},
            }
            sheets.append(sheet)
            preview[ws.title] = {
                "headers": headers,
                "rows": data_rows[:MAX_PREVIEW_ROWS],
                "truncated": (ws.max_row or 0) > MAX_PREVIEW_ROWS or (ws.max_column or 0) > MAX_PREVIEW_COLS,
            }
            if ws.max_row > 50000:
                warnings.append(f"Sheet {ws.title} is large; previews and AI context will be sampled.")
        return sheets, preview, warnings

    def _inspect_csv(self, path: Path) -> tuple[list[dict[str, Any]], dict[str, Any], list[str]]:
        raw = path.read_text(encoding="utf-8-sig", errors="replace")
        delimiter = detect_delimiter(raw[:4096], path.suffix.lower())
        reader = csv.reader(raw.splitlines(), delimiter=delimiter)
        rows = [[_clean_cell(cell) for cell in row[:MAX_PREVIEW_COLS]] for row in reader]
        header_row = detect_header_row(rows)
        headers = rows[header_row - 1] if header_row and header_row <= len(rows) else []
        data_rows = rows[header_row:] if header_row else rows
        column_count = max((len(row) for row in rows), default=0)
        end_column = _column_letter(column_count)
        end_row = max(len(rows), 1)
        sheet_name = "CSV_Data"
        sheets = [
            {
                "name": sheet_name,
                "index": 0,
                "row_count": len(rows),
                "column_count": column_count,
                "header_row": header_row,
                "data_range": f"A1:{end_column}{end_row}",
                "formula_count": 0,
                "merged_range_count": 0,
                "hidden_row_count": 0,
                "hidden_column_count": 0,
                "metadata": {"delimiter": delimiter},
            }
        ]
        warnings = []
        if len(rows) > 50000:
            warnings.append("CSV file is large; previews and AI context will be sampled.")
        return (
            sheets,
            {
                sheet_name: {
                    "headers": headers,
                    "rows": data_rows[:MAX_PREVIEW_ROWS],
                    "truncated": len(rows) > MAX_PREVIEW_ROWS or column_count > MAX_PREVIEW_COLS,
                }
            },
            warnings,
        )

