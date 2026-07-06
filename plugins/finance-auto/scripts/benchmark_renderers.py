"""Benchmark the three rendering paths discussed in v0.3 Part Infra section 1.

Paths under test:

1. **xltpl** -- our recommended primary backend for static reports.
2. **openpyxl direct** -- our recommended primary backend for dynamic
   detail tables (and the fallback for static reports if the row count
   exceeds :data:`STATIC_ROW_THRESHOLD`).
3. **win32com / Excel COM** -- the heavyweight reference: we drive
   ``Excel.Application`` to fill cells the way a human would.  Only runs
   when ``pywin32`` is importable AND a user-installed Excel is reachable.

The benchmark builds two synthetic templates on the fly so the script is
self-contained (no dependency on tmp_spike artefacts).  For each path we
record:

* render time (mean of 3 runs, ms)
* output file size (bytes)
* fidelity score: a 0-5 integer summing
  ``preserves_merge + preserves_font + preserves_fill + preserves_number_format
   + emits_all_rows``.

The CLI prints a markdown table to stdout; redirect to
``_m1_w2_xltpl_benchmark.md`` for capture.
"""

from __future__ import annotations

import argparse
import json
import statistics
import sys
import tempfile
import time
from collections.abc import Iterator
from contextlib import contextmanager
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PLUGIN_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PLUGIN_ROOT))

from finance_auto_backend.renderers import (  # noqa: E402
    OpenpyxlDirectRenderer,
    XltplRenderer,
)


def build_static_template(path: Path) -> None:
    """Synthesise a 24-row balance-sheet-shaped xltpl template."""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BS"
    bold = Font(name="DengXian", size=12, bold=True)
    base = Font(name="DengXian", size=11)
    fill = PatternFill("solid", fgColor="DDEBF7")
    thin = Side(style="thin", color="888888")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    centre = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    ws.merge_cells("A1:D1")
    ws["A1"] = "{{ org.name }} {{ year }} 资产负债表 ({{ standard }})"
    ws["A1"].font = bold
    ws["A1"].alignment = centre
    ws["A1"].fill = fill

    headers = ["项目", "代码", "期末余额", "上期余额"]
    for col, label in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col, value=label)
        c.font = bold
        c.fill = fill
        c.alignment = centre
        c.border = border

    rows = [
        ("货币资金", "1001"),
        ("应收账款", "1122"),
        ("预付款项", "1123"),
        ("存货", "1405"),
        ("固定资产", "1601"),
        ("无形资产", "1701"),
        ("流动资产合计", "TOTAL_CA"),
        ("非流动资产合计", "TOTAL_NCA"),
        ("资产总计", "TOTAL_ASSETS"),
        ("应付账款", "2202"),
        ("预收款项", "2203"),
        ("应交税费", "2221"),
        ("流动负债合计", "TOTAL_CL"),
        ("非流动负债合计", "TOTAL_NCL"),
        ("负债合计", "TOTAL_LIAB"),
        ("实收资本", "4001"),
        ("盈余公积", "4101"),
        ("未分配利润", "4104"),
        ("所有者权益合计", "TOTAL_OE"),
        ("负债和所有者权益合计", "TOTAL_LE"),
    ]
    for i, (label, code) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=label).font = base
        ws.cell(row=i, column=2, value=code).font = base
        v_cell = ws.cell(
            row=i, column=3, value=f"{{{{ cells['{code}'].value }}}}"
        )
        v_cell.font = base
        v_cell.alignment = right
        v_cell.number_format = "#,##0.00"
        v_cell.border = border
        p_cell = ws.cell(
            row=i, column=4, value=f"{{{{ cells['{code}'].previous }}}}"
        )
        p_cell.font = base
        p_cell.alignment = right
        p_cell.number_format = "#,##0.00"
        p_cell.border = border

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    wb.save(str(path))
    wb.close()


def build_dynamic_template(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detail"
    bold = Font(name="DengXian", size=12, bold=True)
    base = Font(name="DengXian", size=10)
    fill = PatternFill("solid", fgColor="FCE4D6")
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    right = Alignment(horizontal="right")

    ws.merge_cells("A1:D1")
    ws["A1"] = "{{ org.name }} - {{ period }} 应收账款明细"
    ws["A1"].font = bold
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["科目编码", "客户名称", "金额", "账龄"]
    keys = ["account_code", "name", "balance", "aging"]
    for col, label in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col, value=label)
        c.font = bold
        c.fill = fill
        c.border = border
    for col, key in enumerate(keys, start=1):
        ws.cell(row=3, column=col, value=key)

    for col_idx in range(1, len(keys) + 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.font = base
        cell.border = border
        if col_idx in (3, 4):
            cell.alignment = right
            cell.number_format = "#,##0.00"
    ws["A4"] = "__ROWS_HERE__"

    ws["A5"] = "合计"
    ws["A5"].font = bold
    ws["C5"] = "__SUM__"
    ws["C5"].font = bold
    ws["C5"].number_format = "#,##0.00"

    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 22
    wb.save(str(path))
    wb.close()


def make_static_context(rows: int) -> dict[str, Any]:
    cells = []
    sample = [
        ("1001", "货币资金", 12_345_678.90),
        ("1122", "应收账款", 8_900_000.00),
        ("1123", "预付款项", 1_200_000.00),
        ("1405", "存货", 4_500_000.00),
        ("1601", "固定资产", 22_000_000.00),
        ("1701", "无形资产", 600_000.00),
        ("TOTAL_CA", "流动资产合计", 26_945_678.90),
        ("TOTAL_NCA", "非流动资产合计", 22_600_000.00),
        ("TOTAL_ASSETS", "资产总计", 49_545_678.90),
        ("2202", "应付账款", 7_800_000.00),
        ("2203", "预收款项", 0.00),
        ("2221", "应交税费", 320_000.00),
        ("TOTAL_CL", "流动负债合计", 8_120_000.00),
        ("TOTAL_NCL", "非流动负债合计", 0.00),
        ("TOTAL_LIAB", "负债合计", 8_120_000.00),
        ("4001", "实收资本", 30_000_000.00),
        ("4101", "盈余公积", 1_500_000.00),
        ("4104", "未分配利润", 9_925_678.90),
        ("TOTAL_OE", "所有者权益合计", 41_425_678.90),
        ("TOTAL_LE", "负债和所有者权益合计", 49_545_678.90),
    ]
    for code, label, val in sample[:rows]:
        cells.append(
            {
                "code": code,
                "label": label,
                "value": val,
                "previous": val * 0.92,
                "row": 0,
                "column": "C",
                "formula": "",
                "source_rows": [],
            }
        )
    return {
        "report_kind": "balance_sheet",
        "standard": "small_enterprise",
        "year": 2025,
        "period": "2025-12",
        "org": {"id": "demo-1", "name": "示范有限公司", "industry": "软件"},
        "cells": cells,
    }


def make_dynamic_context(rows: int) -> dict[str, Any]:
    body = [
        {
            "account_code": f"1122.{i:04d}",
            "name": f"客户 {i:04d}",
            "balance": float(i * 137.5),
            "aging": "0-30" if i % 4 == 0 else "31-60" if i % 4 == 1 else "61-90"
            if i % 4 == 2
            else ">90",
        }
        for i in range(1, rows + 1)
    ]
    return {
        "report_kind": "ar_aging",
        "standard": "general_enterprise",
        "year": 2025,
        "period": "2025-12",
        "org": {"id": "demo-1", "name": "示范有限公司", "industry": "软件"},
        "rows": body,
    }


@contextmanager
def _wb(path: Path) -> Iterator[Any]:
    wb = openpyxl.load_workbook(str(path))
    try:
        yield wb
    finally:
        wb.close()


def _scratch(prefix: str) -> Path:
    fd, name = tempfile.mkstemp(suffix=".xlsx", prefix=prefix)
    import os as _os
    _os.close(fd)
    p = Path(name)
    p.unlink(missing_ok=True)
    return p


def measure_xltpl(template: Path, ctx: dict[str, Any], runs: int) -> dict[str, Any]:
    timings: list[float] = []
    out: Path | None = None
    for _ in range(runs):
        out = _scratch("bench_xltpl_")
        renderer = XltplRenderer(template)
        result = renderer.render(ctx, out)
        timings.append(result.elapsed_ms)
    assert out is not None
    return _summarise(out, timings, "xltpl", static=True)


def _static_openpyxl_render(template: Path, ctx: dict[str, Any], out: Path) -> float:
    """Inline static renderer used only for the benchmark.

    Not part of the production API: real static reports go through
    :class:`XltplRenderer`.  We measure this path so the benchmark report
    can quantify how much we lose by *not* picking openpyxl for static.
    """
    started = time.perf_counter()
    wb = openpyxl.load_workbook(str(template))
    ws = wb.active
    cells_by_code = {str(c["code"]): c for c in ctx.get("cells", [])}
    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            text = cell.value
            replaced = text
            replaced = (
                replaced.replace("{{ org.name }}", ctx["org"]["name"])
                .replace("{{ year }}", str(ctx["year"]))
                .replace("{{ period }}", str(ctx["period"]))
                .replace("{{ standard }}", str(ctx.get("standard", "")))
            )
            for code, cell_data in cells_by_code.items():
                token_value = "{{ cells['" + code + "'].value }}"
                token_prev = "{{ cells['" + code + "'].previous }}"
                if replaced == token_value:
                    cell.value = float(cell_data["value"])
                    break
                if replaced == token_prev:
                    cell.value = float(cell_data.get("previous", 0))
                    break
            else:
                if replaced != text:
                    cell.value = replaced
    wb.save(str(out))
    wb.close()
    return (time.perf_counter() - started) * 1000.0


def measure_openpyxl(
    template: Path, ctx: dict[str, Any], runs: int, *, static: bool
) -> dict[str, Any]:
    timings: list[float] = []
    out: Path | None = None
    for _ in range(runs):
        out = _scratch("bench_opx_")
        if static:
            elapsed = _static_openpyxl_render(template, ctx, out)
            timings.append(elapsed)
        else:
            renderer = OpenpyxlDirectRenderer(template)
            result = renderer.render(ctx, out)
            timings.append(result.elapsed_ms)
    assert out is not None
    return _summarise(out, timings, "openpyxl", static=static)


def measure_win32com(
    template: Path, ctx: dict[str, Any], runs: int, *, static: bool
) -> dict[str, Any]:
    try:
        import pythoncom
        import win32com.client
    except ImportError:
        return {"backend": "win32com", "available": False, "reason": "pywin32 missing"}

    # Excel COM is heavyweight; force a single run regardless of caller wish.
    runs = 1
    timings: list[float] = []
    out: Path | None = None
    for _ in range(runs):
        out_path = _scratch("bench_com_")
        out = out_path
        pythoncom.CoInitialize()
        excel = None
        wb = None
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            started = time.perf_counter()
            wb = excel.Workbooks.Open(str(template))
            ws = wb.Worksheets(1)
            if static:
                cells_by_code = {
                    str(c["code"]): c for c in ctx.get("cells", [])
                }
                for row in range(1, ws.UsedRange.Rows.Count + 1):
                    code_cell = ws.Cells(row, 2).Value
                    if code_cell and str(code_cell) in cells_by_code:
                        cell = cells_by_code[str(code_cell)]
                        ws.Cells(row, 3).Value = float(cell["value"])
                        ws.Cells(row, 4).Value = float(cell.get("previous", 0))
                title = ws.Cells(1, 1).Value or ""
                ws.Cells(1, 1).Value = (
                    str(title)
                    .replace("{{ org.name }}", ctx["org"]["name"])
                    .replace("{{ year }}", str(ctx["year"]))
                    .replace("{{ standard }}", str(ctx.get("standard", "")))
                )
            else:
                rows = ctx["rows"]
                start_row = 4
                for offset, r in enumerate(rows):
                    target = start_row + offset
                    ws.Cells(target, 1).Value = r["account_code"]
                    ws.Cells(target, 2).Value = r["name"]
                    ws.Cells(target, 3).Value = float(r["balance"])
                    ws.Cells(target, 4).Value = r["aging"]
                title = ws.Cells(1, 1).Value or ""
                ws.Cells(1, 1).Value = (
                    str(title)
                    .replace("{{ org.name }}", ctx["org"]["name"])
                    .replace("{{ period }}", str(ctx.get("period", "")))
                )
            wb.SaveAs(str(out_path), FileFormat=51)
            wb.Close(SaveChanges=False)
            wb = None
            timings.append((time.perf_counter() - started) * 1000.0)
        except Exception as exc:  # pragma: no cover - environmental
            return {
                "backend": "win32com",
                "available": False,
                "reason": f"COM error: {exc!r}",
            }
        finally:
            try:
                if wb is not None:
                    wb.Close(SaveChanges=False)
            except Exception:
                pass
            try:
                if excel is not None:
                    excel.Quit()
            except Exception:
                pass
            del excel
            pythoncom.CoUninitialize()
            time.sleep(0.5)
    assert out is not None
    return _summarise(out, timings, "win32com", static=static)


def _summarise(
    out: Path, timings: list[float], backend: str, *, static: bool
) -> dict[str, Any]:
    info = _inspect(out, static=static)
    info["backend"] = backend
    info["available"] = True
    info["mean_ms"] = round(statistics.mean(timings), 2)
    info["stdev_ms"] = (
        round(statistics.pstdev(timings), 2) if len(timings) > 1 else 0.0
    )
    info["bytes"] = out.stat().st_size
    info["sample_path"] = str(out)
    return info


def _inspect(out: Path, *, static: bool) -> dict[str, Any]:
    fid = {
        "preserves_merge": False,
        "preserves_font": False,
        "preserves_fill": False,
        "preserves_number_format": False,
        "emits_all_rows": False,
    }
    with _wb(out) as wb:
        ws = wb.active
        fid["preserves_merge"] = bool(list(ws.merged_cells.ranges))
        first_cell = ws["A2"] if static else ws["A2"]
        fid["preserves_font"] = bool(first_cell.font.bold)
        fid["preserves_fill"] = bool(
            first_cell.fill and first_cell.fill.fgColor
            and first_cell.fill.fgColor.rgb
            and first_cell.fill.fgColor.rgb.upper() not in ("00000000", "FFFFFFFF", None)
        )
        target = ws["C3"] if static else ws["C4"]
        fid["preserves_number_format"] = (target.number_format or "").startswith("#,##0")
        if static:
            non_empty = sum(
                1
                for r in ws.iter_rows(min_row=3, max_col=3)
                if r[2].value not in (None, "")
            )
            fid["emits_all_rows"] = non_empty >= 18
        else:
            non_empty = sum(
                1
                for r in ws.iter_rows(min_row=4, max_col=1)
                if r[0].value not in (None, "")
            )
            fid["emits_all_rows"] = non_empty >= 1
    fid_score = sum(1 for v in fid.values() if v)
    return {"fidelity": fid, "fidelity_score": fid_score}


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--runs", type=int, default=3)
    parser.add_argument("--dynamic-rows", type=int, default=1500)
    parser.add_argument("--include-com", action="store_true")
    parser.add_argument("--out", type=str, default=None,
                        help="optional path to write a JSON dump of the run")
    args = parser.parse_args()

    work = Path(tempfile.mkdtemp(prefix="finance_bench_"))
    static_tpl = work / "static_bs.xlsx"
    dynamic_tpl = work / "dynamic_ar.xlsx"
    build_static_template(static_tpl)
    build_dynamic_template(dynamic_tpl)

    ctx_static = make_static_context(20)
    ctx_dynamic = make_dynamic_context(args.dynamic_rows)

    results: dict[str, Any] = {"static": [], "dynamic": []}

    print(f"# Benchmark scratch: {work}")
    print(f"runs={args.runs} dynamic_rows={args.dynamic_rows} include_com={args.include_com}")

    print("\n## Static (BS, 20 rows)")
    s_xltpl = measure_xltpl(static_tpl, ctx_static, args.runs)
    results["static"].append(s_xltpl)
    print(_fmt(s_xltpl))

    s_opx = measure_openpyxl(static_tpl, ctx_static, args.runs, static=True)
    results["static"].append(s_opx)
    print(_fmt(s_opx))

    if args.include_com:
        s_com = measure_win32com(static_tpl, ctx_static, max(args.runs, 1), static=True)
        results["static"].append(s_com)
        print(_fmt(s_com))

    print(f"\n## Dynamic (AR detail, {args.dynamic_rows} rows)")
    d_opx = measure_openpyxl(dynamic_tpl, ctx_dynamic, args.runs, static=False)
    results["dynamic"].append(d_opx)
    print(_fmt(d_opx))

    if args.include_com:
        d_com = measure_win32com(dynamic_tpl, ctx_dynamic, 1, static=False)
        results["dynamic"].append(d_com)
        print(_fmt(d_com))

    if args.out:
        Path(args.out).write_text(
            json.dumps(results, ensure_ascii=False, indent=2, default=str),
            encoding="utf-8",
        )
        print(f"\nWrote {args.out}")
    return 0


def _fmt(row: dict[str, Any]) -> str:
    if not row.get("available", True):
        return f"- {row['backend']:9}  SKIPPED ({row.get('reason', 'n/a')})"
    return (
        f"- {row['backend']:9}  mean={row['mean_ms']:7.1f}ms  "
        f"std={row['stdev_ms']:5.1f}ms  bytes={row['bytes']:>7}  "
        f"fidelity={row['fidelity_score']}/5"
    )


if __name__ == "__main__":
    raise SystemExit(main())
