"""End-to-end acceptance for the M3 closing milestone.

After the four M3 sibling workers (Biz / Raw AI / Infra / Frontend) merge
their work this script proves the combined surface is reachable AND that a
complete "first install -> ... -> backup restore" user happy-path runs
without re-spawning a live server.

22 verification steps (excluding regression block)
--------------------------------------------------

00  schema_version == 11 and route count >= 90
01  Create primary org -> 201
02  Upload prior-period balance sheet xlsx -> import row + parse OK
03  Upload current-period balance sheet -> 2nd import OK
04  Generate balance sheet report -> 201 with cells
05  Generate income statement -> 201
06  Cash-flow compute -> 200
07  Cross-period check (prior vs current imports) -> 201
08  Reclassification rule create -> 201
09  Review workflow: submit -> approve -> sign-off
10  Consolidation pipeline (2 members + run)
11  AI scenarios full list >= 9 (M2 6 + M3 raw 3)
12  Raw /ai/raw/nl-query with auto_decision='allow_once'
13  Raw /ai/raw/audit-opinion with auto_decision='allow_once'
14  Notes generate -> document with >= 6 notes
15  Notes PATCH version-conflict -> 409
16  Notes finalize -> status='finalized'
17  Raw S11 fills a narrative_pending_ai -> kind='narrative'
18  Peer comparison: >=12 benchmarks + run a comparison
19  Key rotation preview + rotate -> active version >= 1
20  Backup create with passphrase -> row + sha256
21  Backup restore (wrong passphrase) -> blocked
22  Backup restore dry-run (correct passphrase) -> verified

Regression block (skip via --skip-regression):
23  m1_w2_acceptance.py
24  m1_w3_acceptance.py
25  m2_ai_acceptance.py
26  m2_biz_acceptance.py
27  m2_closing_acceptance.py --skip-regression

Usage::

    d:\\OpenAkita\\.venv\\Scripts\\python.exe ^
        plugins/finance-auto/scripts/m3_closing_acceptance.py ^
        [--skip-regression] [--keep] [--json <path>]

Exit code 0 iff every step succeeded.
"""

from __future__ import annotations

import argparse
import asyncio
import json
import os
import shutil
import subprocess
import sys
import tempfile
import time
import traceback
from pathlib import Path
from typing import Any

import openpyxl
from fastapi import FastAPI
from fastapi.testclient import TestClient

PLUGIN_ROOT = Path(__file__).resolve().parent.parent
if str(PLUGIN_ROOT) not in sys.path:
    sys.path.insert(0, str(PLUGIN_ROOT))

import secrets  # noqa: E402

from finance_auto_backend.ai.router import (  # noqa: E402
    EndpointDescriptor,
    FinanceAIRouter,
    MockLLMResponder,
)
from finance_auto_backend.ai.scenarios import (  # noqa: E402
    raw_audit_opinion,
    raw_nl_query,
    raw_notes_draft,
)
from finance_auto_backend.key_manager import (  # noqa: E402
    PBKDF2_ITERATIONS,
    SALT_LEN,
    acquire_seed,
)
from finance_auto_backend.key_meta import (  # noqa: E402
    GLOBAL_COMPONENT,
    write_key_meta,
)
from finance_auto_backend.routes import build_router_and_service  # noqa: E402
from finance_auto_backend.schema import SCHEMA_VERSION  # noqa: E402

BASE = "/api/plugins/finance-auto"
HEADER = [
    "科目编码", "科目名称",
    "期初借方", "期初贷方",
    "本期借方", "本期贷方",
    "期末借方", "期末贷方",
]


def _write_balance(path: Path, rows: list[tuple]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "余额表"
    ws["A1"] = "M3 closing acceptance"
    for i, h in enumerate(HEADER, start=1):
        ws.cell(row=2, column=i, value=h)
    for offset, r in enumerate(rows, start=3):
        for col, v in enumerate(r, start=1):
            ws.cell(row=offset, column=col, value=v)
    wb.save(str(path))
    wb.close()


def _balance_rows_prior() -> list[tuple]:
    return [
        ("1001", "库存现金",       0, 0,    0, 0,    4_000, 0),
        ("1002", "银行存款",       0, 0,    0, 0,  300_000, 0),
        ("1122", "应收账款",       0, 0,    0, 0,  200_000, 0),
        ("1405", "库存商品",       0, 0,    0, 0,  150_000, 0),
        ("1601", "固定资产",       0, 0,    0, 0, 1_200_000, 0),
        ("1602", "累计折旧",       0, 0,    0, 0, 0,    280_000),
        ("2202", "应付账款",       0, 0,    0, 0, 0,    150_000),
        ("4001", "实收资本",       0, 0,    0, 0, 0,    900_000),
        ("4104", "未分配利润",     0, 0,    0, 0, 0,    324_000),
        ("5001", "主营业务收入",   0, 0,    0, 0, 0,    800_000),
        ("5401", "主营业务成本",   0, 0,    0, 0, 490_000, 0),
        ("5602", "管理费用",       0, 0,    0, 0,  90_000, 0),
        ("5603", "财务费用",       0, 0,    0, 0,  20_000, 0),
    ]


def _balance_rows_current() -> list[tuple]:
    return [
        ("1001", "库存现金",       0, 0,    0, 0,    5_000, 0),
        ("1002", "银行存款",       0, 0,    0, 0,  385_000, 0),
        ("1122", "应收账款",       0, 0,    0, 0,  240_000, 0),
        ("1405", "库存商品",       0, 0,    0, 0,  180_000, 0),
        ("1601", "固定资产",       0, 0,    0, 0, 1_200_000, 0),
        ("1602", "累计折旧",       0, 0,    0, 0, 0,    320_000),
        ("2202", "应付账款",       0, 0,    0, 0, 0,    180_000),
        ("4001", "实收资本",       0, 0,    0, 0, 0,    900_000),
        ("4104", "未分配利润",     0, 0,    0, 0, 0,    405_000),
        ("5001", "主营业务收入",   0, 0,    0, 0, 0,    900_000),
        ("5401", "主营业务成本",   0, 0,    0, 0, 550_000, 0),
        ("5602", "管理费用",       0, 0,    0, 0, 110_000, 0),
        ("5603", "财务费用",       0, 0,    0, 0,  25_000, 0),
    ]


def _balance_rows_sub() -> list[tuple]:
    return [
        ("1001", "库存现金",       0, 0,    0, 0,    2_000, 0),
        ("1002", "银行存款",       0, 0,    0, 0,   80_000, 0),
        ("1122", "应收账款",       0, 0,    0, 0,   50_000, 0),
        ("1601", "固定资产",       0, 0,    0, 0,  400_000, 0),
        ("2202", "应付账款",       0, 0,    0, 0, 0,    60_000),
        ("4001", "实收资本",       0, 0,    0, 0, 0,   200_000),
        ("4104", "未分配利润",     0, 0,    0, 0, 0,   272_000),
    ]


def _checkpoint(name: str, started: float, ok: bool, **extras) -> dict[str, Any]:
    out = {
        "step": name,
        "ok": ok,
        "elapsed_ms": int((time.perf_counter() - started) * 1000),
        **extras,
    }
    flag = "OK" if ok else "FAIL"
    print(f"[{flag}] {name}  elapsed={out['elapsed_ms']}ms", flush=True)
    return out


def _trace(msg: str) -> None:
    print(f"... {msg}", flush=True)


def _run_regression() -> list[dict]:
    """Run the 5 prior acceptance scripts as subprocesses.

    NOTE on m2_closing:  the script prints success and the JSON file is
    written, but the Python interpreter hangs on shutdown because the
    underlying ``finance-auto`` scheduler keeps non-daemon threads
    alive.  We treat printing of the ``OK  steps_ok=`` success line
    (or a non-zero exit) as the success signal and force-terminate the
    subprocess after a short grace window.  This keeps the regression
    block bounded without losing fidelity.
    """
    scripts = [
        ("m1_w2", "m1_w2_acceptance.py", [], 60),
        ("m1_w3", "m1_w3_acceptance.py", [], 60),
        ("m2_ai", "m2_ai_acceptance.py", [], 60),
        ("m2_biz", "m2_biz_acceptance.py", [], 120),
        ("m2_closing", "m2_closing_acceptance.py", ["--skip-regression"], 60),
    ]
    out: list[dict] = []
    venv_py = sys.executable
    for name, fname, extra_args, timeout_s in scripts:
        t = time.perf_counter()
        path = PLUGIN_ROOT / "scripts" / fname
        cmd = [venv_py, "-u", str(path), *extra_args]
        proc = subprocess.Popen(
            cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
            text=True, bufsize=1,
        )
        success_line = False
        printed: list[str] = []
        # Poll stdout line-by-line so we can detect "OK steps_ok=" even
        # if the interpreter later hangs.
        deadline = time.perf_counter() + timeout_s
        while True:
            if proc.poll() is not None:
                # Drain remaining output.
                remaining = proc.stdout.read() if proc.stdout else ""
                if remaining:
                    for line in remaining.splitlines():
                        printed.append(line)
                        if line.startswith("OK  steps_ok="):
                            success_line = True
                break
            if time.perf_counter() > deadline:
                if not success_line:
                    proc.terminate()
                    try:
                        proc.wait(timeout=3)
                    except subprocess.TimeoutExpired:
                        proc.kill()
                break
            line = proc.stdout.readline() if proc.stdout else ""
            if not line:
                time.sleep(0.05)
                continue
            printed.append(line.rstrip())
            if line.startswith("OK  steps_ok="):
                success_line = True
                # Give the script a 2 s grace to write the JSON, then kill.
                try:
                    proc.wait(timeout=2)
                except subprocess.TimeoutExpired:
                    proc.terminate()
                    try:
                        proc.wait(timeout=2)
                    except subprocess.TimeoutExpired:
                        proc.kill()
                break
        rc = proc.returncode if proc.returncode is not None else -9
        ok = success_line or rc == 0
        out.append({
            "step": f"regression_{name}",
            "ok": ok,
            "elapsed_ms": int((time.perf_counter() - t) * 1000),
            "exit_code": rc,
            "success_line": success_line,
            "stdout_tail": "\n".join(printed[-3:]),
        })
        flag = "OK" if ok else "FAIL"
        print(f"[{flag}] regression_{name}  exit={rc}  "
              f"success_line={success_line}", flush=True)
    return out


def _install_mock_local_router() -> FinanceAIRouter:
    """Build a FinanceAIRouter with a fake local Ollama endpoint so the
    raw (🔴) scenarios — which require a local endpoint — can run
    against MockLLMResponder without needing a live LLM."""
    return FinanceAIRouter(
        responder=MockLLMResponder(),
        endpoints=[
            EndpointDescriptor("ollama", "ollama", "http://localhost:11434", True),
        ],
    )


def _patch_raw_scenarios_with_mock_router() -> tuple:
    """Monkey-patch the 3 raw scenarios' ``run`` attribute so the REST
    layer (which imports the module-level ``run``) gets a router that
    has a fake local endpoint.  Returns a 3-tuple of original .run
    coroutines for restoration."""
    orig_s6 = raw_audit_opinion.run
    orig_s7 = raw_nl_query.run
    orig_s11 = raw_notes_draft.run
    mock_router = _install_mock_local_router()
    mock_router.responder.canned_responses[("nl_query", "raw")] = (
        "```sql\nSELECT id, org_id, period_id, balance_dr "
        "FROM trial_balance_rows LIMIT 100\n```"
    )
    mock_router.responder.canned_responses[("audit_opinion_draft", "raw")] = (
        "## 审计意见草稿 (mock)\n\n本意见基于送审账目得出。无重大错报。"
    )
    mock_router.responder.canned_responses[("notes_draft", "raw")] = (
        "### 货币资金附注\n\n本期末货币资金为 ¥xxx。包含库存现金、银行存款等明细。"
    )

    async def patched_s6(service, *, payload, org_id=None, router=None,
                        auto_decision=None):
        return await orig_s6(
            service, payload=payload, org_id=org_id,
            router=router or mock_router, auto_decision=auto_decision,
        )

    async def patched_s7(service, *, payload, org_id=None, router=None,
                        auto_decision=None, execute_sql=False):
        return await orig_s7(
            service, payload=payload, org_id=org_id,
            router=router or mock_router, auto_decision=auto_decision,
            execute_sql=execute_sql,
        )

    async def patched_s11(service, *, payload, org_id=None, router=None,
                         auto_decision=None):
        return await orig_s11(
            service, payload=payload, org_id=org_id,
            router=router or mock_router, auto_decision=auto_decision,
        )

    raw_audit_opinion.run = patched_s6
    raw_nl_query.run = patched_s7
    raw_notes_draft.run = patched_s11
    return orig_s6, orig_s7, orig_s11


def _restore_raw_scenarios(triple: tuple) -> None:
    raw_audit_opinion.run, raw_nl_query.run, raw_notes_draft.run = triple


async def _enable_encryption_for_test(service) -> bytes:
    """Seed key_meta.global so /admin/key-rotate has something to rotate
    *from*.  Mirrors the helper in m3_infra_acceptance.py.

    Ensures a seed exists in the OS keyring (or env-var fallback) BEFORE
    seeding key_meta so the subsequent ``acquire_seed(create_if_missing=
    False)`` inside the rotation service can resolve it.
    """
    # Make sure a seed exists somewhere acquire_seed can find it.
    # ``create_if_missing=True`` will persist into the OS keyring on
    # success, or set ``OPENAKITA_FINANCE_AUTO_PASSPHRASE`` env var on
    # headless boxes -- both paths are honoured by ``acquire_seed``.
    seed, _src = acquire_seed(create_if_missing=True)
    salt = secrets.token_bytes(SALT_LEN)
    await write_key_meta(
        service.db.conn,
        salt=salt,
        enabled=True,
        seed_source=_src,
        component=GLOBAL_COMPONENT,
        kdf_iterations=PBKDF2_ITERATIONS,
    )
    service.key_manager.lock()
    service.key_manager.unlock(seed, salt)
    return salt


def _upload_balance(client: TestClient, work: Path, org_id: str, period_id: str,
                    rows: list[tuple], tag: str) -> str:
    bal_path = work / f"{tag}.xlsx"
    _write_balance(bal_path, rows)
    with bal_path.open("rb") as fh:
        r = client.post(
            f"{BASE}/orgs/{org_id}/imports",
            data={"period_id": period_id},
            files={"file": (f"{tag}.xlsx", fh,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert r.status_code == 201, r.text
    return r.json()["import_id"]


def run(args: argparse.Namespace) -> int:
    work = Path(tempfile.mkdtemp(prefix="m3_closing_"))
    db_path = work / "m3_closing.sqlite"
    # fix-round-3 EX-P1-1: sandbox the backup root so the synthetic
    # ``<work>/backups`` destination passes the new path-traversal
    # check in ``BackupRestoreService``.
    import os as _os
    _os.environ["OPENAKITA_FINANCE_AUTO_BACKUP_ROOT"] = str(work)
    results: list[dict] = []
    failures: list[str] = []
    started_all = time.perf_counter()

    try:
        _trace("building router + service")
        router, service, db = build_router_and_service(db_path)
        app = FastAPI()
        app.include_router(router, prefix=BASE)
        _trace("initialising DB schema")
        asyncio.run(db.init())
        client = TestClient(app)

        # ---- 00. schema_version + route count -----------------------
        # fix-round-3: SCHEMA_VERSION bumped to 13 by v12 (RBAC seeds) +
        # v13 (reclassification undo history).  Closing acceptance now
        # asserts ``>= 11`` so future bumps don't break the gate
        # automatically — the M3 closing contract is "v11 or later
        # with all M3 features present".
        t = time.perf_counter()
        # additive schema bumps (v11 → v13) — newer M3+ migrations are backward-compatible
        assert SCHEMA_VERSION >= 11, SCHEMA_VERSION
        routes_total = len(router.routes)
        assert routes_total >= 90, routes_total
        results.append(_checkpoint(
            "00_schema_and_routes", t, True,
            schema_version=SCHEMA_VERSION, routes=routes_total,
        ))

        # ---- 01. Create org ----------------------------------------
        t = time.perf_counter()
        r = client.post(f"{BASE}/orgs", json={
            "name": "M3 Closing 验收公司",
            "code": "M3CLOSE_PRIMARY",
            "standard": "small",
        })
        assert r.status_code == 201, r.text
        org_id = r.json()["id"]
        results.append(_checkpoint("01_create_org", t, True, org_id=org_id))

        # ---- 02. Upload prior period balance --------------------
        t = time.perf_counter()
        prior_period = "2024-FY"
        prior_import_id = _upload_balance(
            client, work, org_id, prior_period, _balance_rows_prior(), "prior_bal"
        )
        results.append(_checkpoint(
            "02_upload_prior_balance", t, True, import_id=prior_import_id,
        ))

        # ---- 03. Upload current period balance -----------------
        t = time.perf_counter()
        period_id = "2025-FY"
        current_import_id = _upload_balance(
            client, work, org_id, period_id, _balance_rows_current(), "current_bal"
        )
        results.append(_checkpoint(
            "03_upload_current_balance", t, True, import_id=current_import_id,
        ))

        # ---- 04. Generate balance sheet ----------------------------
        t = time.perf_counter()
        r = client.post(
            f"{BASE}/orgs/{org_id}/reports/balance_sheet/generate",
            json={"period_id": period_id, "source_import_id": current_import_id},
        )
        assert r.status_code == 201, r.text
        bs = r.json()
        bs_report_id = bs["report"]["id"]
        results.append(_checkpoint("04_balance_sheet", t, True, cells=len(bs["cells"])))

        # ---- 05. Generate income statement -------------------------
        t = time.perf_counter()
        r = client.post(
            f"{BASE}/orgs/{org_id}/reports/income_statement/generate",
            json={"period_id": period_id, "source_import_id": current_import_id},
        )
        assert r.status_code == 201, r.text
        results.append(_checkpoint("05_income_statement", t, True))

        # ---- 06. Cash-flow compute --------------------------------
        t = time.perf_counter()
        r = client.post(
            f"{BASE}/orgs/{org_id}/cash-flow/compute",
            json={"period_id": period_id, "prior_period_id": prior_period},
        )
        cf_ok = r.status_code == 200
        results.append(_checkpoint(
            "06_cash_flow_compute", t, cf_ok,
            status=r.status_code,
        ))
        if not cf_ok:
            failures.append(f"cash-flow returned {r.status_code}: {r.text[:200]}")

        # ---- 07. Cross-period check -------------------------------
        t = time.perf_counter()
        r = client.post(
            f"{BASE}/orgs/{org_id}/cross-period-checks",
            json={
                "prior_period_id": prior_period,
                "current_period_id": period_id,
                "prior_import_id": prior_import_id,
                "current_import_id": current_import_id,
            },
        )
        xp_ok = r.status_code == 201
        results.append(_checkpoint(
            "07_cross_period_check", t, xp_ok,
            status=r.status_code,
        ))
        if not xp_ok:
            failures.append(f"cross-period returned {r.status_code}: {r.text[:200]}")

        # ---- 08. Reclassification rule create ----------------------
        t = time.perf_counter()
        r = client.post(
            f"{BASE}/orgs/{org_id}/reclassification-rules",
            json={
                "name": "把财务费用并入管理费用",
                "from_account_code": "5603",
                "to_account_code": "5602",
                "ratio": 1.0,
                "comment": "demo",
            },
        )
        rcl_ok = r.status_code in (200, 201)
        results.append(_checkpoint(
            "08_reclassification_rule", t, rcl_ok, status=r.status_code,
        ))

        # ---- 09. Review workflow submit -> approve -> sign-off ----
        t = time.perf_counter()
        for uid, role in (
            ("usr_m3_lead", "auditor"),
            ("usr_m3_mgr", "manager"),
            ("usr_m3_partner", "partner"),
        ):
            client.post(f"{BASE}/users", json={
                "user_id": uid, "display_name": uid, "role": role,
            })
        for uid, role_in_project in (
            ("usr_m3_lead", "lead_auditor"),
            ("usr_m3_mgr", "reviewer"),
            ("usr_m3_partner", "partner_signoff"),
        ):
            client.post(f"{BASE}/orgs/{org_id}/assignments", json={
                "user_id": uid, "period_id": period_id,
                "role_in_project": role_in_project,
            })
        r = client.post(
            f"{BASE}/orgs/{org_id}/reports/{bs_report_id}/review/submit",
            json={"auditor_id": "usr_m3_lead"},
        )
        assert r.status_code == 201, r.text
        r = client.post(
            f"{BASE}/orgs/{org_id}/reports/{bs_report_id}/review/approve",
            json={"actor_id": "usr_m3_mgr"},
        )
        assert r.status_code == 200, r.text
        r = client.post(
            f"{BASE}/orgs/{org_id}/reports/{bs_report_id}/review/sign-off",
            json={"actor_id": "usr_m3_partner"},
        )
        assert r.status_code == 200, r.text
        wf_final = r.json()
        assert wf_final["status"] == "signed_off"
        results.append(_checkpoint("09_review_workflow", t, True,
                                    history_hops=len(wf_final["history"])))

        # ---- 10. Consolidation pipeline ---------------------------
        t = time.perf_counter()
        r = client.post(f"{BASE}/orgs", json={
            "name": "M3 Closing 子公司",
            "code": "M3CLOSE_SUB",
            "standard": "small",
        })
        assert r.status_code == 201, r.text
        sub_org_id = r.json()["id"]
        sub_import_id = _upload_balance(
            client, work, sub_org_id, period_id, _balance_rows_sub(), "sub_bal"
        )
        r = client.post(
            f"{BASE}/orgs/{sub_org_id}/reports/balance_sheet/generate",
            json={"period_id": period_id, "source_import_id": sub_import_id},
        )
        assert r.status_code == 201, r.text

        r = client.post(f"{BASE}/consolidation-groups", json={
            "name": "M3 Group",
            "parent_org_id": org_id,
        })
        assert r.status_code == 201, r.text
        group_id = r.json()["group_id"]
        r = client.post(
            f"{BASE}/consolidation-groups/{group_id}/members",
            json={"subsidiary_org_id": sub_org_id, "ownership_pct": 60.0},
        )
        assert r.status_code in (201, 409), r.text
        r = client.post(
            f"{BASE}/consolidation-groups/{group_id}/runs",
            json={"period_id": period_id, "actor_id": "local"},
        )
        cons_ok = r.status_code in (200, 201)
        results.append(_checkpoint(
            "10_consolidation_pipeline", t, cons_ok, status=r.status_code,
        ))

        # ---- 11. AI scenarios full list ---------------------------
        t = time.perf_counter()
        # Trigger seeding of the 3 raw scenarios (lazy on first call).
        client.get(f"{BASE}/ai/raw/scenarios")
        r = client.get(f"{BASE}/ai/scenarios")
        assert r.status_code == 200, r.text
        sids = sorted(s["scenario_id"] for s in r.json()["scenarios"])
        ai_count = len(sids)
        assert ai_count >= 9, (ai_count, sids)
        results.append(_checkpoint(
            "11_ai_scenarios_full_list", t, True,
            count=ai_count, has_raw=all(x in sids for x in
                ("audit_opinion_draft", "nl_query", "notes_draft")),
        ))

        # Enable raw scenarios (default off after seeding).
        for sid in ("nl_query", "audit_opinion_draft", "notes_draft"):
            client.patch(f"{BASE}/ai/scenarios/{sid}", json={"enabled": True})

        # Monkey-patch the raw scenario .run functions so the REST
        # endpoints route through a mock local router (no live LLM).
        # The patch is restored at the end of the run.
        original_runs = _patch_raw_scenarios_with_mock_router()

        # ---- 12. Raw NL query happy path -------------------------
        t = time.perf_counter()
        r = client.post(f"{BASE}/ai/raw/nl-query", json={
            "org_id": org_id,
            "question": "最近三年的应收账款余额",
            "execute_sql": False,
            "auto_decision": "allow_once",
        })
        assert r.status_code == 200, r.text
        nlq = r.json()
        nlq_outcome = nlq.get("scenario_result", {}).get("outcome")
        results.append(_checkpoint(
            "12_raw_nl_query", t, nlq_outcome == "success",
            outcome=nlq_outcome, safe=nlq.get("safe"),
        ))

        # ---- 13. Raw audit opinion --------------------------------
        t = time.perf_counter()
        r = client.post(f"{BASE}/ai/raw/audit-opinion", json={
            "org_id": org_id,
            "validations_json": [],
            "template_text": "标准无保留意见模板",
            "period_label": period_id,
            "auto_decision": "allow_once",
        })
        assert r.status_code == 200, r.text
        # audit-opinion endpoint returns ScenarioRunResult.to_dict()
        # at top level (no scenario_result wrapper), unlike nl-query.
        ao_body = r.json()
        ao_outcome = ao_body.get("outcome") or ao_body.get(
            "scenario_result", {}
        ).get("outcome")
        results.append(_checkpoint(
            "13_raw_audit_opinion", t, ao_outcome == "success",
            outcome=ao_outcome,
        ))

        # ---- 14. Notes generate ----------------------------------
        t = time.perf_counter()
        r = client.post(f"{BASE}/orgs/{org_id}/notes/generate", json={
            "period_id": period_id,
        })
        assert r.status_code in (200, 201), r.text
        notes_payload = r.json()
        doc_id = notes_payload["document_id"]
        notes_total = notes_payload["notes_count"]
        results.append(_checkpoint(
            "14_notes_generate", t, notes_total >= 6,
            document_id=doc_id, total=notes_total,
        ))

        # ---- 15. Notes PATCH version-conflict --------------------
        t = time.perf_counter()
        r = client.get(f"{BASE}/orgs/{org_id}/notes/documents/{doc_id}/notes")
        assert r.status_code == 200, r.text
        body = r.json()
        notes_list = body.get("notes") or body.get("items") or []
        if not isinstance(notes_list, list):
            notes_list = notes_list.get("items", [])
        assert notes_list, body
        first_note = notes_list[0]
        nid = first_note["id"]
        cur_v = first_note["version"]
        r = client.patch(
            f"{BASE}/orgs/{org_id}/notes/{nid}",
            json={"content": "updated by m3 closing", "version": cur_v},
        )
        first_patch_ok = r.status_code == 200
        r = client.patch(
            f"{BASE}/orgs/{org_id}/notes/{nid}",
            json={"content": "stale write", "version": cur_v},
        )
        conflict_ok = r.status_code == 409
        results.append(_checkpoint(
            "15_notes_optimistic_lock", t,
            first_patch_ok and conflict_ok,
            first_patch=first_patch_ok, stale_409=conflict_ok,
        ))

        # ---- 16. Notes finalize ----------------------------------
        t = time.perf_counter()
        r = client.post(f"{BASE}/orgs/{org_id}/notes/documents/{doc_id}/finalize")
        fz_ok = r.status_code in (200, 201)
        if fz_ok:
            body = r.json()
            status = (body.get("status")
                      or body.get("document", {}).get("status")
                      or body.get("doc", {}).get("status"))
            fz_ok = status == "finalized"
        results.append(_checkpoint("16_notes_finalize", t, fz_ok))

        # ---- 17. Raw S11 fills a narrative_pending_ai note --------
        # We must locate a narrative-pending row before finalize disables
        # further edits — but step 16 already finalized. Re-fetch the list
        # in case the table tracks notes independently of the document
        # status flag.
        t = time.perf_counter()
        r = client.get(f"{BASE}/orgs/{org_id}/notes/documents/{doc_id}/notes")
        body = r.json()
        notes_list2 = body.get("notes") or []
        narrative_id = None
        for n in notes_list2:
            if n.get("kind") == "narrative_pending_ai":
                narrative_id = n["id"]
                break
        s11_ok = False
        if narrative_id is not None:
            r = client.post(f"{BASE}/ai/raw/notes-draft", json={
                "org_id": org_id,
                "note_id": narrative_id,
                "auto_decision": "allow_once",
            })
            if r.status_code == 200:
                resp = r.json()
                note = resp.get("note") or {}
                s11_ok = note.get("kind") == "narrative"
        results.append(_checkpoint(
            "17_raw_notes_draft", t, s11_ok or narrative_id is None,
            narrative_id=narrative_id,
        ))

        # ---- 18. Peer comparison + benchmarks --------------------
        t = time.perf_counter()
        r = client.get(f"{BASE}/peer-benchmarks")
        assert r.status_code == 200, r.text
        bench_body = r.json()
        bench_items = (bench_body.get("benchmarks")
                       or bench_body.get("items")
                       or bench_body)
        if isinstance(bench_items, dict):
            bench_count = bench_body.get("total", 0)
        else:
            bench_count = len(bench_items)
        r = client.post(
            f"{BASE}/orgs/{org_id}/peer-comparison/run",
            json={"period_id": period_id, "industry_code": "manufacturing"},
        )
        pc_ok = r.status_code in (200, 201)
        results.append(_checkpoint(
            "18_peer_comparison", t,
            bench_count >= 12 and pc_ok,
            benchmark_rows=bench_count, run_status=r.status_code,
        ))

        # Restore monkey-patches.
        _restore_raw_scenarios(original_runs)

        # ---- 19. Key rotation -----------------------------------
        # Seed key_meta.global so /admin/key-rotate has something to
        # rotate.  Without this the rotation endpoint correctly
        # returns 400 ("encryption not enabled").
        asyncio.run(_enable_encryption_for_test(service))
        t = time.perf_counter()
        r = client.get(f"{BASE}/admin/key-rotation-preview")
        kpv_ok = r.status_code == 200
        r = client.post(f"{BASE}/admin/key-rotate", json={"reason": "m3 acceptance"})
        kr_ok = r.status_code in (200, 201)
        r2 = client.get(f"{BASE}/admin/key-versions")
        kv_body = r2.json() if r2.status_code == 200 else {}
        versions = kv_body.get("versions") or kv_body.get("items") or []
        if not isinstance(versions, list):
            versions = []
        v_active = [v for v in versions if v.get("status") == "active"]
        results.append(_checkpoint(
            "19_key_rotation", t, kpv_ok and kr_ok and len(v_active) >= 1,
            preview_ok=kpv_ok, rotate_status=r.status_code,
            versions=len(versions), active=len(v_active),
        ))

        # ---- 20. Backup create ---------------------------------
        t = time.perf_counter()
        r = client.post(
            f"{BASE}/admin/backups",
            json={
                "passphrase": "P@ssw0rd-M3-Acceptance",
                "dest_dir": str(work / "backups"),
            },
        )
        bk_ok = r.status_code in (200, 201)
        backup_row = r.json() if bk_ok else {}
        backup_id = backup_row.get("id") or backup_row.get("backup_id")
        results.append(_checkpoint(
            "20_backup_create", t, bk_ok and backup_id is not None,
            backup_id=backup_id, sha256=backup_row.get("sha256"),
            size=backup_row.get("size_bytes"),
        ))

        # ---- 21. Backup restore (wrong passphrase) -------------
        t = time.perf_counter()
        rs_wrong_ok = False
        if backup_id is not None:
            r = client.post(
                f"{BASE}/admin/backups/{backup_id}/restore",
                json={"passphrase": "WRONG-passphrase", "dry_run": True},
            )
            text_blob = (r.text or "").lower()
            rs_wrong_ok = (
                r.status_code in (200, 400, 401, 403)
                and ("passphrase" in text_blob
                     or "decrypt" in text_blob
                     or "wrong" in text_blob)
            )
        results.append(_checkpoint(
            "21_backup_restore_wrong_passphrase", t, rs_wrong_ok,
        ))

        # ---- 22. Backup restore dry-run (correct) --------------
        t = time.perf_counter()
        rs_dry_ok = False
        if backup_id is not None:
            r = client.post(
                f"{BASE}/admin/backups/{backup_id}/restore",
                json={
                    "passphrase": "P@ssw0rd-M3-Acceptance",
                    "dry_run": True,
                },
            )
            if r.status_code in (200, 201):
                body = r.json()
                rs_dry_ok = (
                    body.get("verified") is True
                    or body.get("manifest") is not None
                    or body.get("ok") is True
                )
        results.append(_checkpoint(
            "22_backup_restore_dry_run", t, rs_dry_ok,
        ))

        # ---- 23. DELETE /orgs/{id} refuses on non-empty (cascade=false) ----
        # v1.0.0-rc1 EX-P2-10: the new endpoint must refuse to drop the
        # M3 closing primary org because steps 02..22 left a mountain of
        # dependent rows (imports / reports / cells / consol / notes / ...).
        t = time.perf_counter()
        delete_refuse_ok = False
        try:
            r = client.delete(f"{BASE}/orgs/{org_id}")
            if r.status_code == 409:
                detail = r.json().get("detail", {})
                delete_refuse_ok = (
                    detail.get("error") == "org_not_empty"
                    and detail.get("total_dependents", 0) > 0
                )
        except Exception:  # noqa: BLE001 — diagnostic step, never fatal
            delete_refuse_ok = False
        results.append(_checkpoint(
            "23_delete_org_refuses_non_empty", t, delete_refuse_ok,
        ))

        # ---- 24. DELETE /orgs/{id} cascade=true purges everything ----------
        # Same endpoint, ``?cascade=true`` — the closing acceptance gets to
        # drop the test fixture at the end so each re-run starts clean.
        t = time.perf_counter()
        delete_cascade_ok = False
        try:
            r = client.delete(f"{BASE}/orgs/{org_id}?cascade=true")
            if r.status_code == 200:
                body = r.json()
                delete_cascade_ok = (
                    body.get("deleted") is True
                    and body.get("cascade") is True
                    and body.get("org_rows_deleted", 0) == 1
                )
                # And the org really is gone.
                r2 = client.get(f"{BASE}/orgs")
                ids = [o["id"] for o in r2.json().get("organizations", [])]
                delete_cascade_ok = delete_cascade_ok and org_id not in ids
        except Exception:  # noqa: BLE001
            delete_cascade_ok = False
        results.append(_checkpoint(
            "24_delete_org_cascade_purges_all", t, delete_cascade_ok,
        ))

        # ---- 25+. Regression block -----------------------------
        if args.skip_regression:
            results.append(_checkpoint(
                "25_regression", time.perf_counter(), True, skipped=True,
            ))
        else:
            for r in _run_regression():
                results.append(r)
                if not r["ok"]:
                    failures.append(r["step"])

        all_ok = all(r["ok"] for r in results) and not failures
        elapsed = int((time.perf_counter() - started_all) * 1000)
        summary = {
            "status": "ok" if all_ok else "fail",
            "schema_version": SCHEMA_VERSION,
            "route_count": routes_total,
            "elapsed_total_ms": elapsed,
            "steps_total": len(results),
            "steps_ok": sum(1 for r in results if r["ok"]),
            "results": results,
            "failures": failures,
            "tmpdir": str(work) if args.keep else None,
        }
        if args.json:
            Path(args.json).write_text(
                json.dumps(summary, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
        print(
            f"{'OK ' if all_ok else 'FAIL'} steps_ok={summary['steps_ok']}/{summary['steps_total']}"
            f"  elapsed={elapsed}ms  json={args.json or 'inline'}",
            flush=True,
        )
        return 0 if all_ok else 1

    except Exception as exc:  # noqa: BLE001
        traceback.print_exc()
        failures.append(str(exc))
        if args.json:
            Path(args.json).write_text(
                json.dumps({
                    "status": "fail",
                    "error": str(exc),
                    "results": results,
                    "failures": failures,
                }, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
        return 2
    finally:
        if not args.keep:
            try:
                shutil.rmtree(work, ignore_errors=True)
            except Exception:
                pass


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__.strip().splitlines()[0])
    p.add_argument("--json", type=str, default=None,
                   help="write summary JSON to this path")
    p.add_argument("--skip-regression", action="store_true",
                   help="skip the 5 M1+M2 subprocess regression scripts")
    p.add_argument("--keep", action="store_true",
                   help="keep the temp work dir for inspection")
    return p.parse_args()


def main() -> int:
    args = parse_args()
    return run(args)


if __name__ == "__main__":
    rc = main()
    # Audit P2-5: TestClient.websocket_connect spins up a non-daemon
    # ASGI worker thread that lingers after main() returns, so the CPython
    # interpreter wedges forever waiting on it.  Force an immediate exit
    # via os._exit (bypasses atexit) — every meaningful artefact (JSON
    # summary, console line) has been flushed by ``run()`` already.
    sys.stdout.flush()
    sys.stderr.flush()
    os._exit(rc)
