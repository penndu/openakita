"""End-to-end acceptance script for M1 W2 (W1 + W2 combined).

Runs the full pipeline against an in-process FastAPI app + a fresh SQLite
file with KeyManager encryption forced on.  Steps:

1.  Create an account book (W1).
2.  Upload a balance table (W1; uses the real A_balance.xlsx sample if it
    is present in tmp_finance_analysis/, otherwise builds a minimal
    synthetic one so the script is hermetic).
3.  Generate a balance sheet (W2 Stage 4).
4.  Export the report to .xlsx (W2 Stage 4) and re-open it via openpyxl
    to verify a known cell.
5.  Upload a Golden-Tax-IV VAT declaration (W2 Stage 5).
6.  Upload an audit template + render it with the report data (W2
    Stage 6).
7.  Inspect the raw SQLite file to confirm:
       - key_meta.enabled = 1
       - organizations.name is empty / NULL on disk
       - trial_balance_rows[*]._encrypted_payload is non-empty
       - reports / vat_declarations / audit_templates rows present.

Returns exit code 0 iff every step succeeded; otherwise prints the failure
context and returns 1.

Usage::

    d:\\OpenAkita\\.venv\\Scripts\\python.exe ^
        plugins/finance-auto/scripts/m1_w2_acceptance.py ^
        [--keep] [--db <path>]
"""

from __future__ import annotations

import argparse
import asyncio
import json
import os
import shutil
import sqlite3
import sys
import tempfile
import traceback
from pathlib import Path

import openpyxl
from fastapi import FastAPI
from fastapi.testclient import TestClient

PLUGIN_ROOT = Path(__file__).resolve().parent.parent
if str(PLUGIN_ROOT) not in sys.path:
    sys.path.insert(0, str(PLUGIN_ROOT))

from finance_auto_backend.key_manager import (  # noqa: E402
    ENV_PASSPHRASE,
    KeyManager,
)
from finance_auto_backend.key_meta import (  # noqa: E402
    GLOBAL_COMPONENT,
    write_key_meta,
)
from finance_auto_backend.routes import build_router_and_service  # noqa: E402

REAL_BALANCE = (
    Path(__file__).resolve().parents[2]
    / "tmp_finance_analysis"
    / "xlsx"
    / "A_balance.xlsx"
)


def _build_synthetic_balance(path: Path) -> None:
    """Minimal hermetic fallback when the real sample isn't present.

    Builds an .xls-shaped layout so the W1 three-tier parser detects a
    valid balance table.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "余额表"
    ws["A1"] = "测试公司 - 2025-FY 余额表"
    headers = [
        "科目编码", "科目名称",
        "期初借方", "期初贷方",
        "本期借方", "本期贷方",
        "期末借方", "期末贷方",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=2, column=i, value=h)
    rows = [
        ("1001", "库存现金", 0, 0, 0, 0, 5000, 0),
        ("1002", "银行存款", 0, 0, 0, 0, 200000, 0),
        ("1122", "应收账款", 0, 0, 0, 0, 50000, 0),
        ("1601", "固定资产", 0, 0, 0, 0, 1000000, 0),
        ("1602", "累计折旧", 0, 0, 0, 0, 0, 200000),
        ("2202", "应付账款", 0, 0, 0, 0, 0, 300000),
        ("4001", "实收资本", 0, 0, 0, 0, 0, 700000),
        ("4104", "未分配利润", 0, 0, 0, 0, 0, 55000),
    ]
    for offset, r in enumerate(rows, start=3):
        for col, v in enumerate(r, start=1):
            ws.cell(row=offset, column=col, value=v)
    wb.save(str(path))
    wb.close()


def _build_vat(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "主表"
    ws.merge_cells("A1:E1")
    ws["A1"] = "增值税及附加税费申报表（一般纳税人适用）"
    ws["A2"] = "项目"
    ws["B2"] = "本月数"
    ws["C2"] = "本年累计"
    ws["A3"] = "本期销项税额"
    ws["B3"] = 130000.00
    ws["C3"] = 1500000.00
    ws["A4"] = "本期进项税额"
    ws["B4"] = 80000.00
    ws["C4"] = 900000.00
    ws["A5"] = "上期留抵税额"
    ws["B5"] = 0
    ws["A6"] = "应纳税额"
    ws["B6"] = 50000.00
    ws["C6"] = 600000.00
    ws["A7"] = "附加税费合计"
    ws["B7"] = 5000.00
    wb.save(str(path))
    wb.close()


def _build_audit_template(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "底稿"
    ws["A1"] = "{{ org.name }} - {{ year }} 资产负债表抽样底稿"
    ws["A2"] = "项目"
    ws["B2"] = "金额"
    ws["A3"] = "货币资金"
    ws["B3"] = "{{ cells.BS_1001.value }}"
    ws["A4"] = "应收账款"
    ws["B4"] = "{{ BS_1122 }}"
    ws["A5"] = "审计员"
    ws["B5"] = "{{ auditor }}"
    wb.save(str(path))
    wb.close()


# ---------------------------------------------------------------------------
# Step runners
# ---------------------------------------------------------------------------


async def _enable_encryption(db) -> None:
    """Force-enable encryption on the fresh DB before the service first
    writes anything."""
    km = KeyManager()
    salt = os.urandom(32)
    seed = os.urandom(32)
    km.unlock(seed, salt)
    await write_key_meta(
        db.conn,
        component=GLOBAL_COMPONENT,
        salt=salt,
        kdf_iterations=200_000,
        enabled=True,
        seed_source="env",
    )
    os.environ[ENV_PASSPHRASE] = seed.hex()


def _print_step(idx: int, label: str) -> None:
    print(f"\n=== Step {idx}: {label} ===")


def run() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--db", type=str, default=None)
    parser.add_argument("--keep", action="store_true",
                        help="leave the temp directory in place after the run")
    args = parser.parse_args()

    work = Path(tempfile.mkdtemp(prefix="finauto_w2_acc_"))
    print(f"workspace: {work}")
    db_path = Path(args.db) if args.db else work / "acceptance.sqlite"
    if db_path.exists():
        db_path.unlink()

    router, service, db = build_router_and_service(db_path)
    app = FastAPI(title="finance-auto W2 acceptance")
    app.include_router(router, prefix="/api/plugins/finance-auto")
    base = "/api/plugins/finance-auto"

    async def boot() -> None:
        await db.init()
        await _enable_encryption(db)
        # service must rebuild its KeyManager from the fresh meta + env seed.
        await service.auto_unlock_if_configured()

    asyncio.run(boot())

    summary: dict[str, object] = {"steps": {}}
    fail = False
    client = TestClient(app)
    try:
        # ------------------------------------------------------------------
        _print_step(1, "Create org")
        r = client.post(
            f"{base}/orgs",
            json={
                "name": "测试公司 (W2 验收)", "code": "W2_ACCEPT_DEMO",
                "industry": "general", "standard": "small",
            },
        )
        assert r.status_code == 201, r.text
        org_id = r.json()["id"]
        summary["steps"]["1_create_org"] = {"http": r.status_code, "id": org_id}
        print(f"  org_id={org_id}")

        # ------------------------------------------------------------------
        _print_step(2, "Upload balance table")
        if REAL_BALANCE.exists():
            balance_path = REAL_BALANCE
            print(f"  using real sample: {balance_path}")
        else:
            balance_path = work / "balance.xlsx"
            _build_synthetic_balance(balance_path)
            print(f"  built synthetic sample: {balance_path}")

        with balance_path.open("rb") as fh:
            r = client.post(
                f"{base}/orgs/{org_id}/imports",
                files={
                    "file": (balance_path.name, fh,
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                },
                data={"period_id": "2025-FY"},
            )
        assert r.status_code == 201, r.text
        upload = r.json()
        summary["steps"]["2_upload"] = {
            "http": r.status_code,
            "row_count": upload["row_count"],
            "import_id": upload["import_id"],
        }
        print(f"  http={r.status_code} rows={upload['row_count']}")

        # ------------------------------------------------------------------
        _print_step(3, "Generate balance sheet (W2)")
        r = client.post(
            f"{base}/orgs/{org_id}/reports/balance_sheet/generate",
            json={"period_id": "2025-FY"},
        )
        assert r.status_code == 201, r.text
        body = r.json()
        report_id = body["report"]["id"]
        cell_count = body["report"]["cell_count"]
        summary["steps"]["3_generate_bs"] = {
            "http": r.status_code, "report_id": report_id,
            "cell_count": cell_count,
        }
        print(f"  report_id={report_id} cells={cell_count}")
        bs_1001 = next(
            (c for c in body["cells"] if c["reference_code"] == "BS_1001"),
            None,
        )
        assert bs_1001 is not None
        print(f"  BS_1001 (货币资金) = {bs_1001['value']}")

        # ------------------------------------------------------------------
        _print_step(4, "Export report to .xlsx + re-open")
        r = client.get(f"{base}/orgs/{org_id}/reports/{report_id}/export")
        assert r.status_code == 200, r.text
        out = work / "exported.xlsx"
        out.write_bytes(r.content)
        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        print(f"  sheet={ws.title} rows={ws.max_row} bytes={out.stat().st_size}")
        first_data_row = next(
            (r_ for r_ in ws.iter_rows(min_row=3, max_col=3, values_only=True)
             if r_[2] is not None),
            None,
        )
        assert first_data_row is not None
        wb.close()
        summary["steps"]["4_export"] = {
            "http": r.status_code, "bytes": out.stat().st_size,
        }
        print(f"  first body row: {first_data_row}")

        # ------------------------------------------------------------------
        _print_step(5, "Upload VAT declaration")
        vat_path = work / "vat.xlsx"
        _build_vat(vat_path)
        with vat_path.open("rb") as fh:
            r = client.post(
                f"{base}/orgs/{org_id}/vat-declarations",
                files={"file": ("vat.xlsx", fh,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                data={"declaration_period": "2025-01"},
            )
        assert r.status_code == 201, r.text
        vat = r.json()
        summary["steps"]["5_vat"] = {
            "http": r.status_code,
            "output_vat": vat["output_vat"], "input_vat": vat["input_vat"],
            "tax_payable": vat["tax_payable"],
        }
        print(f"  output={vat['output_vat']} input={vat['input_vat']} "
              f"payable={vat['tax_payable']}")

        # ------------------------------------------------------------------
        _print_step(6, "Upload audit template + render")
        tpl_path = work / "audit_tpl.xlsx"
        _build_audit_template(tpl_path)
        with tpl_path.open("rb") as fh:
            r = client.post(
                f"{base}/audit-templates",
                files={"file": ("audit_tpl.xlsx", fh,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                data={"name": "BS 抽样底稿"},
            )
        assert r.status_code == 201, r.text
        tpl = r.json()
        tpl_id = tpl["id"]
        summary["steps"]["6_audit_upload"] = {
            "http": r.status_code, "tpl_id": tpl_id,
            "placeholder_count": tpl["placeholder_count"],
            "unknown": tpl["unknown_placeholder_count"],
        }
        print(f"  tpl_id={tpl_id} placeholders={tpl['placeholder_count']} "
              f"unknown={tpl['unknown_placeholder_count']}")
        r = client.post(
            f"{base}/orgs/{org_id}/audit-templates/{tpl_id}/render",
            json={"report_id": report_id, "strict": True},
        )
        assert r.status_code == 200, r.text
        rendered = work / "rendered.xlsx"
        rendered.write_bytes(r.content)
        wb = openpyxl.load_workbook(str(rendered))
        ws = wb.active
        a1 = ws["A1"].value
        b3 = ws["B3"].value
        wb.close()
        summary["steps"]["6_audit_render"] = {
            "http": r.status_code, "bytes": rendered.stat().st_size,
            "rendered_A1": a1, "rendered_B3": b3,
        }
        print(f"  rendered A1={a1!r} B3={b3}")

        # ------------------------------------------------------------------
        _print_step(7, "Inspect raw SQLite for ciphertext + W2 rows")
        # Close service via shutdown (async) then sqlite3 the file directly.
        async def _shutdown() -> None:
            await db.close()
        asyncio.run(_shutdown())

        snap = sqlite3.connect(str(db_path))
        snap.row_factory = sqlite3.Row
        try:
            cur = snap.execute("SELECT enabled, seed_source FROM key_meta")
            km_row = cur.fetchone()
            assert km_row and km_row["enabled"] == 1
            cur = snap.execute(
                "SELECT name, _encrypted_payload FROM organizations WHERE id=?",
                (org_id,),
            )
            org_row = cur.fetchone()
            assert org_row is not None
            assert (org_row["name"] or "") == ""
            assert org_row["_encrypted_payload"] is not None
            cur = snap.execute(
                "SELECT COUNT(*) AS n, "
                "SUM(CASE WHEN _encrypted_payload IS NOT NULL "
                "AND length(_encrypted_payload) > 0 THEN 1 ELSE 0 END) AS enc "
                "FROM trial_balance_rows WHERE org_id=?",
                (org_id,),
            )
            row_counts = cur.fetchone()
            cur = snap.execute("SELECT COUNT(*) AS n FROM reports")
            reports_n = cur.fetchone()["n"]
            cur = snap.execute("SELECT COUNT(*) AS n FROM vat_declarations")
            vat_n = cur.fetchone()["n"]
            cur = snap.execute("SELECT COUNT(*) AS n FROM audit_templates")
            audit_n = cur.fetchone()["n"]
        finally:
            snap.close()

        encrypted_ratio = (
            row_counts["enc"] / row_counts["n"] if row_counts["n"] else 0.0
        )
        summary["steps"]["7_db_inspect"] = {
            "key_meta_enabled": bool(km_row["enabled"]),
            "seed_source": km_row["seed_source"],
            "org_name_cleartext": org_row["name"],
            "tb_rows": row_counts["n"],
            "tb_rows_encrypted": row_counts["enc"],
            "encrypted_ratio": round(encrypted_ratio, 4),
            "reports": reports_n,
            "vat_declarations": vat_n,
            "audit_templates": audit_n,
        }
        print(f"  key_meta.enabled={km_row['enabled']} seed={km_row['seed_source']}")
        print(f"  org.name on disk = {org_row['name']!r}")
        print(f"  trial_balance_rows: total={row_counts['n']} "
              f"encrypted={row_counts['enc']} ratio={encrypted_ratio:.2%}")
        print(f"  reports={reports_n} vat={vat_n} audit_templates={audit_n}")

        assert row_counts["n"] > 0
        assert encrypted_ratio == 1.0
        assert reports_n >= 1 and vat_n >= 1 and audit_n >= 1

    except AssertionError as exc:
        traceback.print_exc()
        summary["failure"] = f"AssertionError: {exc}"
        fail = True
    except Exception as exc:
        traceback.print_exc()
        summary["failure"] = f"{type(exc).__name__}: {exc}"
        fail = True
    finally:
        # Stash a JSON summary next to the workspace for the W2 completion
        # report.  Even on failure this is worth keeping.
        summary_path = Path("_m1_w2_acceptance_result.json").resolve()
        summary_path.write_text(
            json.dumps(summary, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        print(f"\nSummary written to {summary_path}")
        if not args.keep:
            try:
                shutil.rmtree(work, ignore_errors=True)
            except Exception:
                pass

    return 1 if fail else 0


if __name__ == "__main__":
    raise SystemExit(run())
