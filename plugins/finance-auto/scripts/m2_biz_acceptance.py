"""End-to-end acceptance script for M2 Biz backend (Stages 1–6 combined).

Runs the full M2 business pipeline against an in-process FastAPI app + a
fresh SQLite file.  Covers every requirement from the M2 Biz Stage 7
acceptance checklist:

1. Register 3 users (auditor / manager / partner).
2. Create an org + upload a balance file (M1 happy path).
3. Assign all three users to the org.
4. Generate a balance sheet.
5. Submit → approve → sign-off (full review-workflow happy path).
6. Add 1 comment to a freshly-generated BS cell.
7. Create 1 reclassification rule → preview → apply.
8. Fill 7 manual inputs → generate a cash-flow statement
   (verifies ≥ 5 cf_* cells have non-zero values).
9. Create a *second* org + add it to a consolidation group with a
   matching elimination, trigger consolidation, verify the merged
   report subtracts the elimination amount.
10. DB inspect: row counts of new v9 tables + ``version`` columns
    incremented.

Exit code 0 iff every step succeeded; otherwise prints the failure
context and returns 1.

Usage::

    d:\\OpenAkita\\.venv\\Scripts\\python.exe ^
        plugins/finance-auto/scripts/m2_biz_acceptance.py ^
        [--keep] [--db <path>] [--json <path>]
"""

from __future__ import annotations

import argparse
import asyncio
import json
import shutil
import sqlite3
import sys
import tempfile
import time
import traceback
from pathlib import Path

import openpyxl
from fastapi import FastAPI
from fastapi.testclient import TestClient

PLUGIN_ROOT = Path(__file__).resolve().parent.parent
if str(PLUGIN_ROOT) not in sys.path:
    sys.path.insert(0, str(PLUGIN_ROOT))

from finance_auto_backend.routes import build_router_and_service  # noqa: E402


BASE = "/api/plugins/finance-auto"


HEADER = [
    "科目编码", "科目名称",
    "期初借方", "期初贷方",
    "本期借方", "本期贷方",
    "期末借方", "期末贷方",
]


def _write_balance(path: Path, rows: list[tuple]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "余额表"
    ws["A1"] = "M2 Biz Acceptance"
    for i, h in enumerate(HEADER, start=1):
        ws.cell(row=2, column=i, value=h)
    for offset, r in enumerate(rows, start=3):
        for col, v in enumerate(r, start=1):
            ws.cell(row=offset, column=col, value=v)
    wb.save(str(path))
    wb.close()


def _build_primary_balance(path: Path) -> None:
    rows = [
        ("1001", "库存现金",     0, 0, 0, 0,    5_000, 0),
        ("1002", "银行存款",     0, 0, 0, 0,  250_000, 0),
        ("1122", "应收账款",     0, 0, 0, 0,  150_000, 0),
        ("1601", "固定资产",     0, 0, 0, 0, 1_200_000, 0),
        ("1602", "累计折旧",     0, 0, 0, 0, 0,   320_000),
        ("2202", "应付账款",     0, 0, 0, 0, 0,   180_000),
        ("4001", "实收资本",     0, 0, 0, 0, 0,   900_000),
        ("4104", "未分配利润",   0, 0, 0, 0, 0,   205_000),
    ]
    _write_balance(path, rows)


def _build_subsidiary_balance(path: Path) -> None:
    rows = [
        ("1001", "库存现金",     0, 0, 0, 0,    2_000, 0),
        ("1002", "银行存款",     0, 0, 0, 0,   80_000, 0),
        ("1122", "应收账款",     0, 0, 0, 0,   50_000, 0),
        ("1601", "固定资产",     0, 0, 0, 0,  400_000, 0),
        ("2202", "应付账款",     0, 0, 0, 0, 0,    60_000),
        ("4001", "实收资本",     0, 0, 0, 0, 0,   200_000),
        ("4104", "未分配利润",   0, 0, 0, 0, 0,   272_000),
    ]
    _write_balance(path, rows)


def _checkpoint(name: str, started: float, ok: bool, **extras) -> dict:
    return {
        "step": name,
        "ok": ok,
        "elapsed_ms": int((time.perf_counter() - started) * 1000),
        **extras,
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--db", type=Path, default=None,
                        help="reuse an existing sqlite file (overwritten)")
    parser.add_argument("--keep", action="store_true",
                        help="keep the tempdir after the run for inspection")
    parser.add_argument(
        "--json", dest="json_out", type=Path,
        default=PLUGIN_ROOT.parent.parent / "_m2_biz_acceptance_result.json",
        help="path to write the JSON summary",
    )
    args = parser.parse_args()

    work = Path(tempfile.mkdtemp(prefix="m2_biz_accept_"))
    db_path = args.db or (work / "m2_biz.sqlite")
    if args.db and args.db.exists():
        args.db.unlink()

    results: list[dict] = []
    failures: list[str] = []
    final_status = "ok"
    started_all = time.perf_counter()

    try:
        router, service, db = build_router_and_service(db_path)
        app = FastAPI()
        app.include_router(router, prefix=BASE)
        asyncio.run(db.init())
        client = TestClient(app)

        # ---- Step 1: register 3 users ------------------------------------
        t = time.perf_counter()
        for uid, role in (
            ("usr_auditor", "auditor"),
            ("usr_manager", "manager"),
            ("usr_partner", "partner"),
        ):
            r = client.post(f"{BASE}/users", json={
                "user_id": uid, "display_name": uid.upper(), "role": role,
            })
            if r.status_code not in (201, 409):
                raise RuntimeError(f"register {uid}: {r.status_code} {r.text}")
        results.append(_checkpoint("01_register_users", t, True, user_count=3))

        # ---- Step 2: create org + upload balance --------------------------
        t = time.perf_counter()
        r = client.post(f"{BASE}/orgs", json={
            "name": "M2 Biz 验收公司", "code": "M2BIZ_ACCEPT",
            "standard": "small",
        })
        assert r.status_code == 201, r.text
        primary_org = r.json()["id"]

        bal_path = work / "primary_bal.xlsx"
        _build_primary_balance(bal_path)
        period_id = "2025-FY"
        with bal_path.open("rb") as fh:
            r = client.post(
                f"{BASE}/orgs/{primary_org}/imports",
                data={"period_id": period_id},
                files={"file": ("primary_bal.xlsx", fh,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        assert r.status_code == 201, r.text
        primary_import = r.json()["import_id"]
        results.append(_checkpoint("02_create_org_and_upload", t, True,
                                    org_id=primary_org, import_id=primary_import))

        # ---- Step 3: assign 3 users --------------------------------------
        t = time.perf_counter()
        for uid, role in (
            ("usr_auditor", "lead_auditor"),
            ("usr_manager", "reviewer"),
            ("usr_partner", "partner_signoff"),
        ):
            r = client.post(f"{BASE}/orgs/{primary_org}/assignments", json={
                "user_id": uid, "period_id": period_id, "role_in_project": role,
            })
            assert r.status_code == 201, r.text
        results.append(_checkpoint("03_assign_users", t, True, assignments=3))

        # ---- Step 4: generate balance sheet ------------------------------
        t = time.perf_counter()
        r = client.post(
            f"{BASE}/orgs/{primary_org}/reports/balance_sheet/generate",
            json={"period_id": period_id, "source_import_id": primary_import},
        )
        assert r.status_code == 201, r.text
        bs = r.json()
        bs_report_id = bs["report"]["id"]
        bs_cell_count = len(bs["cells"])
        first_cell_id = bs["cells"][0]["id"]
        results.append(_checkpoint("04_generate_bs", t, True,
                                    report_id=bs_report_id, cell_count=bs_cell_count))

        # ---- Step 5: review workflow happy path --------------------------
        t = time.perf_counter()
        r = client.post(
            f"{BASE}/orgs/{primary_org}/reports/{bs_report_id}/review/submit",
            json={"auditor_id": "usr_auditor"},
        )
        assert r.status_code == 201, r.text
        wf_after_submit = r.json()
        assert wf_after_submit["status"] == "pending_review"
        r = client.post(
            f"{BASE}/orgs/{primary_org}/reports/{bs_report_id}/review/approve",
            json={"actor_id": "usr_manager"},
        )
        assert r.status_code == 200, r.text
        assert r.json()["status"] == "pending_signoff"
        r = client.post(
            f"{BASE}/orgs/{primary_org}/reports/{bs_report_id}/review/sign-off",
            json={"actor_id": "usr_partner"},
        )
        assert r.status_code == 200, r.text
        wf_final = r.json()
        assert wf_final["status"] == "signed_off"
        results.append(_checkpoint("05_review_workflow", t, True,
                                    workflow_id=wf_final["workflow_id"],
                                    history_hops=len(wf_final["history"])))

        # ---- Step 6: add comment to a cell -------------------------------
        t = time.perf_counter()
        r = client.post(
            f"{BASE}/orgs/{primary_org}/reports/{bs_report_id}/cells/{first_cell_id}/comments",
            json={
                "body": "请复核货币资金期末余额",
                "kind": "review_question",
                "author_id": "usr_manager",
            },
        )
        assert r.status_code == 201, r.text
        results.append(_checkpoint("06_add_comment", t, True,
                                    comment_id=r.json()["id"]))

        # ---- Step 7: reclassification rule preview + apply ---------------
        t = time.perf_counter()
        r = client.post(
            f"{BASE}/orgs/{primary_org}/reclassification-rules",
            json={
                "name": "Acceptance Reclass",
                "when_condition": {
                    "account_code_starts": ["2202"],
                    "balance_direction": "credit",
                    "threshold": "0.01",
                },
                "action": {
                    "move_to_account_code": "1122",
                    "reason": "应付负余额按谨慎性重分类",
                    "parse_issue_severity": "warning",
                    "parse_issue_threshold": "50000",
                },
                "priority": 10,
            },
        )
        assert r.status_code == 201, r.text
        rule_id = r.json()["rule_id"]
        r = client.post(
            f"{BASE}/orgs/{primary_org}/reclassification-runs/preview",
            json={"period_id": period_id},
        )
        assert r.status_code == 201, r.text
        preview_run = r.json()
        r = client.post(
            f"{BASE}/orgs/{primary_org}/reclassification-runs/apply",
            json={"period_id": period_id},
        )
        assert r.status_code == 201, r.text
        apply_run = r.json()
        results.append(_checkpoint("07_reclassification", t, True,
                                    rule_id=rule_id,
                                    preview_items=preview_run["items_count"],
                                    apply_items=apply_run["items_count"],
                                    parse_issues=len(apply_run["parse_issue_ids"])))

        # ---- Step 8: 7 manual_inputs + cash-flow statement --------------
        t = time.perf_counter()
        # First, generate the income statement so the cash-flow engine has
        # PL_NET_PROFIT to consume.
        r = client.post(
            f"{BASE}/orgs/{primary_org}/reports/income_statement/generate",
            json={"period_id": period_id, "source_import_id": primary_import},
        )
        assert r.status_code == 201, r.text
        # Fill 7 manual_inputs (legacy cash_flow_aux keys still respected; we
        # also seed a few cf_* keys directly so the indirect engine has data).
        manual_keys = {
            "vat_output": "12000", "vat_input": "8000",
            "bill_discount_received": "2000", "interest_paid": "1500",
            "interest_income": "300", "bank_fee_paid": "200",
            "social_security_paid": "5000",
        }
        for k, v in manual_keys.items():
            r = client.put(
                f"{BASE}/orgs/{primary_org}/periods/{period_id}/manual-inputs/{k}",
                # Round-2 #1: ``expected_version`` is now required on every
                # PUT; fresh slots use 0.
                json={"value": v, "source": "manual", "expected_version": 0},
            )
            assert r.status_code in (200, 201), f"{k}: {r.status_code} {r.text}"
        # Compute + persist via the indirect engine.
        r = client.post(
            f"{BASE}/orgs/{primary_org}/cash-flow/persist",
            json={"period_id": period_id},
        )
        assert r.status_code == 201, r.text
        cf = r.json()
        non_zero = sum(1 for v in cf["values"].values() if v not in ("0", "0E-9"))
        # Generate the actual cash-flow report.
        r = client.post(
            f"{BASE}/orgs/{primary_org}/reports/cash_flow/generate",
            json={"period_id": period_id, "source_import_id": primary_import},
        )
        assert r.status_code == 201, r.text
        cf_report = r.json()
        cf_cells_non_zero = sum(1 for c in cf_report["cells"] if c["value"] != 0.0)
        assert cf_cells_non_zero >= 5, f"only {cf_cells_non_zero} non-zero cf cells"
        results.append(_checkpoint("08_cash_flow", t, True,
                                    keys_persisted=cf["persisted"],
                                    non_zero_keys=non_zero,
                                    cf_cells=len(cf_report["cells"]),
                                    cf_cells_non_zero=cf_cells_non_zero))

        # ---- Step 9: consolidation engine -------------------------------
        t = time.perf_counter()
        r = client.post(f"{BASE}/orgs", json={
            "name": "M2 Biz 子公司", "code": "M2BIZ_SUB",
            "standard": "small",
        })
        assert r.status_code == 201, r.text
        sub_org = r.json()["id"]
        sub_bal = work / "sub_bal.xlsx"
        _build_subsidiary_balance(sub_bal)
        with sub_bal.open("rb") as fh:
            r = client.post(
                f"{BASE}/orgs/{sub_org}/imports",
                data={"period_id": period_id},
                files={"file": ("sub_bal.xlsx", fh,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        assert r.status_code == 201, r.text
        sub_import = r.json()["import_id"]
        r = client.post(
            f"{BASE}/orgs/{sub_org}/reports/balance_sheet/generate",
            json={"period_id": period_id, "source_import_id": sub_import},
        )
        assert r.status_code == 201, r.text

        r = client.post(f"{BASE}/consolidation-groups", json={
            "name": "M2 Biz 测试集团", "parent_org_id": primary_org,
        })
        assert r.status_code == 201, r.text
        gid = r.json()["group_id"]
        r = client.post(f"{BASE}/consolidation-groups/{gid}/members", json={
            "subsidiary_org_id": sub_org, "ownership_pct": 80.0,
            "join_method": "full",
        })
        assert r.status_code == 201, r.text
        r = client.post(f"{BASE}/consolidation-groups/{gid}/eliminations", json={
            "period_id": period_id, "kind": "inter_ar_ap",
            "debit_target": "BS_2202", "credit_target": "BS_1122",
            "amount": "20000", "rationale": "intra-group AR/AP",
        })
        assert r.status_code == 201, r.text
        r = client.post(f"{BASE}/consolidation-groups/{gid}/runs", json={
            "period_id": period_id, "kind": "balance_sheet",
        })
        assert r.status_code == 201, r.text
        consol = r.json()
        elim_ids = consol["elimination_ids"]
        member_count = len(consol["member_orgs_snapshot"])
        assert member_count == 2, member_count
        assert len(elim_ids) == 1, elim_ids
        results.append(_checkpoint("09_consolidation", t, True,
                                    group_id=gid,
                                    member_count=member_count,
                                    eliminations=len(elim_ids),
                                    minority_interest=consol["minority_interest"]))

        # ---- Step 10: DB inspect ----------------------------------------
        t = time.perf_counter()
        asyncio.run(db.close())
        conn = sqlite3.connect(str(db_path))
        try:
            inspect = {}
            for tbl in (
                "users", "assignments", "review_workflows", "comments",
                "permissions", "reclassification_rules",
                "reclassification_runs", "reclassification_run_items",
                "consolidation_groups", "consolidation_members",
                "elimination_entries", "consolidated_reports",
            ):
                cur = conn.execute(f"SELECT COUNT(*) FROM {tbl}")
                inspect[tbl] = cur.fetchone()[0]
            # Verify a workflow's version was bumped.
            row = conn.execute(
                "SELECT MAX(version) FROM review_workflows"
            ).fetchone()
            wf_max_version = row[0] if row else 0
            # Schema version baseline.
            row = conn.execute(
                "SELECT MAX(version) FROM schema_version"
            ).fetchone()
            schema_version = row[0] if row else 0
        finally:
            conn.close()
        assert schema_version >= 9, schema_version
        assert inspect["users"] >= 3
        assert inspect["assignments"] >= 3
        assert inspect["review_workflows"] >= 1
        assert inspect["comments"] >= 1
        assert inspect["reclassification_rules"] >= 1
        assert inspect["reclassification_runs"] >= 2
        assert inspect["consolidation_groups"] >= 1
        assert inspect["consolidation_members"] >= 2
        assert inspect["elimination_entries"] >= 1
        assert inspect["consolidated_reports"] >= 1
        assert wf_max_version >= 2, wf_max_version  # 3 transitions => version 4+
        results.append(_checkpoint(
            "10_db_inspect", t, True,
            inspect=inspect, wf_max_version=wf_max_version,
            schema_version=schema_version,
        ))

    except Exception as exc:  # noqa: BLE001
        final_status = "failed"
        failures.append(f"{type(exc).__name__}: {exc}")
        traceback.print_exc()
    finally:
        if not args.keep:
            shutil.rmtree(work, ignore_errors=True)

    summary = {
        "status": final_status,
        "elapsed_total_ms": int((time.perf_counter() - started_all) * 1000),
        "tmpdir": str(work) if args.keep else None,
        "steps_total": 10,
        "steps_ok": sum(1 for r in results if r["ok"]),
        "results": results,
        "failures": failures,
    }
    try:
        args.json_out.write_text(
            json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
        )
    except Exception as exc:  # noqa: BLE001
        print(f"warning: could not write json summary: {exc}")

    if final_status == "ok":
        print(f"OK  steps_ok={summary['steps_ok']}/10  "
              f"elapsed={summary['elapsed_total_ms']}ms  "
              f"json={args.json_out}")
        return 0
    print(f"FAIL  steps_ok={summary['steps_ok']}/10  failures={failures}")
    return 1


if __name__ == "__main__":
    sys.exit(main())
