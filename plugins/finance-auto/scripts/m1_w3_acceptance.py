"""End-to-end acceptance script for M1 W3 (W1 + W2 + W3 combined).

Runs the full pipeline against an in-process FastAPI app + a fresh SQLite
file with KeyManager encryption forced on, and exercises every W3
backend stage.

Steps (11):

1.  Create a ``restaurant`` org (Stage 5).
2.  List industries.  Confirms the three shipped overlays
    (manufacturing / restaurant / tech_service) + synthetic ``general``.
3.  GET effective-config for the org — restaurant overlay sets
    ``aux_mode=light``.
4.  Upload a balance table with 15 AP sub-accounts (2202.01–2202.15),
    one unknown-class row (``9001``) and one debit-credit-imbalanced
    row, so Stage 1 has something to find.
5.  List parse-issues.  Pick one and POST decide → learn.
6.  Generate balance sheet (W2 + Stage 2 instrumentation).
7.  PATCH BS_2202 cell with simplify enabled, top_n=10
    (Stage 2).
8.  GET cell details — visible_rows should be 11 (10 kept + 1 "其他"),
    full_rows should be 15.
9.  Upload a prior-period balance (2024-FY) so the cross-period
    validator (Stage 3) has two snapshots to compare.
10. POST cross-period-checks — Stage 3.  Asserts non-zero diff count
    and that a ParseIssue with ``issue_type=cross_period_mismatch``
    was emitted for each error-graded diff.
11. PUT 7 manual_inputs (Stage 4) and generate a cash-flow statement.
    Asserts every slot now has ``filled=true`` and the cash_flow
    report has at least 7 cells whose source row references
    ``manual_input:<key>``.

DB inspect at the end: ``schema_version`` table reports 7,
``parse_issues`` count > 0, ``cross_period_check_results`` count >= 1,
``manual_inputs`` count == 7.

Returns exit code 0 iff every step succeeded; otherwise prints the
failure context and returns 1.

Usage::

    d:\\OpenAkita\\.venv\\Scripts\\python.exe ^
        plugins/finance-auto/scripts/m1_w3_acceptance.py ^
        [--keep] [--db <path>]
"""

from __future__ import annotations

import argparse
import asyncio
import json
import os
import shutil
import sqlite3
import sys
import tempfile
import traceback
from pathlib import Path

import openpyxl
from fastapi import FastAPI
from fastapi.testclient import TestClient

PLUGIN_ROOT = Path(__file__).resolve().parent.parent
if str(PLUGIN_ROOT) not in sys.path:
    sys.path.insert(0, str(PLUGIN_ROOT))

from finance_auto_backend.key_manager import (  # noqa: E402
    ENV_PASSPHRASE,
    KeyManager,
)
from finance_auto_backend.key_meta import (  # noqa: E402
    GLOBAL_COMPONENT,
    write_key_meta,
)
from finance_auto_backend.routes import build_router_and_service  # noqa: E402


HEADER = [
    "科目编码", "科目名称",
    "期初借方", "期初贷方",
    "本期借方", "本期贷方",
    "期末借方", "期末贷方",
]


def _write_balance(path: Path, rows: list[tuple]) -> None:
    """Build a minimal .xlsx that the W1 three-tier parser accepts.

    Each row is ``(code, name, op_d, op_c, p_d, p_c, cl_d, cl_c)``.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "余额表"
    ws["A1"] = "测试公司 - 余额表"
    for i, h in enumerate(HEADER, start=1):
        ws.cell(row=2, column=i, value=h)
    for offset, r in enumerate(rows, start=3):
        for col, v in enumerate(r, start=1):
            ws.cell(row=offset, column=col, value=v)
    wb.save(str(path))
    wb.close()


def _build_w3_balance(path: Path, *, period_label: str) -> dict:
    """Build the W3-flavoured balance: top-level summary + 15 AP
    sub-accounts (so Stage 2 simplify has detail), one unknown-code
    row + one imbalanced row (so Stage 1 fires).

    Returns a small ``{"ap_total", "ap_details"}`` dict so callers can
    sanity-check their assertions against the synthetic numbers.
    """
    ap_amounts = [
        50000, 42000, 38000, 31000, 27000,
        21000, 18000, 15000, 12500, 10800,
        9200, 7500, 6100, 4800, 3300,
    ]
    ap_total = sum(ap_amounts)
    rows: list[tuple] = [
        # Header summary lines for the balance sheet.
        ("1001", "库存现金",     0, 0, 0, 0,  5_000, 0),
        ("1002", "银行存款",     0, 0, 0, 0, 250_000, 0),
        ("1122", "应收账款",     0, 0, 0, 0, 150_000, 0),
        ("1601", "固定资产",     0, 0, 0, 0, 1_200_000, 0),
        ("1602", "累计折旧",     0, 0, 0, 0, 0, 320_000),
    ]
    # 15 AP sub-accounts so the simplifier has something to fold; no
    # zero-amount parent placeholder so the row count under the
    # account_filter '^2202' is exactly 15.
    for i, amt in enumerate(ap_amounts, start=1):
        rows.append((
            f"2202.{i:02d}",
            f"应付账款-供应商{i:02d}",
            0, 0, 0, 0, 0, amt,
        ))
    rows.extend([
        # Stage 1 trigger #1 — unknown-class code 9001.
        ("9001", "未知科目",      0, 0, 0, 0, 1_000, 0),
        # Stage 1 trigger #2 — debit/credit imbalance: closing != opening + period.
        ("6601", "销售费用",      0, 0,  500, 0,  900, 0),
        # Equity rows to keep the balance sheet vaguely sensible.
        ("4001", "实收资本",      0, 0, 0, 0, 0,  900_000),
        ("4104", "未分配利润",    0, 0, 0, 0, 0,  17_700),
    ])
    _write_balance(path, rows)
    return {"ap_total": ap_total, "ap_details": ap_amounts, "period": period_label}


def _build_prior_balance(path: Path, *, ap_total: float) -> None:
    """Prior-period balance — same AP parent total but small drift on
    a couple of accounts so the cross-period validator finds at least
    one warning + one error diff."""
    rows = [
        ("1001", "库存现金",     0, 0, 0, 0,  4_500, 0),
        # Bank balance drifts 5_000 vs current's 250_000 -> 245_000 closing.
        # We want the *closing of prior* to equal the *opening of current*;
        # the validator compares closing_prior vs opening_current.  Since we
        # left openings of the current balance at zero, mismatches are obvious.
        ("1002", "银行存款",     0, 0, 0, 0, 245_000, 0),
        # AP closing same as current's running total — but we adjust two
        # sub-accounts so a couple of diffs surface.
        ("2202", "应付账款",     0, 0, 0, 0, 0, ap_total),
        ("4001", "实收资本",     0, 0, 0, 0, 0, 900_000),
        ("4104", "未分配利润",   0, 0, 0, 0, 0,   8_900),
    ]
    _write_balance(path, rows)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


async def _enable_encryption(db) -> None:
    km = KeyManager()
    salt = os.urandom(32)
    seed = os.urandom(32)
    km.unlock(seed, salt)
    await write_key_meta(
        db.conn,
        component=GLOBAL_COMPONENT,
        salt=salt,
        kdf_iterations=200_000,
        enabled=True,
        seed_source="env",
    )
    os.environ[ENV_PASSPHRASE] = seed.hex()


def _print_step(idx: int, label: str) -> None:
    print(f"\n=== Step {idx}: {label} ===")


def _upload(client: TestClient, base: str, *, org_id: str, period_id: str,
            path: Path) -> dict:
    with path.open("rb") as fh:
        r = client.post(
            f"{base}/orgs/{org_id}/imports",
            files={
                "file": (
                    path.name, fh,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            },
            data={"period_id": period_id},
        )
    assert r.status_code == 201, r.text
    return r.json()


# ---------------------------------------------------------------------------
# Main entrypoint
# ---------------------------------------------------------------------------


def run() -> int:  # noqa: PLR0912, PLR0915
    parser = argparse.ArgumentParser()
    parser.add_argument("--db", type=str, default=None)
    parser.add_argument("--keep", action="store_true",
                        help="leave the temp directory in place after the run")
    args = parser.parse_args()

    work = Path(tempfile.mkdtemp(prefix="finauto_w3_acc_"))
    print(f"workspace: {work}")
    db_path = Path(args.db) if args.db else work / "acceptance.sqlite"
    if db_path.exists():
        db_path.unlink()

    router, service, db = build_router_and_service(db_path)
    app = FastAPI(title="finance-auto W3 acceptance")
    app.include_router(router, prefix="/api/plugins/finance-auto")
    base = "/api/plugins/finance-auto"

    async def boot() -> None:
        await db.init()
        await _enable_encryption(db)
        await service.auto_unlock_if_configured()

    asyncio.run(boot())

    summary: dict[str, object] = {"steps": {}}
    fail = False
    client = TestClient(app)

    try:
        # ----------------------------------------------------------- Step 1
        _print_step(1, "Create restaurant org")
        r = client.post(
            f"{base}/orgs",
            json={
                "name": "W3 验收·餐饮试点", "code": "W3_ACCEPT_REST",
                "industry": "restaurant", "standard": "small",
            },
        )
        assert r.status_code == 201, r.text
        org_id = r.json()["id"]
        summary["steps"]["1_create_org"] = {
            "http": r.status_code, "id": org_id, "industry": "restaurant",
        }
        print(f"  org_id={org_id} industry=restaurant")

        # ----------------------------------------------------------- Step 2
        _print_step(2, "List industries (Stage 5)")
        r = client.get(f"{base}/industries")
        assert r.status_code == 200, r.text
        ind_payload = r.json()
        industries = {i["industry"] for i in ind_payload["items"]}
        assert {"general", "manufacturing", "restaurant", "tech_service"} <= industries
        summary["steps"]["2_list_industries"] = {
            "http": r.status_code,
            "industries": sorted(industries),
            "total": ind_payload["total"],
        }
        print(f"  industries={sorted(industries)}")

        # ----------------------------------------------------------- Step 3
        _print_step(3, "Effective-config for restaurant org (Stage 5)")
        r = client.get(f"{base}/orgs/{org_id}/effective-config")
        assert r.status_code == 200, r.text
        eff = r.json()
        assert eff["overlay_loaded"] is True, eff
        assert eff["effective"]["org_defaults"]["aux_mode"] == "light", eff
        # 7 base manual-input slots from cash_flow_aux.yaml; restaurant overlay
        # does not add or remove any so the count stays at 7.
        slot_keys = {s["key"] for s in eff["manual_input_slots_after_overlay"]}
        assert len(slot_keys) == 7, slot_keys
        summary["steps"]["3_effective_config"] = {
            "http": r.status_code,
            "overlay_keys": eff["overlay_keys"],
            "aux_mode_effective": eff["effective"]["org_defaults"]["aux_mode"],
            "manual_input_slot_count": len(slot_keys),
        }
        print(f"  overlay_keys={eff['overlay_keys']} "
              f"aux_mode={eff['effective']['org_defaults']['aux_mode']}")

        # ----------------------------------------------------------- Step 4
        _print_step(4, "Upload 2025-FY balance with 15 AP sub-accounts (Stage 1+2 setup)")
        bal_path = work / "balance_2025.xlsx"
        meta = _build_w3_balance(bal_path, period_label="2025-FY")
        upload = _upload(client, base, org_id=org_id, period_id="2025-FY",
                         path=bal_path)
        # Parse-issue counts surfaced by the upload route (Stage 1 hooks in).
        assert upload["parse_issues_detected"] >= 1, upload
        summary["steps"]["4_upload_current"] = {
            "http": 201,
            "row_count": upload["row_count"],
            "parse_issues_detected": upload["parse_issues_detected"],
            "parse_issues_must_fix": upload["parse_issues_must_fix"],
            "ap_total_input": meta["ap_total"],
        }
        current_import_id = upload["import_id"]
        print(f"  rows={upload['row_count']} "
              f"parse_issues={upload['parse_issues_detected']} "
              f"must_fix={upload['parse_issues_must_fix']}")

        # ----------------------------------------------------------- Step 5
        _print_step(5, "List + decide + learn one parse-issue (Stage 1)")
        r = client.get(f"{base}/orgs/{org_id}/parse-issues?status=pending")
        assert r.status_code == 200, r.text
        issues = r.json()["issues"]
        assert len(issues) >= 1, issues
        # Prefer the unknown_code issue (we know it exists).
        target = next((i for i in issues if i["issue_type"] == "unknown_code"),
                      issues[0])
        decide = client.post(
            f"{base}/orgs/{org_id}/parse-issues/{target['id']}/decide",
            json={"decision": "manual_fix",
                  "payload": {"action": "remap", "to": "1604"},
                  "decided_by": "acceptance-script"},
        )
        assert decide.status_code == 200, decide.text
        learn = client.post(
            f"{base}/orgs/{org_id}/parse-issues/{target['id']}/learn",
            json={"auto_apply": True, "share_globally": False,
                  "confidence": 0.95},
        )
        assert learn.status_code == 200, learn.text
        # learning-samples should now list at least the one we created.
        r = client.get(f"{base}/orgs/{org_id}/learning-samples")
        assert r.status_code == 200, r.text
        ls = r.json()
        assert ls["total"] >= 1, ls
        summary["steps"]["5_triage"] = {
            "pending_before": len(issues),
            "decided_issue_type": target["issue_type"],
            "decided_issue_id": target["id"],
            "learning_samples_total": ls["total"],
        }
        print(f"  decided issue={target['issue_type']!r} "
              f"learning_samples={ls['total']}")

        # ----------------------------------------------------------- Step 6
        _print_step(6, "Generate balance sheet (W2 + Stage 2 cells)")
        r = client.post(
            f"{base}/orgs/{org_id}/reports/balance_sheet/generate",
            json={"period_id": "2025-FY"},
        )
        assert r.status_code == 201, r.text
        body = r.json()
        report_id = body["report"]["id"]
        cells = body["cells"]
        bs_2202 = next(c for c in cells if c["reference_code"] == "BS_2202")
        # Total should round to our 15 sub-account sum.
        assert abs(bs_2202["value"] - meta["ap_total"]) < 0.5, bs_2202
        # Out of the gate the cell should NOT be simplified yet.
        assert bs_2202["simplified"] is False, bs_2202
        summary["steps"]["6_generate_bs"] = {
            "http": r.status_code, "report_id": report_id,
            "cell_count": body["report"]["cell_count"],
            "BS_2202_value": bs_2202["value"],
        }
        print(f"  report_id={report_id} cells={body['report']['cell_count']} "
              f"BS_2202={bs_2202['value']}")

        # ----------------------------------------------------------- Step 7
        _print_step(7, "PATCH BS_2202 simplify top_n=10 (Stage 2)")
        r = client.patch(
            f"{base}/orgs/{org_id}/reports/{report_id}/"
            f"cells/{bs_2202['id']}/simplify",
            json={"enabled": True, "strategy": "top_n", "top_n": 10,
                  "sort_by": "amount_desc",
                  "merge_label": "其他供应商",
                  "keep_negative_separate": True,
                  "footnote_template": "其他 {count} 家供应商合计 {amount}"},
        )
        assert r.status_code == 200, r.text
        patched = r.json()
        assert patched["simplified"] is True, patched
        assert patched["simplified_top_n"] == 10, patched
        merged_ids = patched["merged_row_ids"]
        # 15 detail rows – top 10 = 5 merged into "其他".
        assert len(merged_ids) == 5, merged_ids
        summary["steps"]["7_simplify_patch"] = {
            "http": r.status_code,
            "simplified": patched["simplified"],
            "top_n": patched["simplified_top_n"],
            "merged_count": len(merged_ids),
            "footnote": patched["footnote"],
        }
        print(f"  simplified={patched['simplified']} top_n=10 "
              f"merged={len(merged_ids)}")

        # ----------------------------------------------------------- Step 8
        _print_step(8, "GET cell details — top_n=10 + 1 'others' (Stage 2)")
        r = client.get(
            f"{base}/orgs/{org_id}/reports/{report_id}/"
            f"cells/{bs_2202['id']}/details"
        )
        assert r.status_code == 200, r.text
        details = r.json()
        # full_rows always carries every detail row.
        assert len(details["full_rows"]) == 15, details
        # visible_rows = 10 top + 1 'others'.
        assert len(details["visible_rows"]) == 11, details
        last = details["visible_rows"][-1]
        assert last["is_merged"] is True, last
        assert last["merged_count"] == 5, last
        # The 'others' amount should equal the sum of the 5 smallest details.
        smallest5 = sorted(meta["ap_details"])[:5]
        assert abs(last["amount"] - sum(smallest5)) < 0.5, last
        summary["steps"]["8_cell_details"] = {
            "http": r.status_code,
            "full_rows": len(details["full_rows"]),
            "visible_rows": len(details["visible_rows"]),
            "others_amount": last["amount"],
            "others_merged_count": last["merged_count"],
        }
        print(f"  full={len(details['full_rows'])} "
              f"visible={len(details['visible_rows'])} "
              f"others={last['amount']}/{last['merged_count']}")

        # ----------------------------------------------------------- Step 9
        _print_step(9, "Upload 2024-FY prior balance (Stage 3 setup)")
        prior_path = work / "balance_2024.xlsx"
        _build_prior_balance(prior_path, ap_total=meta["ap_total"])
        prior_upload = _upload(client, base, org_id=org_id,
                               period_id="2024-FY", path=prior_path)
        prior_import_id = prior_upload["import_id"]
        summary["steps"]["9_upload_prior"] = {
            "row_count": prior_upload["row_count"],
            "import_id": prior_import_id,
        }
        print(f"  prior rows={prior_upload['row_count']} "
              f"prior_import={prior_import_id}")

        # ----------------------------------------------------------- Step 10
        _print_step(10, "POST cross-period-checks (Stage 3)")
        r = client.post(
            f"{base}/orgs/{org_id}/cross-period-checks",
            json={
                "prior_period_id": "2024-FY",
                "current_period_id": "2025-FY",
                "prior_import_id": prior_import_id,
                "current_import_id": current_import_id,
                "tolerance": 1.0,
                "warn_threshold": 100.0,
                "emit_parse_issues": True,
            },
        )
        assert r.status_code == 201, r.text
        check = r.json()
        # We expect at least one error-graded diff (the 1002 bank
        # balance drift between 245k prior closing and 0 current
        # opening) and at least one parse_issue emitted from it.
        assert check["error_count"] >= 1, check
        assert len(check["parse_issue_ids"]) >= 1, check
        # And the listed issues now include cross_period_mismatch entries.
        r2 = client.get(
            f"{base}/orgs/{org_id}/parse-issues"
            f"?status=pending&issue_type=cross_period_mismatch"
        )
        assert r2.status_code == 200, r2.text
        xperiod_issues = r2.json()["issues"]
        assert len(xperiod_issues) >= 1, r2.json()
        summary["steps"]["10_cross_period"] = {
            "http": r.status_code,
            "check_id": check["id"],
            "total_accounts": check["total_accounts"],
            "error_count": check["error_count"],
            "warning_count": check["warning_count"],
            "parse_issues_emitted": len(check["parse_issue_ids"]),
            "xperiod_issues_listed": len(xperiod_issues),
        }
        print(f"  check_id={check['id']} errors={check['error_count']} "
              f"warnings={check['warning_count']} "
              f"emitted_issues={len(check['parse_issue_ids'])}")

        # ----------------------------------------------------------- Step 11
        _print_step(11, "PUT 7 manual_inputs + generate cash_flow (Stage 4)")
        r = client.get(
            f"{base}/orgs/{org_id}/periods/2025-FY/manual-inputs"
        )
        assert r.status_code == 200, r.text
        slots = r.json()["slots"]
        assert len(slots) == 7, slots
        # Values chosen so the SUM_LINES total is non-zero and predictable
        # for the test_manual_input_api regression set.
        seed_values = {
            "vat_output": "130000",
            "vat_input": "80000",
            "bill_discount_received": "5000",
            "interest_paid": "1200",
            "interest_income": "800",
            "bank_fee_paid": "600",
            "social_security_paid": "9000",
        }
        for k, v in seed_values.items():
            r = client.put(
                f"{base}/orgs/{org_id}/periods/2025-FY/manual-inputs/{k}",
                json={
                    "value": v,
                    "decided_by": "acceptance-script",
                    # Round-2 #1: ``expected_version`` is now required on every
                    # PUT.  Fresh slots use 0.
                    "expected_version": 0,
                },
            )
            assert r.status_code == 200, r.text
        # Now relist; every slot must be filled.
        r = client.get(
            f"{base}/orgs/{org_id}/periods/2025-FY/manual-inputs"
        )
        relisted = r.json()
        assert relisted["filled_count"] == 7, relisted
        # Generate cash_flow; should pull every manual_input value.
        r = client.post(
            f"{base}/orgs/{org_id}/reports/cash_flow/generate",
            json={"period_id": "2025-FY"},
        )
        assert r.status_code == 201, r.text
        cf = r.json()
        cf_cells = cf["cells"]
        # At least one cell must reference manual_input as its source.
        manual_refs = [
            c for c in cf_cells
            if any(s.startswith("manual_input:") for s in c.get("source_rows", []))
        ]
        assert manual_refs, cf_cells
        summary["steps"]["11_manual_inputs_and_cf"] = {
            "manual_inputs_filled": relisted["filled_count"],
            "cash_flow_report_id": cf["report"]["id"],
            "cash_flow_cell_count": cf["report"]["cell_count"],
            "manual_input_referencing_cells": len(manual_refs),
        }
        print(f"  manual_inputs filled={relisted['filled_count']}/7 "
              f"cash_flow cells={cf['report']['cell_count']} "
              f"manual_refs={len(manual_refs)}")

        # --------------------------------------------------------- DB inspect
        _print_step(12, "Inspect raw SQLite for W3 invariants")

        async def _shutdown() -> None:
            await db.close()
        asyncio.run(_shutdown())

        snap = sqlite3.connect(str(db_path))
        snap.row_factory = sqlite3.Row
        try:
            cur = snap.execute(
                "SELECT version FROM schema_version WHERE component='finance_auto'"
            )
            sv = cur.fetchone()
            schema_version = int(sv["version"]) if sv else 0

            cur = snap.execute("SELECT COUNT(*) AS n FROM parse_issues")
            n_parse = cur.fetchone()["n"]
            cur = snap.execute(
                "SELECT COUNT(*) AS n FROM parse_issues WHERE issue_type='cross_period_mismatch'"
            )
            n_xperiod_issues = cur.fetchone()["n"]
            cur = snap.execute("SELECT COUNT(*) AS n FROM learning_samples")
            n_learn = cur.fetchone()["n"]
            cur = snap.execute(
                "SELECT COUNT(*) AS n FROM cross_period_check_results"
            )
            n_xp = cur.fetchone()["n"]
            cur = snap.execute("SELECT COUNT(*) AS n FROM manual_inputs")
            n_mi = cur.fetchone()["n"]
            cur = snap.execute(
                "SELECT COUNT(*) AS n FROM report_cells WHERE simplified=1"
            )
            n_simplified_cells = cur.fetchone()["n"]
            cur = snap.execute(
                "SELECT COUNT(*) AS n FROM reports WHERE sheet_kind='cash_flow'"
            )
            n_cf_reports = cur.fetchone()["n"]
        finally:
            snap.close()

        # W3 baseline is schema v7; M2 workers bump it forward (AI→v8, Biz→v9).
        # Acceptance only requires that the W3 features still work, so accept
        # any version >= 7.
        assert schema_version >= 7, schema_version
        assert n_parse > 0
        assert n_xperiod_issues >= 1
        assert n_learn >= 1
        assert n_xp >= 1
        assert n_mi == 7
        assert n_simplified_cells >= 1
        assert n_cf_reports >= 1

        summary["steps"]["12_db_inspect"] = {
            "schema_version": schema_version,
            "parse_issues_total": n_parse,
            "cross_period_parse_issues": n_xperiod_issues,
            "learning_samples": n_learn,
            "cross_period_check_results": n_xp,
            "manual_inputs": n_mi,
            "simplified_cells": n_simplified_cells,
            "cash_flow_reports": n_cf_reports,
        }
        print(f"  schema_version={schema_version} parse_issues={n_parse} "
              f"xperiod_issues={n_xperiod_issues} samples={n_learn} "
              f"xp_checks={n_xp} manual_inputs={n_mi} "
              f"simplified_cells={n_simplified_cells} "
              f"cf_reports={n_cf_reports}")

    except AssertionError as exc:
        traceback.print_exc()
        summary["failure"] = f"AssertionError: {exc}"
        fail = True
    except Exception as exc:
        traceback.print_exc()
        summary["failure"] = f"{type(exc).__name__}: {exc}"
        fail = True
    finally:
        summary_path = Path("_m1_w3_acceptance_result.json").resolve()
        summary_path.write_text(
            json.dumps(summary, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        print(f"\nSummary written to {summary_path}")
        if not args.keep:
            try:
                shutil.rmtree(work, ignore_errors=True)
            except Exception:
                pass

    return 1 if fail else 0


if __name__ == "__main__":
    raise SystemExit(run())
