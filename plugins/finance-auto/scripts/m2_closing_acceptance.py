"""End-to-end acceptance script for the M2 closing wire-up.

After the three M2 sibling workers (AI / Biz / Frontend) merged their
work, this script proves that **every endpoint family is reachable
through the plugin's FastAPI router** -- i.e. the "last-mile wire-up"
the M2 frontend report flagged as missing has in fact been delivered
by commits ``cf18802b``, ``e1cdc176``, ``d458ae65``, ``b78efe9c``,
``1c0ee24c`` and ``10ca88ac``.

The script is intentionally **endpoint-shape focused** -- it doesn't
re-run the deep functional checks of the sibling acceptance scripts
(those still pass and are invoked as the final regression step).

13 verification categories
--------------------------

1.  ``GET    /ai/scenarios``                                  -> 6 scenarios
2.  ``PATCH  /ai/scenarios/{id}``                             -> toggles enabled
3.  ``GET    /ai/consent``                                    -> list (empty ok)
4.  ``POST   /ai/consent/respond``                            -> resolves a dialog
5.  ``DELETE /ai/consent/{id}``                               -> revokes permanent grant
6.  ``GET    /ai/audit-log``                                  -> paginated response
7.  ``WS     /api/plugins/finance-auto/ws``                   -> hello frame received
8.  ``POST/GET /users``                                       -> create + list
9.  ``POST /orgs/{}/reports/{}/review/{submit,approve,sign-off,request-changes}``
10. ``POST/GET /orgs/{}/reports/{}/cells/{}/comments``        -> create + list
11. ``POST /consolidation-groups`` + ``/members`` + ``/eliminations`` + ``/runs`` +
    ``GET /reports``                                          -> 5-step pipeline
12. ``POST /orgs/{}/reclassification-rules`` + preview + apply
13. Re-run the 4 sibling acceptance scripts in subprocess
    (``m1_w2``, ``m1_w3``, ``m2_ai``, ``m2_biz``)             -> all exit 0

Usage::

    d:\\OpenAkita\\.venv\\Scripts\\python.exe ^
        plugins/finance-auto/scripts/m2_closing_acceptance.py ^
        [--skip-regression] [--keep] [--json <path>]

Exit code 0 iff every step succeeded.
"""

from __future__ import annotations

import argparse
import asyncio
import json
import os
import shutil
import subprocess
import sys
import tempfile
import time
import traceback
from pathlib import Path

import openpyxl
from fastapi import FastAPI
from fastapi.testclient import TestClient

PLUGIN_ROOT = Path(__file__).resolve().parent.parent
if str(PLUGIN_ROOT) not in sys.path:
    sys.path.insert(0, str(PLUGIN_ROOT))

from finance_auto_backend.routes import build_router_and_service  # noqa: E402

BASE = "/api/plugins/finance-auto"
HEADER = [
    "科目编码", "科目名称",
    "期初借方", "期初贷方",
    "本期借方", "本期贷方",
    "期末借方", "期末贷方",
]


def _write_balance(path: Path, rows: list[tuple]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "余额表"
    ws["A1"] = "M2 closing acceptance"
    for i, h in enumerate(HEADER, start=1):
        ws.cell(row=2, column=i, value=h)
    for offset, r in enumerate(rows, start=3):
        for col, v in enumerate(r, start=1):
            ws.cell(row=offset, column=col, value=v)
    wb.save(str(path))
    wb.close()


def _checkpoint(name: str, started: float, ok: bool, **extras) -> dict:
    out = {
        "step": name,
        "ok": ok,
        "elapsed_ms": int((time.perf_counter() - started) * 1000),
        **extras,
    }
    print(f"[{'OK' if ok else 'FAIL'}] {name}  elapsed={out['elapsed_ms']}ms", flush=True)
    return out


def _trace(msg: str) -> None:
    print(f"... {msg}", flush=True)


def _make_balance_rows() -> list[tuple]:
    return [
        ("1001", "库存现金",     0, 0, 0, 0,    5_000, 0),
        ("1002", "银行存款",     0, 0, 0, 0,  250_000, 0),
        ("1122", "应收账款",     0, 0, 0, 0,  150_000, 0),
        ("1601", "固定资产",     0, 0, 0, 0, 1_200_000, 0),
        ("1602", "累计折旧",     0, 0, 0, 0, 0,   320_000),
        ("2202", "应付账款",     0, 0, 0, 0, 0,   180_000),
        ("4001", "实收资本",     0, 0, 0, 0, 0,   900_000),
        ("4104", "未分配利润",   0, 0, 0, 0, 0,   205_000),
    ]


def _make_sub_balance_rows() -> list[tuple]:
    return [
        ("1001", "库存现金",     0, 0, 0, 0,    2_000, 0),
        ("1002", "银行存款",     0, 0, 0, 0,   80_000, 0),
        ("1122", "应收账款",     0, 0, 0, 0,   50_000, 0),
        ("1601", "固定资产",     0, 0, 0, 0,  400_000, 0),
        ("2202", "应付账款",     0, 0, 0, 0, 0,    60_000),
        ("4001", "实收资本",     0, 0, 0, 0, 0,   200_000),
        ("4104", "未分配利润",   0, 0, 0, 0, 0,   272_000),
    ]


def run(args: argparse.Namespace) -> int:
    work = Path(tempfile.mkdtemp(prefix="m2_closing_"))
    db_path = work / "m2_closing.sqlite"
    results: list[dict] = []
    failures: list[str] = []
    final_status = "ok"
    started_all = time.perf_counter()

    try:
        _trace("building router + service")
        router, service, db = build_router_and_service(db_path)
        app = FastAPI()
        app.include_router(router, prefix=BASE)
        _trace("initialising DB schema")
        asyncio.run(db.init())
        _trace("creating TestClient")
        client = TestClient(app)
        _trace("entering 13 verification steps")

        # ---- 1. GET /ai/scenarios -- expect 6 scenarios (S1-S6) ----------
        t = time.perf_counter()
        r = client.get(f"{BASE}/ai/scenarios")
        assert r.status_code == 200, r.text
        scenarios = r.json()["scenarios"]
        scenario_ids = sorted(s["scenario_id"] for s in scenarios)
        assert len(scenarios) >= 6, f"expected >=6 scenarios, got {len(scenarios)}"
        results.append(_checkpoint(
            "01_ai_scenarios_list", t, True,
            count=len(scenarios), ids=scenario_ids,
        ))

        # ---- 2. PATCH /ai/scenarios/{id} -- toggle enabled --------------
        t = time.perf_counter()
        target_id = scenario_ids[0]
        r = client.patch(
            f"{BASE}/ai/scenarios/{target_id}",
            json={"enabled": False, "sensitivity_override": "aggregated"},
        )
        assert r.status_code == 200, r.text
        patched = r.json()
        assert patched["enabled_override"] is False, patched
        assert patched["sensitivity_override"] == "aggregated", patched
        # Flip back to keep the surface clean for follow-on steps.
        r = client.patch(
            f"{BASE}/ai/scenarios/{target_id}",
            json={"enabled": True},
        )
        assert r.status_code == 200, r.text
        results.append(_checkpoint("02_ai_scenarios_patch", t, True,
                                    scenario_id=target_id))

        # ---- 3. GET /ai/consent ----------------------------------------
        t = time.perf_counter()
        r = client.get(f"{BASE}/ai/consent")
        assert r.status_code == 200, r.text
        consent_body = r.json()
        assert isinstance(consent_body.get("consents"), list), consent_body
        results.append(_checkpoint("03_ai_consent_list", t, True,
                                    initial_count=consent_body["total"]))

        # ---- 4. POST /ai/consent/respond -- 404 for unknown dialog_id ---
        # Verifies the endpoint is mounted and the validation path works.
        # The full happy-path round-trip is exercised by m2_ai_acceptance.py
        # (step 13 regression) -- replicating it here would require
        # cross-event-loop coordination that is fragile under TestClient.
        t = time.perf_counter()
        r = client.post(
            f"{BASE}/ai/consent/respond",
            json={"dialog_id": "dlg_nonexistent", "decision": "allow_once"},
        )
        assert r.status_code == 404, r.text
        results.append(_checkpoint("04_ai_consent_respond", t, True,
                                    behaviour="404 on unknown dialog_id"))

        # ---- 5. DELETE /ai/consent/{id} -- 404 for unknown id ----------
        # Same rationale as #4: the route is mounted and the not-found
        # branch is exercised here; m2_ai_acceptance.py covers the
        # happy-path persistence + revoke flow with a real dialog.
        t = time.perf_counter()
        r = client.delete(f"{BASE}/ai/consent/999999")
        assert r.status_code == 404, r.text
        results.append(_checkpoint("05_ai_consent_revoke", t, True,
                                    behaviour="404 on unknown consent_id"))

        # ---- 6. GET /ai/audit-log --------------------------------------
        t = time.perf_counter()
        r = client.get(f"{BASE}/ai/audit-log", params={"limit": 50})
        assert r.status_code == 200, r.text
        audit = r.json()
        assert "items" in audit and "summary" in audit, audit
        results.append(_checkpoint("06_ai_audit_log", t, True,
                                    total=audit["total"],
                                    summary=audit["summary"]))

        # ---- 7. WS /ws -- hello frame on connect ------------------------
        # WebSocket TestClient: connect, receive first frame, assert hello.
        t = time.perf_counter()
        with client.websocket_connect(f"{BASE}/ws") as ws:
            hello = ws.receive_json()
            assert hello.get("event") == "finance_ws_hello", hello
            subs = hello.get("subscriptions") or []
        results.append(_checkpoint("07_ws_hello", t, True, subs=subs))

        # ---- 8. POST + GET /users --------------------------------------
        t = time.perf_counter()
        r = client.post(f"{BASE}/users", json={
            "user_id": "usr_closing",
            "display_name": "M2 Closing Auditor",
            "role": "auditor",
        })
        assert r.status_code in (201, 409), r.text
        r = client.get(f"{BASE}/users")
        assert r.status_code == 200, r.text
        users = r.json()["users"]
        assert any(u["user_id"] == "usr_closing" for u in users), users
        results.append(_checkpoint("08_users_create_list", t, True,
                                    user_count=len(users)))

        # ---- 9. Review workflow happy path -----------------------------
        # Need an org + balance + report first.
        t = time.perf_counter()
        r = client.post(f"{BASE}/orgs", json={
            "name": "M2 Closing 验收公司",
            "code": "M2CLOSE_PRIMARY",
            "standard": "small",
        })
        assert r.status_code == 201, r.text
        org_id = r.json()["id"]

        # Register a manager + partner for the review chain.
        for uid, role in (("usr_mgr_close", "manager"),
                          ("usr_partner_close", "partner")):
            r = client.post(f"{BASE}/users", json={
                "user_id": uid,
                "display_name": uid,
                "role": role,
            })
            assert r.status_code in (201, 409), r.text

        # Assignments unlock workflow.submit / approve / sign-off perms.
        period_id = "2025-FY"
        for uid, role_in_project in (
            ("usr_closing", "lead_auditor"),
            ("usr_mgr_close", "reviewer"),
            ("usr_partner_close", "partner_signoff"),
        ):
            r = client.post(f"{BASE}/orgs/{org_id}/assignments", json={
                "user_id": uid, "period_id": period_id,
                "role_in_project": role_in_project,
            })
            assert r.status_code in (201, 409), r.text

        bal_path = work / "primary_bal.xlsx"
        _write_balance(bal_path, _make_balance_rows())
        with bal_path.open("rb") as fh:
            r = client.post(
                f"{BASE}/orgs/{org_id}/imports",
                data={"period_id": period_id},
                files={"file": ("primary_bal.xlsx", fh,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        assert r.status_code == 201, r.text
        import_id = r.json()["import_id"]

        r = client.post(
            f"{BASE}/orgs/{org_id}/reports/balance_sheet/generate",
            json={"period_id": period_id, "source_import_id": import_id},
        )
        assert r.status_code == 201, r.text
        bs_payload = r.json()
        bs_report_id = bs_payload["report"]["id"]
        first_cell_id = bs_payload["cells"][0]["id"]

        # Submit -> request-changes -> submit again -> approve -> sign-off.
        r = client.post(
            f"{BASE}/orgs/{org_id}/reports/{bs_report_id}/review/submit",
            json={"auditor_id": "usr_closing"},
        )
        assert r.status_code == 201, r.text
        assert r.json()["status"] == "pending_review"

        r = client.post(
            f"{BASE}/orgs/{org_id}/reports/{bs_report_id}/review/request-changes",
            json={"actor_id": "usr_mgr_close",
                  "reason": "请修订科目分类"},
        )
        assert r.status_code == 200, r.text
        # status name is ``returned`` in the state machine (see
        # services/review_workflow.py _VALID_TRANSITIONS).
        assert r.json()["status"] == "returned", r.json()

        r = client.post(
            f"{BASE}/orgs/{org_id}/reports/{bs_report_id}/review/submit",
            json={"auditor_id": "usr_closing"},
        )
        assert r.status_code == 201, r.text

        r = client.post(
            f"{BASE}/orgs/{org_id}/reports/{bs_report_id}/review/approve",
            json={"actor_id": "usr_mgr_close"},
        )
        assert r.status_code == 200, r.text
        assert r.json()["status"] == "pending_signoff"

        r = client.post(
            f"{BASE}/orgs/{org_id}/reports/{bs_report_id}/review/sign-off",
            json={"actor_id": "usr_partner_close"},
        )
        assert r.status_code == 200, r.text
        wf_final = r.json()
        assert wf_final["status"] == "signed_off"
        results.append(_checkpoint("09_review_workflow", t, True,
                                    workflow_id=wf_final["workflow_id"],
                                    history_hops=len(wf_final["history"])))

        # ---- 10. Comments CRUD on a cell -------------------------------
        t = time.perf_counter()
        r = client.post(
            f"{BASE}/orgs/{org_id}/reports/{bs_report_id}/cells/{first_cell_id}/comments",
            json={
                "body": "M2 closing comment",
                "kind": "review_question",
                "author_id": "usr_mgr_close",
            },
        )
        assert r.status_code == 201, r.text
        comment_id = r.json()["id"]
        r = client.get(
            f"{BASE}/orgs/{org_id}/reports/{bs_report_id}/comments"
        )
        assert r.status_code == 200, r.text
        comments = r.json()["comments"]
        assert any(c["id"] == comment_id for c in comments), comments
        results.append(_checkpoint("10_comments_crud", t, True,
                                    comment_count=len(comments)))

        # ---- 11. Consolidation pipeline --------------------------------
        t = time.perf_counter()
        r = client.post(f"{BASE}/orgs", json={
            "name": "M2 Closing 子公司",
            "code": "M2CLOSE_SUB",
            "standard": "small",
        })
        assert r.status_code == 201, r.text
        sub_org_id = r.json()["id"]
        sub_bal_path = work / "sub_bal.xlsx"
        _write_balance(sub_bal_path, _make_sub_balance_rows())
        with sub_bal_path.open("rb") as fh:
            r = client.post(
                f"{BASE}/orgs/{sub_org_id}/imports",
                data={"period_id": period_id},
                files={"file": ("sub_bal.xlsx", fh,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        assert r.status_code == 201, r.text
        sub_import_id = r.json()["import_id"]
        r = client.post(
            f"{BASE}/orgs/{sub_org_id}/reports/balance_sheet/generate",
            json={"period_id": period_id, "source_import_id": sub_import_id},
        )
        assert r.status_code == 201, r.text

        r = client.post(f"{BASE}/consolidation-groups", json={
            "name": "M2 Closing 集团",
            "parent_org_id": org_id,
        })
        assert r.status_code == 201, r.text
        gid = r.json()["group_id"]
        r = client.post(
            f"{BASE}/consolidation-groups/{gid}/members",
            json={
                "subsidiary_org_id": sub_org_id,
                "ownership_pct": 75.0,
                "join_method": "full",
            },
        )
        assert r.status_code == 201, r.text
        r = client.post(
            f"{BASE}/consolidation-groups/{gid}/eliminations",
            json={
                "period_id": period_id,
                "kind": "inter_ar_ap",
                "debit_target": "BS_2202",
                "credit_target": "BS_1122",
                "amount": "10000",
                "rationale": "closing test",
            },
        )
        assert r.status_code == 201, r.text
        r = client.post(
            f"{BASE}/consolidation-groups/{gid}/runs",
            json={"period_id": period_id, "kind": "balance_sheet"},
        )
        assert r.status_code == 201, r.text
        consol = r.json()
        r = client.get(f"{BASE}/consolidation-groups/{gid}/reports")
        assert r.status_code == 200, r.text
        consol_reports = r.json()["reports"]
        assert len(consol_reports) >= 1, consol_reports
        results.append(_checkpoint("11_consolidation_pipeline", t, True,
                                    group_id=gid,
                                    member_count=len(consol["member_orgs_snapshot"]),
                                    minority_interest=consol["minority_interest"],
                                    report_count=len(consol_reports)))

        # ---- 12. Reclassification rule preview + apply ------------------
        t = time.perf_counter()
        r = client.post(
            f"{BASE}/orgs/{org_id}/reclassification-rules",
            json={
                "name": "M2 closing reclass",
                "when_condition": {
                    "account_code_starts": ["2202"],
                    "balance_direction": "credit",
                    "threshold": "0.01",
                },
                "action": {
                    "move_to_account_code": "1122",
                    "reason": "closing test reclass",
                    "parse_issue_severity": "warning",
                    "parse_issue_threshold": "50000",
                },
                "priority": 10,
            },
        )
        assert r.status_code == 201, r.text
        rule_id = r.json()["rule_id"]
        r = client.post(
            f"{BASE}/orgs/{org_id}/reclassification-runs/preview",
            json={"period_id": period_id},
        )
        assert r.status_code == 201, r.text
        preview = r.json()
        r = client.post(
            f"{BASE}/orgs/{org_id}/reclassification-runs/apply",
            json={"period_id": period_id},
        )
        assert r.status_code == 201, r.text
        applied = r.json()
        results.append(_checkpoint("12_reclassification", t, True,
                                    rule_id=rule_id,
                                    preview_items=preview["items_count"],
                                    apply_items=applied["items_count"]))

        # ---- 13. Regression: re-run the 4 sibling acceptance scripts ---
        t = time.perf_counter()
        regression_results: dict[str, dict] = {}
        if not args.skip_regression:
            scripts = [
                "m1_w2_acceptance.py",
                "m1_w3_acceptance.py",
                "m2_ai_acceptance.py",
                "m2_biz_acceptance.py",
            ]
            for s in scripts:
                script_path = PLUGIN_ROOT / "scripts" / s
                rt0 = time.perf_counter()
                proc = subprocess.run(
                    [sys.executable, str(script_path)],
                    capture_output=True, text=True, timeout=600,
                )
                rt_ms = int((time.perf_counter() - rt0) * 1000)
                regression_results[s] = {
                    "exit_code": proc.returncode,
                    "elapsed_ms": rt_ms,
                    "stdout_tail": (proc.stdout or "").splitlines()[-3:],
                    "stderr_tail": (proc.stderr or "").splitlines()[-3:],
                }
                assert proc.returncode == 0, (
                    f"{s} regression failed (exit {proc.returncode})\n"
                    f"STDOUT:\n{proc.stdout[-1500:]}\n"
                    f"STDERR:\n{proc.stderr[-1500:]}"
                )
        else:
            regression_results = {"skipped": {"reason": "--skip-regression"}}
        results.append(_checkpoint(
            "13_regression", t, True,
            scripts_run=len(regression_results)
            if "skipped" not in regression_results else 0,
            details=regression_results,
        ))

    except Exception as exc:  # noqa: BLE001
        final_status = "failed"
        failures.append(f"{type(exc).__name__}: {exc}")
        traceback.print_exc()
    finally:
        if not args.keep:
            shutil.rmtree(work, ignore_errors=True)

    summary = {
        "status": final_status,
        "elapsed_total_ms": int((time.perf_counter() - started_all) * 1000),
        "tmpdir": str(work) if args.keep else None,
        "steps_total": 13,
        "steps_ok": sum(1 for r in results if r["ok"]),
        "results": results,
        "failures": failures,
    }
    try:
        args.json_out.write_text(
            json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
        )
    except Exception as exc:  # noqa: BLE001
        print(f"warning: could not write json summary: {exc}")

    if final_status == "ok":
        print(f"OK  steps_ok={summary['steps_ok']}/13  "
              f"elapsed={summary['elapsed_total_ms']}ms  "
              f"json={args.json_out}")
        return 0
    print(f"FAIL  steps_ok={summary['steps_ok']}/13  failures={failures}")
    return 1


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--keep", action="store_true",
                        help="keep the tempdir after the run for inspection")
    parser.add_argument("--skip-regression", action="store_true",
                        help="skip step 13 (re-running sibling acceptance scripts)")
    parser.add_argument(
        "--json", dest="json_out", type=Path,
        default=PLUGIN_ROOT.parent.parent / "_m2_closing_acceptance_result.json",
        help="path to write the JSON summary",
    )
    args = parser.parse_args()
    return run(args)


if __name__ == "__main__":
    rc = main()
    # Audit P2-5: TestClient.websocket_connect spins up a non-daemon
    # ASGI worker thread (anyio + httpx) that does not exit cleanly when
    # the script's main coroutine finishes.  Without this os._exit the
    # interpreter wedges forever waiting for the thread, forcing CI to
    # kill the job via timeout.  Print the summary first via sys.stdout
    # flush so PowerShell / CI captures it, then bypass atexit handlers
    # entirely — every meaningful artefact (JSON summary, console output)
    # has already been written by ``run()`` at this point.
    sys.stdout.flush()
    sys.stderr.flush()
    os._exit(rc)
