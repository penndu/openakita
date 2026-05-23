"""Tests for audit-template upload / placeholder validation / render."""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient

PLUGIN_ROOT = Path(__file__).resolve().parent.parent
if str(PLUGIN_ROOT) not in sys.path:
    sys.path.insert(0, str(PLUGIN_ROOT))

from finance_auto_backend.parsers.xls_parser import ParsedRow  # noqa: E402
from finance_auto_backend.routes import build_router_and_service  # noqa: E402
from finance_auto_backend.services.audit_template import (  # noqa: E402
    build_allowlist,
    scan_template,
    validate_placeholders,
)


def _build_audit_template(path: Path, *, with_unknown: bool = False) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "底稿"
    ws["A1"] = "{{ org.name }} - {{ year }} 资产负债表抽样"
    ws["A2"] = "项目"
    ws["B2"] = "金额"
    ws["A3"] = "货币资金"
    ws["B3"] = "{{ cells.BS_1001.value }}"
    ws["A4"] = "应收账款"
    ws["B4"] = "{{ BS_1122 }}"
    ws["A5"] = "审计员"
    ws["B5"] = "{{ auditor }}"
    if with_unknown:
        ws["A6"] = "未知字段"
        ws["B6"] = "{{ cells.BOGUS_CODE.value }}"
    ws["A7"] = "{% if report.warnings %}有警告{% endif %}"
    wb.save(str(path))
    wb.close()


def test_scan_template_finds_placeholders(tmp_path: Path) -> None:
    p = tmp_path / "tpl.xlsx"
    _build_audit_template(p)
    placeholders = scan_template(p)
    primaries = {ph.primary_name for ph in placeholders}
    assert "org.name" in primaries
    assert "year" in primaries
    assert "cells.BS_1001.value" in primaries
    assert "BS_1122" in primaries
    assert "auditor" in primaries
    assert "report.warnings" in primaries


def test_validate_with_known_codes(tmp_path: Path) -> None:
    p = tmp_path / "tpl.xlsx"
    _build_audit_template(p)
    placeholders = scan_template(p)
    allow = build_allowlist(["BS_1001", "BS_1122", "BS_TOTAL_ASSETS"])
    rep = validate_placeholders(placeholders, allow)
    assert rep.unknown == []
    assert rep.is_strict_clean()
    assert "BS_1122" in rep.known


def test_validate_flags_unknown(tmp_path: Path) -> None:
    p = tmp_path / "tpl.xlsx"
    _build_audit_template(p, with_unknown=True)
    placeholders = scan_template(p)
    allow = build_allowlist(["BS_1001", "BS_1122"])
    rep = validate_placeholders(placeholders, allow)
    assert "cells.BOGUS_CODE.value" in rep.unknown
    assert not rep.is_strict_clean()


@pytest.mark.asyncio
async def test_audit_template_full_lifecycle(tmp_path: Path):
    db_path = tmp_path / "audit.sqlite"
    router, service, db = build_router_and_service(db_path)
    await db.init()
    app = FastAPI()
    app.include_router(router, prefix="/api/plugins/finance-auto")
    client = TestClient(app)
    try:
        # Seed: org, balance import, generated report so the allowlist
        # picks up the codes we use in the template.
        r = client.post(
            "/api/plugins/finance-auto/orgs",
            json={
                "name": "测试公司",
                "code": "AUDIT_DEMO",
                "industry": "general",
                "standard": "small",
            },
        )
        assert r.status_code == 201, r.text
        org_id = r.json()["id"]

        await service.ensure_period(org_id=org_id, period_id="2025-FY")
        imp = await service.insert_pending_import(
            org_id=org_id, period_id="2025-FY", source_file="seed.xlsx",
            file_size=0, file_sha256=None,
        )
        await service.persist_rows(
            import_id=imp.id, org_id=org_id, period_id="2025-FY",
            rows=[
                ParsedRow(
                    row_index=1, raw_code="1001", parent_code="1001",
                    child_code=None, full_code="1001", account_name="货币资金",
                    opening_debit=0, opening_credit=0, period_debit=0,
                    period_credit=0, closing_debit=10000, closing_credit=0,
                    aux_text=None,
                ),
                ParsedRow(
                    row_index=2, raw_code="1122", parent_code="1122",
                    child_code=None, full_code="1122", account_name="应收账款",
                    opening_debit=0, opening_credit=0, period_debit=0,
                    period_credit=0, closing_debit=50000, closing_credit=0,
                    aux_text=None,
                ),
            ],
        )
        await service.finalise_import(
            import_id=imp.id, parser_used="seed", row_count=2, status="ok",
            error_message=None,
        )
        r = client.post(
            f"/api/plugins/finance-auto/orgs/{org_id}/reports/balance_sheet/generate",
            json={"period_id": "2025-FY"},
        )
        assert r.status_code == 201
        report_id = r.json()["report"]["id"]

        # Upload audit template.
        tpl_path = tmp_path / "audit_tpl.xlsx"
        _build_audit_template(tpl_path)
        with tpl_path.open("rb") as fh:
            r = client.post(
                "/api/plugins/finance-auto/audit-templates",
                files={"file": ("audit_tpl.xlsx", fh,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                data={"name": "BS 抽样底稿"},
            )
        assert r.status_code == 201, r.text
        tpl = r.json()
        assert tpl["unknown_placeholder_count"] == 0
        tpl_id = tpl["id"]

        r = client.get("/api/plugins/finance-auto/audit-templates")
        assert r.status_code == 200
        assert r.json()["total"] == 1

        # Render.
        r = client.post(
            f"/api/plugins/finance-auto/orgs/{org_id}/audit-templates/{tpl_id}/render",
            json={"report_id": report_id, "strict": True},
        )
        assert r.status_code == 200, r.text
        out = tmp_path / "rendered.xlsx"
        out.write_bytes(r.content)
        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        assert ws["A1"].value == "测试公司 - 2025 资产负债表抽样"
        assert ws["B3"].value == pytest.approx(10000.0)
        assert ws["B4"].value == pytest.approx(50000.0)
        wb.close()
    finally:
        await db.close()


@pytest.mark.asyncio
async def test_strict_mode_rejects_unknown(tmp_path: Path):
    db_path = tmp_path / "audit_strict.sqlite"
    router, service, db = build_router_and_service(db_path)
    await db.init()
    app = FastAPI()
    app.include_router(router, prefix="/api/plugins/finance-auto")
    client = TestClient(app)
    try:
        r = client.post(
            "/api/plugins/finance-auto/orgs",
            json={
                "name": "测试公司",
                "code": "AUDIT_STRICT",
                "industry": "general",
                "standard": "small",
            },
        )
        org_id = r.json()["id"]

        # No report has been generated yet, so the allowlist is sparse and
        # cells.BS_1001.value is unknown.
        tpl_path = tmp_path / "audit_tpl.xlsx"
        _build_audit_template(tpl_path, with_unknown=False)
        with tpl_path.open("rb") as fh:
            r = client.post(
                "/api/plugins/finance-auto/audit-templates",
                files={"file": ("audit_tpl.xlsx", fh,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                data={"name": "Sparse template"},
            )
        assert r.status_code == 201
        tpl = r.json()
        assert tpl["unknown_placeholder_count"] >= 1

        r = client.post(
            f"/api/plugins/finance-auto/orgs/{org_id}/audit-templates/{tpl['id']}/render",
            json={"report_id": "rep_none", "strict": True},
        )
        assert r.status_code == 400
        assert "unknown" in r.json()["detail"]
    finally:
        await db.close()
