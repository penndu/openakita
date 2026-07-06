"""Integration tests for the W3 Stage 1 parse-issue + learning-sample APIs."""

from __future__ import annotations

import io
import sys
from pathlib import Path

import openpyxl
import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient

PLUGIN_ROOT = Path(__file__).resolve().parent.parent
if str(PLUGIN_ROOT) not in sys.path:
    sys.path.insert(0, str(PLUGIN_ROOT))

from finance_auto_backend.routes import build_router_and_service  # noqa: E402


def _build_balance_xlsx_with_anomalies() -> bytes:
    """Build a small xlsx that triggers several issue families."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "余额表"
    ws["A1"] = "演示账套 - 余额表"
    headers = ["科目编码", "科目名称",
               "期初借方", "期初贷方",
               "本期借方", "本期贷方",
               "期末借方", "期末贷方"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=2, column=i, value=h)
    rows = [
        ("1001", "库存现金",        0, 0, 0, 0, 5000, 0),    # clean
        ("1002", "银行存款",        0, 0, 0, 0, 200000, 0),  # clean
        # unknown account class — 8888 not in 1-7
        ("8888", "未知挂账",        0, 0, 0, 0, 1000, 0),
        # direction anomaly — 1001 with closing credit
        ("1010", "现金 - 一银行",    0, 0, 0, 0, 0, 500),
        # missing name (account_name = None)
        ("1601", None,            0, 0, 0, 0, 1000, 0),
    ]
    for offset, r in enumerate(rows, start=3):
        for col, v in enumerate(r, start=1):
            ws.cell(row=offset, column=col, value=v)
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


@pytest.fixture
async def api(tmp_path: Path):
    db_path = tmp_path / "parse_issue_api.sqlite"
    router, service, db = build_router_and_service(db_path)
    await db.init()
    app = FastAPI()
    app.include_router(router, prefix="/api/plugins/finance-auto")
    client = TestClient(app)
    try:
        yield client, service
    finally:
        await db.close()


@pytest.mark.asyncio
async def test_upload_detects_and_lists_parse_issues(api):
    client, _service = api
    base = "/api/plugins/finance-auto"

    r = client.post(
        f"{base}/orgs",
        json={"name": "演示账套", "code": "PI_TEST_1", "industry": "general",
              "standard": "small"},
    )
    assert r.status_code == 201, r.text
    org_id = r.json()["id"]

    data = _build_balance_xlsx_with_anomalies()
    r = client.post(
        f"{base}/orgs/{org_id}/imports",
        files={"file": ("balance.xlsx", data,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"period_id": "2025-FY"},
    )
    assert r.status_code == 201, r.text
    body = r.json()
    assert body["parse_issues_detected"] >= 3
    assert body["parse_issues_must_fix"] >= 2

    r = client.get(f"{base}/orgs/{org_id}/parse-issues")
    assert r.status_code == 200
    page = r.json()
    types = {i["issue_type"] for i in page["issues"]}
    assert "unknown_code" in types
    assert "direction_anomaly" in types
    assert "field_missing" in types
    assert page["pending"] == page["total"]


@pytest.mark.asyncio
async def test_decide_and_learn_then_auto_apply(api):
    client, _service = api
    base = "/api/plugins/finance-auto"

    r = client.post(
        f"{base}/orgs",
        json={"name": "演示账套2", "code": "PI_TEST_2", "industry": "general",
              "standard": "small"},
    )
    org_id = r.json()["id"]

    data = _build_balance_xlsx_with_anomalies()
    r = client.post(
        f"{base}/orgs/{org_id}/imports",
        files={"file": ("balance.xlsx", data,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"period_id": "2025-FY"},
    )
    assert r.status_code == 201
    issues = client.get(f"{base}/orgs/{org_id}/parse-issues").json()["issues"]
    target = next(i for i in issues if i["issue_type"] == "unknown_code")

    # Decide the issue first
    r = client.post(
        f"{base}/orgs/{org_id}/parse-issues/{target['id']}/decide",
        json={"decision": "manual_fix",
              "payload": {"target_parent_code": "1221"},
              "decided_by": "tester"},
    )
    assert r.status_code == 200, r.text
    assert r.json()["user_decision"] == "manual_fix"

    # Promote to a learning sample with auto_apply
    r = client.post(
        f"{base}/orgs/{org_id}/parse-issues/{target['id']}/learn",
        json={"auto_apply": True, "share_globally": False, "confidence": 0.95},
    )
    assert r.status_code == 200, r.text
    sample_id = r.json()["id"]
    assert r.json()["auto_apply"] is True

    samples = client.get(f"{base}/orgs/{org_id}/learning-samples").json()
    assert samples["total"] >= 1
    assert any(s["id"] == sample_id for s in samples["samples"])

    # Re-upload same file — auto_applied should fire for the 8888 row
    r2 = client.post(
        f"{base}/orgs/{org_id}/imports",
        files={"file": ("balance2.xlsx", data,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"period_id": "2025-FY"},
    )
    assert r2.status_code == 201
    assert r2.json()["parse_issues_auto_applied"] >= 1

    auto = client.get(
        f"{base}/orgs/{org_id}/parse-issues?status=auto_applied"
    ).json()
    assert auto["total"] >= 1
    assert auto["issues"][0]["auto_applied"] is True
