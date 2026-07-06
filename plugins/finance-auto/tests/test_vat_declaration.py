"""Tests for the Golden-Tax-IV VAT declaration parser + API.

Builds a synthetic .xlsx that mimics the canonical layout the central STA
template uses (title row + 5 labelled rows), runs both the pure parser and
the upload endpoint against it.
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient

PLUGIN_ROOT = Path(__file__).resolve().parent.parent
if str(PLUGIN_ROOT) not in sys.path:
    sys.path.insert(0, str(PLUGIN_ROOT))

from finance_auto_backend.parsers.vat_declaration import (  # noqa: E402
    VatParseError,
    parse_workbook,
    probe_source,
)
from finance_auto_backend.routes import build_router_and_service  # noqa: E402


def _build_synthetic_form(path: Path, *, province_hint: str | None = None) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "主表"

    title = "增值税及附加税费申报表（一般纳税人适用）"
    if province_hint:
        title = f"{province_hint} {title}"
    ws.merge_cells("A1:E1")
    ws["A1"] = title

    ws["A2"] = "项目"
    ws["B2"] = "一般项目"
    ws["C2"] = "即征即退项目"
    ws["D2"] = "本月数"
    ws["E2"] = "本年累计"

    ws["A3"] = "本期销项税额"
    ws["D3"] = 1_200_000.00
    ws["E3"] = 12_000_000.00

    ws["A4"] = "本期进项税额"
    ws["D4"] = 800_000.00
    ws["E4"] = 9_000_000.00

    ws["A5"] = "上期留抵税额"
    ws["D5"] = 50_000.00
    ws["E5"] = 0.00

    ws["A6"] = "应纳税额"
    ws["D6"] = 350_000.00
    ws["E6"] = 3_000_000.00

    ws["A7"] = "附加税费合计"
    ws["D7"] = 35_000.00

    wb.save(str(path))
    wb.close()


def test_probe_recognises_form(tmp_path: Path) -> None:
    p = tmp_path / "vat.xlsx"
    _build_synthetic_form(p)
    probe = probe_source(p)
    assert probe.dialect == "golden_tax_iv_generic"
    assert probe.confidence > 0.5


def test_probe_rejects_unrelated_workbook(tmp_path: Path) -> None:
    p = tmp_path / "balance.xlsx"
    wb = openpyxl.Workbook()
    wb.active["A1"] = "余额表"
    wb.save(str(p))
    wb.close()
    probe = probe_source(p)
    assert probe.dialect == "unknown"


def test_parse_extracts_canonical_fields(tmp_path: Path) -> None:
    p = tmp_path / "vat.xlsx"
    _build_synthetic_form(p)
    decl = parse_workbook(p, declaration_period="2025-01")
    assert decl.output_vat == pytest.approx(12_000_000.0)
    assert decl.input_vat == pytest.approx(9_000_000.0)
    assert decl.prev_credit == pytest.approx(50_000.0)
    assert decl.tax_payable == pytest.approx(3_000_000.0)
    assert decl.surtax_total == pytest.approx(35_000.0)


def test_parse_raises_on_unrecognised(tmp_path: Path) -> None:
    p = tmp_path / "balance.xlsx"
    wb = openpyxl.Workbook()
    wb.active["A1"] = "余额表"
    wb.save(str(p))
    wb.close()
    with pytest.raises(VatParseError):
        parse_workbook(p, declaration_period="2025-01")


def test_province_hint_recorded(tmp_path: Path) -> None:
    p = tmp_path / "vat_bj.xlsx"
    _build_synthetic_form(p, province_hint="北京")
    probe = probe_source(p)
    assert any("BJ" in n for n in probe.notes)


@pytest.mark.asyncio
async def test_vat_upload_endpoint(tmp_path: Path):
    db_path = tmp_path / "vat.sqlite"
    router, service, db = build_router_and_service(db_path)
    await db.init()
    app = FastAPI()
    app.include_router(router, prefix="/api/plugins/finance-auto")
    client = TestClient(app)
    try:
        r = client.post(
            "/api/plugins/finance-auto/orgs",
            json={
                "name": "测试公司",
                "code": "VAT_DEMO",
                "industry": "general",
                "standard": "small",
            },
        )
        assert r.status_code == 201, r.text
        org_id = r.json()["id"]

        form_path = tmp_path / "vat.xlsx"
        _build_synthetic_form(form_path)
        with form_path.open("rb") as fh:
            r = client.post(
                f"/api/plugins/finance-auto/orgs/{org_id}/vat-declarations",
                files={
                    "file": (
                        "vat.xlsx",
                        fh,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
                data={"declaration_period": "2025-01"},
            )
        assert r.status_code == 201, r.text
        body = r.json()
        assert body["dialect"] == "golden_tax_iv_generic"
        assert body["output_vat"] == pytest.approx(12_000_000.0)
        assert body["tax_payable"] == pytest.approx(3_000_000.0)

        r = client.get(f"/api/plugins/finance-auto/orgs/{org_id}/vat-declarations")
        assert r.status_code == 200
        assert r.json()["total"] == 1
    finally:
        await db.close()
