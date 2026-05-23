"""Sanity tests for the ReportRenderer abstraction.

Covers the small surface that is *not* exercised by the benchmark itself:

* :func:`make_renderer` selects the right backend based on the report kind
  + rows estimate.
* :class:`XltplRenderer` and :class:`OpenpyxlDirectRenderer` enforce the
  single-use contract.
* The openpyxl renderer raises a friendly ValueError when the template
  template lacks the ``__ROWS_HERE__`` anchor.
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
import pytest
from openpyxl.styles import Font

PLUGIN_ROOT = Path(__file__).resolve().parent.parent
if str(PLUGIN_ROOT) not in sys.path:
    sys.path.insert(0, str(PLUGIN_ROOT))

from finance_auto_backend.renderers import (  # noqa: E402
    OpenpyxlDirectRenderer,
    XltplRenderer,
    make_renderer,
)


def _build_xltpl_template(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BS"
    ws["A1"] = "{{ org.name }}"
    ws["A2"] = "项目"
    ws["B2"] = "代码"
    ws["C2"] = "金额"
    ws["A3"] = "货币资金"
    ws["B3"] = "1001"
    ws["C3"] = "{{ cells['1001'].value }}"
    wb.save(str(path))
    wb.close()


def _build_openpyxl_template(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AR"
    ws["A1"] = "{{ org.name }}"
    ws["A2"] = "code"
    ws["B2"] = "name"
    ws["C2"] = "amount"
    ws["A3"] = "account_code"
    ws["B3"] = "name"
    ws["C3"] = "amount"
    ws["A4"] = "__ROWS_HERE__"
    ws["A4"].font = Font(bold=True)
    ws["A5"] = "合计"
    ws["C5"] = "__SUM__"
    wb.save(str(path))
    wb.close()


def test_factory_picks_xltpl_for_static_small(tmp_path: Path) -> None:
    template = tmp_path / "bs.xlsx"
    _build_xltpl_template(template)
    renderer = make_renderer("balance_sheet", 20, template)
    assert isinstance(renderer, XltplRenderer)
    assert renderer.backend == "xltpl"


def test_factory_picks_openpyxl_for_large_static(tmp_path: Path) -> None:
    template = tmp_path / "bs.xlsx"
    _build_openpyxl_template(template)
    renderer = make_renderer("balance_sheet", 200, template)
    assert isinstance(renderer, OpenpyxlDirectRenderer)


def test_factory_picks_openpyxl_for_dynamic_kind(tmp_path: Path) -> None:
    template = tmp_path / "ar.xlsx"
    _build_openpyxl_template(template)
    renderer = make_renderer("ar_aging", 3, template)
    assert isinstance(renderer, OpenpyxlDirectRenderer)


def test_factory_rejects_negative_rows(tmp_path: Path) -> None:
    template = tmp_path / "bs.xlsx"
    _build_xltpl_template(template)
    with pytest.raises(ValueError):
        make_renderer("balance_sheet", -1, template)


def test_renderer_rejects_missing_template(tmp_path: Path) -> None:
    with pytest.raises(FileNotFoundError):
        XltplRenderer(tmp_path / "does-not-exist.xlsx")


def test_xltpl_renderer_writes_value(tmp_path: Path) -> None:
    template = tmp_path / "bs.xlsx"
    _build_xltpl_template(template)
    out = tmp_path / "bs_out.xlsx"
    renderer = XltplRenderer(template)
    ctx = {
        "report_kind": "balance_sheet",
        "standard": "small_enterprise",
        "year": 2025,
        "period": "2025-12",
        "org": {"name": "测试公司"},
        "cells": [
            {"code": "1001", "label": "货币资金", "value": 12345.67},
        ],
    }
    result = renderer.render(ctx, out)
    assert result.backend == "xltpl"
    assert out.exists()
    assert result.rows_written == 1
    wb = openpyxl.load_workbook(str(out))
    ws = wb.active
    assert ws["A1"].value == "测试公司"
    assert float(ws["C3"].value) == pytest.approx(12345.67)
    wb.close()

    with pytest.raises(RuntimeError):
        renderer.render(ctx, out)


def test_openpyxl_renderer_writes_rows(tmp_path: Path) -> None:
    template = tmp_path / "ar.xlsx"
    _build_openpyxl_template(template)
    out = tmp_path / "ar_out.xlsx"
    renderer = OpenpyxlDirectRenderer(template)
    rows = [
        {"account_code": "1122.001", "name": "客户 A", "amount": 100.0},
        {"account_code": "1122.002", "name": "客户 B", "amount": 200.0},
        {"account_code": "1122.003", "name": "客户 C", "amount": 300.0},
    ]
    ctx = {
        "report_kind": "ar_aging",
        "standard": "general_enterprise",
        "year": 2025,
        "period": "2025-12",
        "org": {"name": "测试公司"},
        "rows": rows,
    }
    result = renderer.render(ctx, out)
    assert result.rows_written == 3
    wb = openpyxl.load_workbook(str(out))
    ws = wb.active
    assert ws["A1"].value == "测试公司"
    assert ws["A4"].value == "1122.001"
    assert ws["A5"].value == "1122.002"
    assert ws["A6"].value == "1122.003"
    sum_cell = ws["C7"].value
    assert isinstance(sum_cell, str) and sum_cell.startswith("=SUM(C4:C6")
    wb.close()


def test_openpyxl_renderer_raises_without_marker(tmp_path: Path) -> None:
    template = tmp_path / "bad.xlsx"
    wb = openpyxl.Workbook()
    wb.active["A1"] = "no marker"
    wb.save(str(template))
    wb.close()
    with pytest.raises(ValueError, match="__ROWS_HERE__"):
        OpenpyxlDirectRenderer(template).render({"rows": []}, tmp_path / "x.xlsx")
