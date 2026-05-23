"""Audit-template upload + Jinja2 placeholder validation + render.

Auditors today maintain a per-engagement folder of working-paper templates
(``审计底稿`` / 套表).  Each template is a hand-laid .xlsx that uses
Jinja-style placeholders to pull values out of a balance sheet / PL /
cash-flow report.  The plugin's audit-template service:

1. Accepts the .xlsx upload.
2. Scans every cell of every sheet for Jinja placeholders
   (``{{ ... }}`` and ``{% ... %}``).
3. Cross-checks the placeholder names against a known-field allow-list
   (built dynamically from the four shipped YAML report templates' rule
   reference_codes plus a tiny "context" prefix list).
4. Stores the placeholder report on the ``audit_templates`` row.
5. Provides a render endpoint that loads the latest matching report by
   accounting standard / sheet kind for the given org and substitutes the
   placeholders, then streams the rendered .xlsx back.

The placeholder validator is intentionally conservative -- any token that
looks like Jinja but does not match a known field is reported as ``unknown``
rather than discarded.  The render endpoint refuses to render templates
that have ``unknown`` tokens unless the caller passes ``strict=False``.
"""

from __future__ import annotations

import json
import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Placeholder scan
# ---------------------------------------------------------------------------


_JINJA_VAR_RE = re.compile(r"\{\{\s*(.+?)\s*\}\}")
_JINJA_TAG_RE = re.compile(r"\{%\s*(.+?)\s*%\}")
_IDENT_RE = re.compile(r"[A-Za-z_][A-Za-z0-9_]*(?:\.[A-Za-z_][A-Za-z0-9_]*)*")


@dataclass(frozen=True)
class Placeholder:
    """One placeholder occurrence found in a template cell."""

    sheet: str
    cell: str           # A1 reference
    raw: str            # full ``{{ ... }}`` / ``{% ... %}`` text
    kind: str           # "var" | "tag"
    primary_name: str   # the dotted identifier that we cross-check


@dataclass
class PlaceholderReport:
    placeholders: list[Placeholder] = field(default_factory=list)
    known: list[str] = field(default_factory=list)
    unknown: list[str] = field(default_factory=list)

    def is_strict_clean(self) -> bool:
        return not self.unknown


# ---------------------------------------------------------------------------
# Allowlist builder -- combines the four shipped YAML report-template rule
# reference codes with a small static context prefix list.
# ---------------------------------------------------------------------------


_STATIC_CONTEXT_NAMES: frozenset[str] = frozenset(
    {
        "org",                # whole Organization model
        "org.id",
        "org.name",
        "org.code",
        "org.industry",
        "org.standard",
        "org.fiscal_start",
        "report",             # whole ReportInstance
        "report.id",
        "report.period_id",
        "report.sheet_kind",
        "report.accounting_standard",
        "report.template_id",
        "report.generated_at",
        "year",
        "period",
        "standard",
        "today",
        "now",
        "auditor",            # free-form identity field auditors fill in
        "engagement_id",
    }
)


def build_allowlist(reference_codes: list[str]) -> set[str]:
    """Return the set of dotted identifiers a template may safely reference.

    Each report-cell ``BS_1001`` becomes both the bare code (so
    ``{{ BS_1001 }}`` works) and a prefixed form ``cells.BS_1001`` (the
    namespaced form we recommend to template authors).
    """

    allow = set(_STATIC_CONTEXT_NAMES)
    for code in reference_codes:
        allow.add(code)
        allow.add(f"cells.{code}")
        allow.add(f"cells.{code}.value")
        allow.add(f"cells.{code}.label")
    return allow


# ---------------------------------------------------------------------------
# Template scanning
# ---------------------------------------------------------------------------


def scan_template(path: Path | str) -> list[Placeholder]:
    """Collect every Jinja-style placeholder in the template."""

    import openpyxl

    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"audit template not found: {p}")

    wb = openpyxl.load_workbook(str(p), data_only=False)
    out: list[Placeholder] = []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if not isinstance(cell.value, str):
                        continue
                    for raw, primary, kind in _scan_cell(cell.value):
                        out.append(
                            Placeholder(
                                sheet=sheet_name,
                                cell=cell.coordinate,
                                raw=raw,
                                kind=kind,
                                primary_name=primary,
                            )
                        )
    finally:
        wb.close()
    return out


def _scan_cell(text: str) -> list[tuple[str, str, str]]:
    """Find all ``{{ ... }}`` and ``{% ... %}`` occurrences in a cell."""
    found: list[tuple[str, str, str]] = []
    for m in _JINJA_VAR_RE.finditer(text):
        primary = _primary_identifier(m.group(1))
        found.append((m.group(0), primary, "var"))
    for m in _JINJA_TAG_RE.finditer(text):
        primary = _primary_identifier(m.group(1))
        found.append((m.group(0), primary, "tag"))
    return found


def _primary_identifier(expr: str) -> str:
    """Return the leading dotted identifier of a placeholder expression.

    Examples::

        cells.BS_1001.value | round(2)   ->   "cells.BS_1001.value"
        for cell in cells                ->   "cells"
        if report.warnings | length > 0  ->   "report.warnings"
    """
    cleaned = expr.strip()
    # Tag forms: drop the leading keyword ("for", "if", etc.).
    if cleaned.startswith(("for ", "if ", "elif ", "endfor", "endif", "set ")):
        cleaned = cleaned.split(" ", 1)[1] if " " in cleaned else ""
    if " in " in cleaned:
        cleaned = cleaned.split(" in ", 1)[1]
    m = _IDENT_RE.match(cleaned.strip())
    return m.group(0) if m else cleaned.strip()


def validate_placeholders(
    placeholders: list[Placeholder], allowlist: set[str]
) -> PlaceholderReport:
    report = PlaceholderReport(placeholders=placeholders)
    seen_known: set[str] = set()
    seen_unknown: set[str] = set()
    for ph in placeholders:
        name = ph.primary_name
        if not name:
            continue
        # A few common Jinja2 builtins we always allow.
        if name in {"loop", "_", "true", "false", "none", "True", "False", "None"}:
            continue
        if name in allowlist or any(name.startswith(a + ".") for a in allowlist):
            seen_known.add(name)
        else:
            seen_unknown.add(name)
    report.known = sorted(seen_known)
    report.unknown = sorted(seen_unknown)
    return report


# ---------------------------------------------------------------------------
# Render
# ---------------------------------------------------------------------------


def render_template(
    template_path: Path | str,
    output_path: Path | str,
    *,
    context: dict[str, Any],
) -> Path:
    """Render an audit template by stamping cell text through Jinja2.

    Implementation note: we reach into Jinja2 directly (not openpyxl-jinja or
    xltpl) because the placeholder vocabulary is much smaller than what
    xltpl supports and we want consistent behaviour across cells, including
    formula cells -- xltpl rewrites formulas in confusing ways.
    """

    import openpyxl
    from jinja2 import Environment, StrictUndefined, TemplateError

    template_path = Path(template_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    env = Environment(
        autoescape=False,
        undefined=StrictUndefined,
        keep_trailing_newline=True,
    )

    wb = openpyxl.load_workbook(str(template_path))
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if not isinstance(cell.value, str):
                        continue
                    text = cell.value
                    if "{{" not in text and "{%" not in text:
                        continue
                    try:
                        rendered = env.from_string(text).render(**context)
                    except TemplateError as exc:
                        logger.warning(
                            "audit-template render: cell %s/%s failed: %s",
                            ws.title, cell.coordinate, exc,
                        )
                        continue
                    if rendered != text:
                        # Try float coercion for numeric cells (lets
                        # auditors do {{ cells.BS_1001.value }} and still
                        # get a number-formatted result).
                        try:
                            cell.value = float(rendered)
                        except (TypeError, ValueError):
                            cell.value = rendered
        wb.save(str(output_path))
        return output_path
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Storage helpers
# ---------------------------------------------------------------------------


def serialise_report(report: PlaceholderReport) -> str:
    return json.dumps(
        {
            "placeholders": [
                {
                    "sheet": ph.sheet,
                    "cell": ph.cell,
                    "raw": ph.raw,
                    "kind": ph.kind,
                    "primary_name": ph.primary_name,
                }
                for ph in report.placeholders
            ],
            "known": report.known,
            "unknown": report.unknown,
        },
        ensure_ascii=False,
    )


def deserialise_report(blob: str | None) -> PlaceholderReport:
    if not blob:
        return PlaceholderReport()
    raw = json.loads(blob)
    return PlaceholderReport(
        placeholders=[
            Placeholder(
                sheet=p["sheet"], cell=p["cell"], raw=p["raw"],
                kind=p["kind"], primary_name=p["primary_name"],
            )
            for p in raw.get("placeholders", [])
        ],
        known=list(raw.get("known", [])),
        unknown=list(raw.get("unknown", [])),
    )
