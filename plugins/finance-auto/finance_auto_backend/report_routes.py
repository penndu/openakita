"""Report-generation endpoints (M1 W2 Stage 4).

Mounted onto the same APIRouter created by :func:`routes.build_router` -- we
keep the W1 endpoint surface byte-identical and bolt these on via
:func:`register_report_endpoints` so the W1 file stays small.

Four endpoints:

* ``POST /orgs/{org_id}/reports/{kind}/generate``
* ``GET  /orgs/{org_id}/reports``
* ``GET  /orgs/{org_id}/reports/{report_id}``
* ``GET  /orgs/{org_id}/reports/{report_id}/export?format=xlsx``

Decimals + Chinese labels round-trip through JSON cleanly.
"""

from __future__ import annotations

import json
import logging
import tempfile
from pathlib import Path
from typing import TYPE_CHECKING, Any

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import FileResponse

from .rbac import require_permission

from .config.yaml_loader import LoadedTemplate, load_template
from .models import (
    CellDetailRow,
    CellDetailsResponse,
    CellSimplifyPatchRequest,
    ReportCell,
    ReportDetailResponse,
    ReportGenerateRequest,
    ReportInstance,
    ReportListResponse,
)
from .renderers.simplifier import (
    DetailRow,
    SimplifyConfig,
    simplify_aux_details,
)
from .report_generator import (
    GeneratedReport,
    TrialBalanceLine,
    _balance_kind_to_amount,
    generate_report,
)

if TYPE_CHECKING:
    from .routes import FinanceAutoService

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Template-resolution helpers
# ---------------------------------------------------------------------------


_TEMPLATE_DIR_CANDIDATES = (
    Path(__file__).resolve().parent.parent / "templates" / "reports",
)
"""Where the YAML templates live.  Relative to the backend package; the host
plugin loader resolves the same path because plugin.py copies templates/
alongside the backend code into the runtime tree."""

_KIND_TO_FILENAME = {
    "balance_sheet:small_enterprise": "balance_sheet_small_enterprise.yaml",
    "balance_sheet:general_enterprise": "balance_sheet_general_enterprise.yaml",
    "income_statement:small_enterprise": "income_statement_small_enterprise.yaml",
    "income_statement:general_enterprise": "income_statement_general_enterprise.yaml",
    # M1 W3 Stage 4: minimal cash-flow template wired to manual_inputs.
    "cash_flow:small_enterprise": "cash_flow_small_enterprise.yaml",
    # M2 Biz Stage 4: indirect-method cash-flow for the general / CAS
    # standard.  The template was shipped (cf_indirect_ge_v1) but never
    # wired into the resolver, so the default CAS org -> general_enterprise
    # path returned 400.  It consumes the cf_* synthetic keys produced by
    # IndirectCashFlowEngine (via /cash-flow/persist) through the existing
    # manual_input data_source -- no generator change needed.
    "cash_flow:general_enterprise": "cash_flow_indirect_general_enterprise.yaml",
}


def _resolve_template_path(kind: str, standard: str) -> Path:
    key = f"{kind}:{standard}"
    filename = _KIND_TO_FILENAME.get(key)
    if filename is None:
        raise HTTPException(
            status_code=400,
            detail=f"unsupported (kind, standard) combination: {key!r}",
        )
    for root in _TEMPLATE_DIR_CANDIDATES:
        candidate = root / filename
        if candidate.exists():
            return candidate
    raise HTTPException(
        status_code=500, detail=f"template file not found: {filename}"
    )


def _standard_for_org(org_standard: str, override: str | None) -> str:
    if override:
        return override
    if org_standard == "small":
        return "small_enterprise"
    return "general_enterprise"


# ---------------------------------------------------------------------------
# DB helpers (leveraging the W1 service for connection access)
# ---------------------------------------------------------------------------


async def _load_balance_lines(
    service: FinanceAutoService,
    *,
    org_id: str,
    period_id: str,
    source_import_id: str | None,
) -> tuple[list[TrialBalanceLine], str]:
    if source_import_id is None:
        async with service.db.conn.execute(
            "SELECT id FROM trial_balance_imports "
            "WHERE org_id=? AND period_id=? AND status='ok' "
            "ORDER BY uploaded_at DESC LIMIT 1",
            (org_id, period_id),
        ) as cur:
            row = await cur.fetchone()
        if row is None:
            raise HTTPException(
                status_code=404,
                detail=(
                    f"no successful balance-table import for org={org_id} "
                    f"period={period_id}"
                ),
            )
        source_import_id = row[0]

    rows_typed = await service.list_all_rows(
        org_id=org_id, import_id=source_import_id
    )
    lines = [
        TrialBalanceLine(
            id=r.id,
            full_code=r.full_code,
            parent_code=r.parent_code,
            child_code=r.child_code,
            account_name=r.account_name,
            opening_debit=r.opening_debit,
            opening_credit=r.opening_credit,
            period_debit=r.period_debit,
            period_credit=r.period_credit,
            closing_debit=r.closing_debit,
            closing_credit=r.closing_credit,
            aux_text=r.aux_text,
        )
        for r in rows_typed
    ]
    return lines, source_import_id


async def _persist_report(
    service: FinanceAutoService,
    *,
    template: LoadedTemplate,
    generated: GeneratedReport,
) -> None:
    inst = generated.instance
    await service.db.conn.execute(
        "INSERT INTO reports(id, org_id, period_id, sheet_kind, "
        "accounting_standard, template_id, template_version, status, "
        "cell_count, warnings_json, source_import_id, backend_used, "
        "output_path, generated_at) "
        "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        (
            inst.id,
            inst.org_id,
            inst.period_id,
            inst.sheet_kind,
            inst.accounting_standard,
            inst.template_id,
            inst.template_version,
            inst.status,
            inst.cell_count,
            json.dumps(inst.warnings, ensure_ascii=False),
            inst.source_import_id,
            inst.backend_used,
            inst.output_path,
            inst.generated_at,
        ),
    )
    await service.db.conn.executemany(
        "INSERT INTO report_cells(id, report_id, reference_code, target_line_no, "
        "target_label, indent_level, data_source, code, value, sign, is_total, "
        "is_tbd, formula, notes, source_rows, simplified, simplified_top_n, "
        "simplify_config_json, merged_row_ids_json, footnote, version) "
        "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        [
            (
                c.id,
                c.report_id,
                c.reference_code,
                c.target_line_no,
                c.target_label,
                c.indent_level,
                c.data_source,
                c.code,
                c.value,
                c.sign,
                int(c.is_total),
                int(c.is_tbd),
                c.formula,
                c.notes,
                json.dumps(c.source_rows, ensure_ascii=False),
                int(c.simplified),
                c.simplified_top_n,
                (
                    json.dumps(c.simplify_config, ensure_ascii=False)
                    if c.simplify_config
                    else None
                ),
                json.dumps(c.merged_row_ids, ensure_ascii=False) if c.merged_row_ids else None,
                c.footnote,
                1,  # initial version
            )
            for c in generated.cells
        ],
    )
    await service.db.conn.commit()


def _row_to_report(row: Any) -> ReportInstance:
    warnings = json.loads(row["warnings_json"] or "[]")
    return ReportInstance(
        id=row["id"],
        org_id=row["org_id"],
        period_id=row["period_id"],
        sheet_kind=row["sheet_kind"],
        accounting_standard=row["accounting_standard"],
        template_id=row["template_id"],
        template_version=row["template_version"] or 1,
        status=row["status"] or "ok",
        cell_count=row["cell_count"] or 0,
        warnings=warnings,
        source_import_id=row["source_import_id"],
        backend_used=row["backend_used"],
        output_path=row["output_path"],
        generated_at=row["generated_at"],
    )


def _row_to_cell(row: Any) -> ReportCell:
    sources = json.loads(row["source_rows"] or "[]")
    keys = row.keys()
    simplified_cfg = None
    if "simplify_config_json" in keys and row["simplify_config_json"]:
        try:
            simplified_cfg = json.loads(row["simplify_config_json"])
        except json.JSONDecodeError:
            simplified_cfg = None
    merged_ids: list[str] = []
    if "merged_row_ids_json" in keys and row["merged_row_ids_json"]:
        try:
            merged_ids = json.loads(row["merged_row_ids_json"]) or []
        except json.JSONDecodeError:
            merged_ids = []
    return ReportCell(
        id=row["id"],
        report_id=row["report_id"],
        reference_code=row["reference_code"],
        target_line_no=row["target_line_no"],
        target_label=row["target_label"],
        indent_level=row["indent_level"],
        data_source=row["data_source"],
        code=row["code"],
        value=row["value"] or 0.0,
        sign=row["sign"] or 1,
        is_total=bool(row["is_total"]),
        is_tbd=bool(row["is_tbd"]),
        formula=row["formula"],
        notes=row["notes"],
        source_rows=sources,
        simplified=bool(row["simplified"]) if "simplified" in keys else False,
        simplified_top_n=int(row["simplified_top_n"] or 0) if "simplified_top_n" in keys else 0,
        simplify_config=simplified_cfg,
        merged_row_ids=merged_ids,
        footnote=row["footnote"] if "footnote" in keys else None,
    )


# ---------------------------------------------------------------------------
# Excel export (programmatic openpyxl writer; the YAML xltpl_file references
# remain forward-looking until the design team ships hand-laid templates).
# ---------------------------------------------------------------------------


def _render_report_sheet(
    ws, template: LoadedTemplate, instance: ReportInstance,
    cells: list[ReportCell],
) -> None:
    """Lay one report (header + cell rows) onto an openpyxl worksheet.

    Extracted from ``_build_workbook`` so the single-report export and the
    multi-sheet bundle export share byte-identical formatting.
    """
    from openpyxl.styles import Alignment, Font, PatternFill

    title = (
        f"{template.name} - {instance.period_id} "
        f"({instance.accounting_standard})"
    )
    ws.merge_cells("A1:D1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["项目", "代码", "金额", "备注"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DDEBF7")

    cells_sorted = sorted(cells, key=lambda c: (c.target_line_no, c.reference_code))
    section_font = Font(bold=True, color="305496")
    total_font = Font(bold=True)
    total_fill = PatternFill("solid", fgColor="FFF2CC")
    tbd_fill = PatternFill("solid", fgColor="F8CBAD")
    # W3 Stage 2: 简化行 (含"其他"合并) 用灰色背景 + 斜体 区分
    simplified_fill = PatternFill("solid", fgColor="EEEEEE")
    simplified_font = Font(italic=True, color="595959")

    row_idx = 3
    for cell in cells_sorted:
        indent = "  " * (cell.indent_level or 0)
        label = f"{indent}{cell.target_label}"
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=cell.code or cell.reference_code)
        ws.cell(row=row_idx, column=3, value=float(cell.value))
        ws.cell(row=row_idx, column=3).number_format = "#,##0.00"
        notes = cell.notes or ""
        if cell.is_tbd:
            notes = f"[TBD] {notes}".strip()
        if cell.simplified and cell.footnote:
            notes = (f"{notes} | {cell.footnote}").strip(" |")
        ws.cell(row=row_idx, column=4, value=notes)
        if cell.data_source == "section":
            ws.cell(row=row_idx, column=1).font = section_font
        if cell.is_total:
            for col in range(1, 5):
                ws.cell(row=row_idx, column=col).font = total_font
                ws.cell(row=row_idx, column=col).fill = total_fill
        if cell.is_tbd:
            for col in range(1, 5):
                ws.cell(row=row_idx, column=col).fill = tbd_fill
        if cell.simplified and not cell.is_total:
            for col in range(1, 5):
                ws.cell(row=row_idx, column=col).fill = simplified_fill
                ws.cell(row=row_idx, column=col).font = simplified_font
        row_idx += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 40


def _build_workbook(template: LoadedTemplate, instance: ReportInstance,
                    cells: list[ReportCell]) -> Path:
    import openpyxl

    fd, name = tempfile.mkstemp(suffix=".xlsx", prefix="finauto_report_")
    import os as _os
    _os.close(fd)
    out = Path(name)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = template.sheet_kind[:31]
    _render_report_sheet(ws, template, instance, cells)
    wb.save(str(out))
    wb.close()
    return out


# Stable sheet-name labels for the one-click bundle (Excel caps at 31 chars).
_BUNDLE_SHEET_TITLES = {
    "balance_sheet": "资产负债表",
    "income_statement": "利润表",
    "cash_flow": "现金流量表",
}


def _build_bundle_workbook(
    org_name: str,
    period_id: str,
    sections: list[tuple[LoadedTemplate, ReportInstance, list[ReportCell]]],
) -> Path:
    """Pack BS / IS / CF (plus a cover + audit summary) into one workbook.

    ``sections`` is an ordered list of ``(template, instance, cells)``; one
    worksheet is produced per section reusing :func:`_render_report_sheet`,
    preceded by a cover sheet that doubles as the audit summary (per-report
    status, cell count, and any generation warnings).
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    fd, name = tempfile.mkstemp(suffix=".xlsx", prefix="finauto_bundle_")
    import os as _os
    _os.close(fd)
    out = Path(name)

    wb = openpyxl.Workbook()
    cover = wb.active
    cover.title = "汇总"
    cover["A1"] = f"{org_name} · {period_id} · 财务报表汇总"
    cover["A1"].font = Font(bold=True, size=14)
    hdr = ["报表", "准则", "状态", "行数", "生成时间", "提示/警告"]
    for i, h in enumerate(hdr, start=1):
        c = cover.cell(row=3, column=i, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DDEBF7")
    r = 4
    used_titles: set[str] = set()
    for template, instance, cells in sections:
        base = _BUNDLE_SHEET_TITLES.get(instance.sheet_kind, instance.sheet_kind)
        sheet_title = base[:31]
        # Guard against duplicate sheet names (e.g. two of the same kind).
        suffix = 2
        while sheet_title in used_titles:
            sheet_title = f"{base[:28]}_{suffix}"
            suffix += 1
        used_titles.add(sheet_title)

        cover.cell(row=r, column=1, value=base)
        cover.cell(row=r, column=2, value=instance.accounting_standard)
        cover.cell(row=r, column=3, value=instance.status or "ok")
        cover.cell(row=r, column=4, value=instance.cell_count or len(cells))
        cover.cell(row=r, column=5, value=instance.generated_at or "")
        cover.cell(row=r, column=6, value="; ".join(instance.warnings or []) or "—")
        r += 1

        ws = wb.create_sheet(title=sheet_title)
        _render_report_sheet(ws, template, instance, cells)

    cover.column_dimensions["A"].width = 16
    cover.column_dimensions["B"].width = 18
    cover.column_dimensions["C"].width = 10
    cover.column_dimensions["D"].width = 8
    cover.column_dimensions["E"].width = 22
    cover.column_dimensions["F"].width = 50
    wb.save(str(out))
    wb.close()
    return out


# ---------------------------------------------------------------------------
# Endpoint registration
# ---------------------------------------------------------------------------


def register_report_endpoints(
    router: APIRouter, service: FinanceAutoService
) -> None:
    @router.get(
        "/orgs/{org_id}/report-bundle/export",
        summary="一键打包导出：BS/IS/CF（含汇总/审计）为单一多 sheet xlsx",
        response_class=FileResponse,
    )
    async def export_report_bundle(
        org_id: str,
        period_id: str = Query(..., description="结账期间，如 2025-FY"),
        format: str = Query(default="xlsx", pattern="^(xlsx)$"),
        _user: str = Depends(require_permission("report", "read")),
    ) -> FileResponse:
        org = await service.get_org(org_id)
        # Gather every report for (org, period); keep the most recent of
        # each kind so the bundle holds at most one BS / IS / CF sheet.
        async with service.db.conn.execute(
            "SELECT * FROM reports WHERE org_id=? AND period_id=? "
            "ORDER BY generated_at DESC",
            (org_id, period_id),
        ) as cur:
            rows = await cur.fetchall()
        latest_by_kind: dict[str, Any] = {}
        for row in rows:
            kind = row["sheet_kind"]
            if kind not in latest_by_kind:
                latest_by_kind[kind] = row
        if not latest_by_kind:
            raise HTTPException(
                status_code=404,
                detail=(
                    f"no generated reports for org={org_id} period={period_id}; "
                    "generate the statements first"
                ),
            )

        # Stable BS → IS → CF ordering, with any other kinds appended.
        kind_order = ["balance_sheet", "income_statement", "cash_flow"]
        ordered_kinds = [k for k in kind_order if k in latest_by_kind]
        ordered_kinds += [k for k in latest_by_kind if k not in kind_order]

        sections: list[tuple[LoadedTemplate, ReportInstance, list[ReportCell]]] = []
        for kind in ordered_kinds:
            row = latest_by_kind[kind]
            instance = _row_to_report(row)
            async with service.db.conn.execute(
                "SELECT * FROM report_cells WHERE report_id=? "
                "ORDER BY target_line_no ASC",
                (instance.id,),
            ) as cur:
                cell_rows = await cur.fetchall()
            cells = [_row_to_cell(c) for c in cell_rows]
            template = load_template(
                _resolve_template_path(
                    instance.sheet_kind, instance.accounting_standard
                )
            )
            sections.append((template, instance, cells))

        out_path = _build_bundle_workbook(
            org_name=org.name or org_id,
            period_id=period_id,
            sections=sections,
        )
        safe_org = (org.code or org_id)[-12:]
        filename = f"bundle_{safe_org}_{period_id}.xlsx"
        return FileResponse(
            str(out_path),
            media_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
            filename=filename,
        )

    @router.post(
        "/orgs/{org_id}/reports/{kind}/generate",
        status_code=201,
        summary="按 YAML 模板生成主表（小企业 / 企业准则资产负债表 + 利润表）",
    )
    async def generate(
        org_id: str, kind: str, payload: ReportGenerateRequest
    ) -> ReportDetailResponse:
        if kind not in {"balance_sheet", "income_statement", "cash_flow"}:
            raise HTTPException(
                status_code=400,
                detail=f"unsupported report kind: {kind!r}; allowed: "
                "balance_sheet | income_statement | cash_flow",
            )
        org = await service.get_org(org_id)
        standard = _standard_for_org(org.standard, payload.accounting_standard)
        template_path = _resolve_template_path(kind, standard)
        template = load_template(template_path)

        balance_lines, source_id = await _load_balance_lines(
            service,
            org_id=org_id,
            period_id=payload.period_id,
            source_import_id=payload.source_import_id,
        )
        manual_input_values: dict[str, float] = {}
        if kind == "cash_flow":
            # W3 Stage 4: pre-load every manual_input filled for this
            # (org, period) so the generator can substitute them by key.
            async with service.db.conn.execute(
                "SELECT field_key, value FROM manual_inputs WHERE org_id=? "
                "AND period_id=?",
                (org_id, payload.period_id),
            ) as cur:
                async for row in cur:
                    raw = (row["value"] or "").strip()
                    if not raw:
                        continue
                    try:
                        manual_input_values[row["field_key"]] = float(raw)
                    except (TypeError, ValueError):
                        continue
        generated = generate_report(
            template=template,
            org_id=org_id,
            period_id=payload.period_id,
            accounting_standard=standard,
            balance_lines=balance_lines,
            source_import_id=source_id,
            manual_input_values=manual_input_values,
        )
        await _persist_report(service, template=template, generated=generated)

        return ReportDetailResponse(
            report=generated.instance, cells=generated.cells
        )

    @router.get(
        "/orgs/{org_id}/reports",
        summary="列出某账套已生成的报表实例",
    )
    async def list_reports(org_id: str) -> ReportListResponse:
        await service.get_org(org_id)
        async with service.db.conn.execute(
            "SELECT * FROM reports WHERE org_id=? ORDER BY generated_at DESC",
            (org_id,),
        ) as cur:
            rows = await cur.fetchall()
        items = [_row_to_report(r) for r in rows]
        return ReportListResponse(reports=items, total=len(items))

    @router.get(
        "/orgs/{org_id}/reports/{report_id}",
        summary="读取一份报表（含全部 ReportCell）",
    )
    async def get_report(org_id: str, report_id: str) -> ReportDetailResponse:
        await service.get_org(org_id)
        async with service.db.conn.execute(
            "SELECT * FROM reports WHERE org_id=? AND id=?",
            (org_id, report_id),
        ) as cur:
            row = await cur.fetchone()
        if row is None:
            raise HTTPException(status_code=404, detail="report not found")
        instance = _row_to_report(row)
        async with service.db.conn.execute(
            "SELECT * FROM report_cells WHERE report_id=? "
            "ORDER BY target_line_no ASC",
            (report_id,),
        ) as cur:
            cell_rows = await cur.fetchall()
        cells = [_row_to_cell(r) for r in cell_rows]
        return ReportDetailResponse(report=instance, cells=cells)

    @router.patch(
        "/orgs/{org_id}/reports/{report_id}/cells/{cell_id}/simplify",
        summary="单 cell 切换 / 调整简化开关 (W3 Stage 2)",
    )
    async def patch_cell_simplify(
        org_id: str,
        report_id: str,
        cell_id: str,
        payload: CellSimplifyPatchRequest,
    ) -> ReportCell:
        await service.get_org(org_id)
        async with service.db.conn.execute(
            "SELECT * FROM reports WHERE org_id=? AND id=?",
            (org_id, report_id),
        ) as cur:
            r_row = await cur.fetchone()
        if r_row is None:
            raise HTTPException(status_code=404, detail="report not found")
        async with service.db.conn.execute(
            "SELECT * FROM report_cells WHERE id=? AND report_id=?",
            (cell_id, report_id),
        ) as cur:
            c_row = await cur.fetchone()
        if c_row is None:
            raise HTTPException(status_code=404, detail="cell not found")

        # Look the underlying account_filter + balance_kind back up from the
        # template so we can re-run the simplifier with the new config.
        instance = _row_to_report(r_row)
        template = load_template(
            _resolve_template_path(instance.sheet_kind, instance.accounting_standard)
        )
        rule = template.rule_by_code(c_row["reference_code"])
        if rule is None or rule.data_source != "account":
            raise HTTPException(
                status_code=400,
                detail=(
                    "cell does not back onto a simplifiable account rule; "
                    "only data_source=account rules support simplify toggle"
                ),
            )
        if not (rule.account_filter and rule.balance_kind):
            raise HTTPException(
                status_code=400,
                detail="rule lacks account_filter / balance_kind",
            )

        # Reload the source detail rows (W2 already encrypted/decrypted).
        balance_lines, _src = await _load_balance_lines(
            service,
            org_id=org_id,
            period_id=instance.period_id,
            source_import_id=instance.source_import_id,
        )
        from .report_generator import _filter_by_pattern  # local to avoid cycle
        matched = _filter_by_pattern(balance_lines, rule.account_filter)
        detail_rows = [
            DetailRow(
                row_id=ln.id,
                name=(ln.aux_text or ln.account_name or ln.full_code) or ln.full_code,
                amount=_balance_kind_to_amount(ln, rule.balance_kind),
                extra={"account_code": ln.full_code, "aux_text": ln.aux_text},
            )
            for ln in matched
        ]
        cfg = SimplifyConfig(
            enabled=payload.enabled,
            strategy=payload.strategy,
            top_n=payload.top_n,
            sort_by=payload.sort_by,
            merge_label=payload.merge_label,
            min_threshold=payload.min_threshold,
            keep_negative_separate=payload.keep_negative_separate,
            footnote_template=payload.footnote_template,
        )
        result = simplify_aux_details(detail_rows, cfg)
        merged_ids_json = (
            json.dumps(result.merged_row_ids, ensure_ascii=False)
            if result.merged_row_ids
            else None
        )
        cfg_json = json.dumps(cfg.to_dict(), ensure_ascii=False)
        new_simplified = bool(result.merged_count > 0)
        current_version = int(c_row["version"] or 1) if "version" in c_row.keys() else 1
        await service.db.conn.execute(
            "UPDATE report_cells SET simplified=?, simplified_top_n=?, "
            "simplify_config_json=?, merged_row_ids_json=?, footnote=?, "
            "version=version+1 WHERE id=? AND report_id=?",
            (
                int(new_simplified),
                cfg.top_n,
                cfg_json,
                merged_ids_json,
                result.footnote or None,
                cell_id,
                report_id,
            ),
        )
        await service.db.conn.commit()
        async with service.db.conn.execute(
            "SELECT * FROM report_cells WHERE id=?", (cell_id,),
        ) as cur:
            updated = await cur.fetchone()
        _ = current_version  # reserved for future optimistic-lock 409 path
        return _row_to_cell(updated)

    @router.get(
        "/orgs/{org_id}/reports/{report_id}/cells/{cell_id}/details",
        summary="展开报表 cell 的完整明细（含 \"其他\" 行的合并细节）",
    )
    async def get_cell_details(
        org_id: str, report_id: str, cell_id: str
    ) -> CellDetailsResponse:
        await service.get_org(org_id)
        async with service.db.conn.execute(
            "SELECT * FROM report_cells WHERE id=? AND report_id=?",
            (cell_id, report_id),
        ) as cur:
            c_row = await cur.fetchone()
        if c_row is None:
            raise HTTPException(status_code=404, detail="cell not found")
        cell = _row_to_cell(c_row)

        async with service.db.conn.execute(
            "SELECT org_id, period_id, source_import_id, sheet_kind, "
            "accounting_standard FROM reports WHERE id=?", (report_id,),
        ) as cur:
            r_row = await cur.fetchone()
        if r_row is None:
            raise HTTPException(status_code=404, detail="report not found")
        instance_period = r_row["period_id"]
        source_import = r_row["source_import_id"]

        template = load_template(
            _resolve_template_path(r_row["sheet_kind"], r_row["accounting_standard"])
        )
        rule = template.rule_by_code(cell.reference_code)
        full_rows: list[CellDetailRow] = []
        visible_rows: list[CellDetailRow] = []

        if rule and rule.data_source == "account" and rule.account_filter and rule.balance_kind:
            balance_lines, _src = await _load_balance_lines(
                service,
                org_id=org_id,
                period_id=instance_period,
                source_import_id=source_import,
            )
            from .report_generator import _filter_by_pattern
            matched = _filter_by_pattern(balance_lines, rule.account_filter)
            by_id = {ln.id: ln for ln in matched}
            for ln in matched:
                full_rows.append(CellDetailRow(
                    trial_balance_row_id=ln.id,
                    name=(ln.aux_text or ln.account_name or ln.full_code) or ln.full_code,
                    amount=_balance_kind_to_amount(ln, rule.balance_kind),
                    aux_text=ln.aux_text,
                    account_code=ln.full_code,
                ))
            if cell.simplified:
                merged_set = set(cell.merged_row_ids or [])
                for ln in matched:
                    if ln.id in merged_set:
                        continue
                    visible_rows.append(CellDetailRow(
                        trial_balance_row_id=ln.id,
                        name=(ln.aux_text or ln.account_name or ln.full_code) or ln.full_code,
                        amount=_balance_kind_to_amount(ln, rule.balance_kind),
                        aux_text=ln.aux_text,
                        account_code=ln.full_code,
                    ))
                if merged_set:
                    merged_amount = sum(
                        _balance_kind_to_amount(by_id[mid], rule.balance_kind)
                        for mid in merged_set
                        if mid in by_id
                    )
                    visible_rows.append(CellDetailRow(
                        trial_balance_row_id=None,
                        name=(cell.simplify_config or {}).get("merge_label", "其他"),
                        amount=round(merged_amount, 2),
                        is_merged=True,
                        merged_count=len(merged_set),
                        merged_row_ids=list(merged_set),
                    ))
            else:
                visible_rows = list(full_rows)

        return CellDetailsResponse(
            report_id=report_id,
            cell_id=cell_id,
            reference_code=cell.reference_code,
            target_label=cell.target_label,
            simplified=cell.simplified,
            simplify_config=cell.simplify_config,
            visible_rows=visible_rows,
            full_rows=full_rows,
            footnote=cell.footnote,
        )

    @router.get(
        "/orgs/{org_id}/reports/{report_id}/export",
        summary="导出报表为 Excel (.xlsx)",
        response_class=FileResponse,
    )
    async def export_report(
        org_id: str,
        report_id: str,
        format: str = Query(default="xlsx", pattern="^(xlsx)$"),
    ) -> FileResponse:
        await service.get_org(org_id)
        async with service.db.conn.execute(
            "SELECT * FROM reports WHERE org_id=? AND id=?",
            (org_id, report_id),
        ) as cur:
            row = await cur.fetchone()
        if row is None:
            raise HTTPException(status_code=404, detail="report not found")
        instance = _row_to_report(row)
        async with service.db.conn.execute(
            "SELECT * FROM report_cells WHERE report_id=? "
            "ORDER BY target_line_no ASC",
            (report_id,),
        ) as cur:
            cell_rows = await cur.fetchall()
        cells = [_row_to_cell(r) for r in cell_rows]

        template_path = _resolve_template_path(
            instance.sheet_kind, instance.accounting_standard
        )
        template = load_template(template_path)
        out_path = _build_workbook(template, instance, cells)

        await service.db.conn.execute(
            "UPDATE reports SET output_path=?, backend_used=? WHERE id=?",
            (str(out_path), "openpyxl", report_id),
        )
        await service.db.conn.commit()

        filename = (
            f"{instance.template_id}_{instance.period_id}_"
            f"{instance.id[-8:]}.xlsx"
        )
        return FileResponse(
            str(out_path),
            media_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
            filename=filename,
        )
