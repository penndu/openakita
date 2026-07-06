"""openpyxl-backed renderer for dynamic detail reports.

Selected by the factory when ``rows_estimate`` exceeds
:data:`STATIC_ROW_THRESHOLD` or when the report kind is on the dynamic-only
list.  The renderer copies the styling of the *first* data row in the
template (font / alignment / borders / number_format / fill) and replicates
it onto every appended row, so designers can lay out a single sample row in
the template and trust the renderer to preserve the look across N rows.
"""

from __future__ import annotations

import time
from copy import copy
from pathlib import Path
from typing import Any

from .base import RenderResult, ReportRenderer

STATIC_ROW_THRESHOLD = 50
"""Reports with up to this many rows are routed to xltpl by default."""


class OpenpyxlDirectRenderer(ReportRenderer):
    """Append rows after a marker, copying the marker row's style.

    Template conventions:

    * Header / title cells use plain text (no Jinja).  Header substitutions
      (``{{ org.name }}`` etc.) are processed as a string ``.replace()`` pass
      after the row append, so designers can use the same placeholder syntax
      as the xltpl track.
    * The marker row contains, in column A, the literal string
      ``__ROWS_HERE__``.  All rows after the marker are dynamic data; the
      marker row's style is the per-row template.
    * Optional total row: any row containing ``__SUM__`` is rewritten with a
      ``=SUM(<first-data-row>:<last-data-row>)`` formula in the same column.
    """

    BACKEND = "openpyxl"

    MARKER = "__ROWS_HERE__"
    SUM_MARKER = "__SUM__"

    def _render(self, context: dict[str, Any], output_path: Path) -> RenderResult:
        import openpyxl

        rows = list(context.get("rows") or [])
        warnings: list[str] = []

        started = time.perf_counter()
        wb = openpyxl.load_workbook(str(self._template_path))
        ws = wb.active

        marker_row = self._find_marker_row(ws)
        if marker_row is None:
            wb.close()
            raise ValueError(
                f"openpyxl template missing '{self.MARKER}' anchor: "
                f"{self._template_path}"
            )

        column_keys = self._collect_column_keys(ws, marker_row)
        if not column_keys:
            warnings.append(
                "no column keys found on marker row; rows written verbatim"
            )

        style_template = [
            (
                cell.font,
                cell.alignment,
                cell.border,
                cell.fill,
                cell.number_format,
                cell.protection,
            )
            for cell in ws[marker_row]
        ]

        # Capture the trailing rows (e.g. ``__SUM__`` totals) that live
        # *after* the marker row.  We need to push them past the last
        # dynamic data row.  Each entry is a list of (col_idx, value, style
        # tuple) so we can re-emit them.
        max_col = ws.max_column
        trailing_rows: list[list[tuple[int, Any, tuple]]] = []
        if ws.max_row > marker_row:
            for src_row in range(marker_row + 1, ws.max_row + 1):
                snapshot: list[tuple[int, Any, tuple]] = []
                for col_idx in range(1, max_col + 1):
                    src_cell = ws.cell(row=src_row, column=col_idx)
                    snapshot.append(
                        (
                            col_idx,
                            src_cell.value,
                            (
                                copy(src_cell.font),
                                copy(src_cell.alignment),
                                copy(src_cell.border),
                                copy(src_cell.fill),
                                src_cell.number_format,
                                copy(src_cell.protection),
                            ),
                        )
                    )
                trailing_rows.append(snapshot)
            for src_row in range(marker_row + 1, ws.max_row + 1):
                for col_idx in range(1, max_col + 1):
                    cell = ws.cell(row=src_row, column=col_idx)
                    cell.value = None

        # Wipe the marker text from the marker row but keep it as a styling
        # template for the first dynamic row.
        for cell in ws[marker_row]:
            if isinstance(cell.value, str) and self.MARKER in cell.value:
                cell.value = None

        first_data_row = marker_row
        for offset, row_data in enumerate(rows):
            target = first_data_row + offset
            for col_idx, key in enumerate(column_keys, start=1):
                if key is None:
                    continue
                value = row_data.get(key)
                cell = ws.cell(row=target, column=col_idx, value=value)
                if col_idx - 1 < len(style_template):
                    font, align, border, fill, fmt, prot = style_template[
                        col_idx - 1
                    ]
                    cell.font = copy(font)
                    cell.alignment = copy(align)
                    cell.border = copy(border)
                    cell.fill = copy(fill)
                    cell.number_format = fmt
                    cell.protection = copy(prot)

        last_data_row = first_data_row + max(len(rows) - 1, 0)

        for offset, snapshot in enumerate(trailing_rows, start=1):
            target_row = last_data_row + offset
            for col_idx, value, styles in snapshot:
                cell = ws.cell(row=target_row, column=col_idx, value=value)
                font, align, border, fill, fmt, prot = styles
                cell.font = font
                cell.alignment = align
                cell.border = border
                cell.fill = fill
                cell.number_format = fmt
                cell.protection = prot

        # Header substitutions on every cell above the data block.
        org = context.get("org", {}) or {}
        substitutions = {
            "{{ org.name }}": str(org.get("name", "")),
            "{{ year }}": str(context.get("year", "")),
            "{{ period }}": str(context.get("period", "")),
            "{{ standard }}": str(context.get("standard", "")),
        }
        self._apply_header_substitutions(ws, marker_row, substitutions)
        self._rewrite_sum_markers(ws, first_data_row, last_data_row)

        wb.save(str(output_path))
        wb.close()
        elapsed_ms = (time.perf_counter() - started) * 1000.0

        return RenderResult(
            output_path=output_path,
            rows_written=len(rows),
            elapsed_ms=elapsed_ms,
            backend=self.BACKEND,
            warnings=warnings,
        )

    def _find_marker_row(self, ws: Any) -> int | None:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and self.MARKER in cell.value:
                    return cell.row
        return None

    def _collect_column_keys(self, ws: Any, marker_row: int) -> list[str | None]:
        # Convention: the row immediately above the marker holds the column
        # keys (e.g. "account_code", "name", "balance").  Header labels live
        # one row higher.
        if marker_row <= 1:
            return []
        keys: list[str | None] = []
        for cell in ws[marker_row - 1]:
            if isinstance(cell.value, str) and cell.value.strip():
                keys.append(cell.value.strip())
            else:
                keys.append(None)
        return keys

    def _apply_header_substitutions(
        self, ws: Any, marker_row: int, substitutions: dict[str, str]
    ) -> None:
        if marker_row <= 1:
            return
        for row in ws.iter_rows(min_row=1, max_row=marker_row - 1):
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                text = cell.value
                for key, val in substitutions.items():
                    if key in text:
                        text = text.replace(key, val)
                if text != cell.value:
                    cell.value = text

    def _rewrite_sum_markers(self, ws: Any, first: int, last: int) -> None:
        if last < first:
            return
        for row in ws.iter_rows(min_row=last + 1, max_row=ws.max_row):
            for cell in row:
                if isinstance(cell.value, str) and self.SUM_MARKER in cell.value:
                    col_letter = cell.column_letter
                    cell.value = f"=SUM({col_letter}{first}:{col_letter}{last})"


# ---------------------------------------------------------------------------
# M2 Biz Stage 5 — template-free dynamic detail writer.
#
# OpenpyxlDirectRenderer above wants a pre-laid-out Excel template with a
# ``__ROWS_HERE__`` marker.  For account / aux-vendor / cash-account detail
# sheets (Part Biz §3.5, 货币资金 / 应收 / 应付 / 辅助核算 明细) we don't have
# a template per detail kind — we want to slot rows into an existing workbook
# in a brand-new sheet with explicit styles.  :class:`OpenpyxlDirectWriter`
# fills that gap.
#
# Design contract (Part Infra C2):
#   * pure-function-on-workbook entry point ``write_detail_rows``
#   * three style buckets (header / row / total) with sensible defaults
#   * row count up to 1500; benchmarked < 5 seconds end-to-end
#   * optional simplifier integration: caller passes ``SimplifyConfig`` and a
#     pre-aggregated DetailRow list and we render the simplified output
# ---------------------------------------------------------------------------


from dataclasses import dataclass, field
from typing import Iterable, Sequence

from .simplifier import DetailRow, SimplifyConfig, SimplifyResult, simplify_aux_details


@dataclass
class DetailStyle:
    """Style spec for one of the three row buckets.

    All optional; ``None`` means "leave openpyxl defaults alone".  Callers
    typically reuse a single :class:`DetailStyle` instance across many sheets
    so the look is consistent.
    """

    font_name: str | None = "Microsoft YaHei"
    font_size: float | None = 11.0
    bold: bool = False
    italic: bool = False
    fill_color: str | None = None        # 'FFEFEFEF' etc; None = no fill
    horizontal: str | None = None        # 'center'|'left'|'right' (defaults to left for text, right for numbers)
    vertical: str = "center"
    number_format: str | None = None     # for numeric cells; e.g. '#,##0.00'
    border: bool = True
    wrap_text: bool = False

    def is_default(self) -> bool:
        return (
            self.fill_color is None
            and self.font_name == "Microsoft YaHei"
            and self.font_size == 11.0
            and not self.bold
            and not self.italic
            and not self.wrap_text
            and self.number_format is None
            and self.border
        )


def _default_header_style() -> DetailStyle:
    return DetailStyle(
        bold=True, fill_color="FFD9E1F2", horizontal="center",
        border=True,
    )


def _default_row_style() -> DetailStyle:
    return DetailStyle(number_format="#,##0.00")


def _default_total_style() -> DetailStyle:
    return DetailStyle(
        bold=True, fill_color="FFFFF2CC", number_format="#,##0.00", border=True,
    )


@dataclass
class WriteDetailResult:
    """Returned by :meth:`OpenpyxlDirectWriter.write_detail_rows`."""

    sheet_name: str
    row_count: int
    elapsed_ms: float
    simplify_applied: bool = False
    grouped_into: int | None = None
    warnings: list[str] = field(default_factory=list)


class OpenpyxlDirectWriter:
    """Template-free dynamic-detail row writer.

    Compared with :class:`OpenpyxlDirectRenderer`, this writer does not need
    an Excel template at all — caller passes a live ``openpyxl.Workbook``,
    target sheet name, header definitions and the row data.  We append rows
    and apply explicit per-bucket styles.

    Typical pipeline:

    .. code-block:: python

        wb = openpyxl.Workbook()
        writer = OpenpyxlDirectWriter()
        writer.write_detail_rows(
            workbook=wb, sheet_name='货币资金明细',
            columns=[
                {'key': 'bank_name', 'label': '银行/账户', 'width': 28},
                {'key': 'currency',  'label': '币种',     'width': 8},
                {'key': 'closing',   'label': '期末余额', 'width': 18, 'numeric': True},
            ],
            rows=[{'bank_name': '工行 ...', 'currency': 'CNY', 'closing': 12345.67}, ...],
            total_label='合计',
            total_columns=['closing'],
        )
    """

    MAX_ROWS = 1500
    """Soft cap recommended by Part Biz §3.5.  Callers exceeding this should
    pre-aggregate via :class:`SimplifyConfig` rather than dump more rows."""

    def write_detail_rows(
        self,
        *,
        workbook: Any,
        sheet_name: str,
        columns: Sequence[dict[str, Any]],
        rows: Iterable[dict[str, Any]],
        header_style: DetailStyle | None = None,
        row_style: DetailStyle | None = None,
        total_style: DetailStyle | None = None,
        total_label: str = "合计",
        total_columns: Sequence[str] | None = None,
        simplify_config: SimplifyConfig | None = None,
    ) -> WriteDetailResult:
        """Write a dynamic-detail block onto ``workbook``.

        ``simplify_config`` (M1 W3 simplifier integration): when supplied,
        rows are first run through ``simplify_aux_details`` so very-long
        vendor lists collapse into the configured top-N + "其他".  We pass
        the simplifier the columns marked as ``numeric=True`` (default: the
        last numeric column) as the rank field.
        """
        import time as _time

        try:
            import openpyxl  # noqa: F401  (verifies the dep is installed)
            from openpyxl.styles import (
                Alignment, Border, Font, PatternFill, Side,
            )
        except ImportError as exc:  # pragma: no cover
            raise RuntimeError(
                "openpyxl is required for OpenpyxlDirectWriter"
            ) from exc

        started = _time.perf_counter()
        rows_list = list(rows)
        warnings: list[str] = []
        simplify_applied = False
        grouped_into: int | None = None

        # Optional simplifier pre-step.
        if simplify_config is not None and simplify_config.enabled and rows_list:
            rank_key = self._pick_rank_key(columns)
            name_key = columns[0]["key"] if columns else "name"
            detail_rows = [
                DetailRow(
                    row_id=str(r.get("row_id") or i),
                    name=str(r.get(name_key) or r.get("name") or ""),
                    amount=float(r.get(rank_key, 0) or 0),
                    extra=r,
                )
                for i, r in enumerate(rows_list)
            ]
            res: SimplifyResult = simplify_aux_details(detail_rows, simplify_config)
            new_rows: list[dict] = []
            for kr in res.kept_rows:
                if kr.extra.get("is_merged"):
                    # Build a synthetic row dict honouring every column key.
                    merged = {c["key"]: "" for c in columns}
                    if columns:
                        merged[columns[0]["key"]] = kr.name
                        merged[rank_key] = kr.amount
                    new_rows.append(merged)
                else:
                    new_rows.append(kr.extra)
            rows_list = new_rows
            simplify_applied = True
            grouped_into = len(rows_list)

        if len(rows_list) > self.MAX_ROWS:
            warnings.append(
                f"row count {len(rows_list)} exceeds MAX_ROWS={self.MAX_ROWS}; "
                "truncated to keep workbook performance acceptable"
            )
            rows_list = rows_list[: self.MAX_ROWS]

        h_style = header_style or _default_header_style()
        r_style = row_style or _default_row_style()
        t_style = total_style or _default_total_style()

        if sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
        else:
            ws = workbook.create_sheet(sheet_name)

        # ---- compile re-usable openpyxl objects ----
        def _font(s: DetailStyle) -> Font:
            return Font(
                name=s.font_name, size=s.font_size, bold=s.bold, italic=s.italic
            )

        def _fill(s: DetailStyle) -> PatternFill | None:
            if not s.fill_color:
                return None
            return PatternFill(
                start_color=s.fill_color, end_color=s.fill_color, fill_type="solid"
            )

        def _border(s: DetailStyle) -> Border | None:
            if not s.border:
                return None
            side = Side(border_style="thin", color="FFB4B4B4")
            return Border(left=side, right=side, top=side, bottom=side)

        def _align(s: DetailStyle, *, numeric: bool) -> Alignment:
            horiz = s.horizontal or ("right" if numeric else "left")
            return Alignment(
                horizontal=horiz, vertical=s.vertical, wrap_text=s.wrap_text
            )

        # ---- header row ----
        header_row_idx = ws.max_row + 1 if ws.max_row > 1 or any(
            ws.cell(row=1, column=c + 1).value for c in range(len(columns))
        ) else 1
        h_font, h_fill, h_border = _font(h_style), _fill(h_style), _border(h_style)
        for col_idx, col in enumerate(columns, start=1):
            cell = ws.cell(row=header_row_idx, column=col_idx, value=col.get("label", col["key"]))
            cell.font = h_font
            cell.alignment = _align(h_style, numeric=False)
            if h_fill is not None:
                cell.fill = h_fill
            if h_border is not None:
                cell.border = h_border
            if "width" in col:
                ws.column_dimensions[cell.column_letter].width = float(col["width"])

        # ---- data rows ----
        r_font, r_fill, r_border = _font(r_style), _fill(r_style), _border(r_style)
        first_data_row = header_row_idx + 1
        for ridx, row_data in enumerate(rows_list, start=first_data_row):
            for col_idx, col in enumerate(columns, start=1):
                value = row_data.get(col["key"])
                cell = ws.cell(row=ridx, column=col_idx, value=value)
                cell.font = r_font
                numeric = bool(col.get("numeric"))
                cell.alignment = _align(r_style, numeric=numeric)
                if numeric and (r_style.number_format or col.get("number_format")):
                    cell.number_format = col.get("number_format") or r_style.number_format
                if r_fill is not None:
                    cell.fill = r_fill
                if r_border is not None:
                    cell.border = r_border
        last_data_row = first_data_row + max(len(rows_list) - 1, 0)

        # ---- total row ----
        if total_columns and rows_list:
            t_row = last_data_row + 1
            t_font, t_fill, t_border = _font(t_style), _fill(t_style), _border(t_style)
            label_col = 1
            for col_idx, col in enumerate(columns, start=1):
                cell = ws.cell(row=t_row, column=col_idx)
                cell.font = t_font
                cell.alignment = _align(t_style, numeric=bool(col.get("numeric")))
                if t_fill is not None:
                    cell.fill = t_fill
                if t_border is not None:
                    cell.border = t_border
                if col_idx == label_col:
                    cell.value = total_label
                elif col["key"] in total_columns:
                    col_letter = cell.column_letter
                    cell.value = f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})"
                    cell.number_format = t_style.number_format or "#,##0.00"
            written = len(rows_list) + 1
        else:
            written = len(rows_list)

        elapsed_ms = (_time.perf_counter() - started) * 1000.0
        return WriteDetailResult(
            sheet_name=sheet_name,
            row_count=written,
            elapsed_ms=elapsed_ms,
            simplify_applied=simplify_applied,
            grouped_into=grouped_into,
            warnings=warnings,
        )

    @staticmethod
    def _pick_rank_key(columns: Sequence[dict[str, Any]]) -> str:
        """Pick the numeric column the simplifier ranks rows by."""
        for col in reversed(list(columns)):
            if col.get("numeric"):
                return col["key"]
        # No numeric column → rank by 'amount' if present, else the last key.
        if any(c["key"] == "amount" for c in columns):
            return "amount"
        return columns[-1]["key"] if columns else "amount"


__all__ = [
    "DetailStyle",
    "OpenpyxlDirectRenderer",
    "OpenpyxlDirectWriter",
    "STATIC_ROW_THRESHOLD",
    "WriteDetailResult",
]
