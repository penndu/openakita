"""Trial-balance parser (.xls / .xlsx) — three-tier fallback chain.

Strategy (mirrors ``tmp_spike/02_xls_parse`` summary, validated on 3 samples):

1. ``.xlsx``  → ``openpyxl``  (read-only, data_only)
2. ``.xls``   → ``xlrd 1.2``  (the last version that supports legacy BIFF)
3. ``.xls``   → ``pywin32`` + Excel COM (auto-convert to xlsx then openpyxl)

Each tier returns a tuple ``(rows, parser_used)`` where ``rows`` is a list of
:class:`ParsedRow` dataclasses ready for persistence.  Header detection is
heuristic — see :func:`_detect_header_layout` — and tolerant to:

* sheets where the column header sits at row 4 with a two-row sub-header
  for ``借/贷`` (Sample A pattern: ``期初余额`` over ``借`` / ``贷``);
* sheets without sub-headers (Sample B pattern: each column has a single label
  like ``期初借方`` / ``期初贷方``);
* sheets with a couple of extra free-text rows above the data (Sample C).
"""

from __future__ import annotations

import logging
import re
import shutil
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from ..normalizers.account_code import join_full_code, normalize_account_code

logger = logging.getLogger(__name__)


@dataclass
class ParsedRow:
    row_index: int
    raw_code: str
    parent_code: str
    child_code: str | None
    full_code: str
    account_name: str | None
    aux_text: str | None = None
    opening_debit: float = 0.0
    opening_credit: float = 0.0
    period_debit: float = 0.0
    period_credit: float = 0.0
    closing_debit: float = 0.0
    closing_credit: float = 0.0


@dataclass
class ParseResult:
    rows: list[ParsedRow] = field(default_factory=list)
    parser_used: str = ""
    sheet_name: str = ""
    skipped: int = 0


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------


def parse_trial_balance(path: Path | str) -> ParseResult:
    """Parse an Excel trial-balance file via the three-tier fallback chain.

    Raises :class:`ValueError` if none of the tiers can read the file.
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"trial-balance file does not exist: {p}")

    suffix = p.suffix.lower()
    errors: list[str] = []

    # Tier 1 — openpyxl for .xlsx / .xlsm
    if suffix in {".xlsx", ".xlsm"}:
        try:
            return _parse_xlsx_openpyxl(p)
        except Exception as exc:
            errors.append(f"openpyxl: {type(exc).__name__}: {exc}")
            logger.warning("finance-auto: openpyxl parse failed for %s: %s", p.name, exc)

    # Tier 2 — xlrd 1.2 for .xls
    if suffix == ".xls":
        try:
            return _parse_xls_xlrd(p)
        except Exception as exc:
            errors.append(f"xlrd: {type(exc).__name__}: {exc}")
            logger.warning("finance-auto: xlrd parse failed for %s: %s", p.name, exc)

    # Tier 3 — pywin32 COM (only on Windows; converts .xls → temp .xlsx)
    if suffix == ".xls":
        try:
            with _convert_xls_to_xlsx_via_com(p) as tmp_xlsx:
                result = _parse_xlsx_openpyxl(tmp_xlsx)
                result.parser_used = "pywin32-com+openpyxl"
                return result
        except Exception as exc:
            errors.append(f"pywin32-com: {type(exc).__name__}: {exc}")
            logger.warning("finance-auto: pywin32-com parse failed for %s: %s", p.name, exc)

    raise ValueError(f"all parser tiers failed for {p.name}: {'; '.join(errors)}")


# ---------------------------------------------------------------------------
# Tier 1 — openpyxl
# ---------------------------------------------------------------------------


def _parse_xlsx_openpyxl(path: Path) -> ParseResult:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = _pick_balance_sheet_openpyxl(wb)
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        wb.close()

    layout = _detect_header_layout(rows)
    parsed_rows = _build_rows(rows, layout)
    return ParseResult(
        rows=parsed_rows,
        parser_used="openpyxl",
        sheet_name=sheet.title,
        skipped=layout.data_start,
    )


def _pick_balance_sheet_openpyxl(wb: Any) -> Any:
    """Pick the most likely 余额表 sheet — defaults to the first sheet."""
    preferred_keywords = ("余额", "balance", "trial")
    for name in wb.sheetnames:
        low = name.lower()
        if any(kw in low for kw in preferred_keywords):
            return wb[name]
    return wb[wb.sheetnames[0]]


# ---------------------------------------------------------------------------
# Tier 2 — xlrd 1.2
# ---------------------------------------------------------------------------


def _parse_xls_xlrd(path: Path) -> ParseResult:
    import xlrd

    book = xlrd.open_workbook(str(path))
    sheet = _pick_balance_sheet_xlrd(book)
    rows = [
        tuple(sheet.cell_value(r, c) for c in range(sheet.ncols))
        for r in range(sheet.nrows)
    ]
    layout = _detect_header_layout(rows)
    parsed_rows = _build_rows(rows, layout)
    return ParseResult(
        rows=parsed_rows,
        parser_used="xlrd",
        sheet_name=sheet.name,
        skipped=layout.data_start,
    )


def _pick_balance_sheet_xlrd(book: Any) -> Any:
    preferred_keywords = ("余额", "balance", "trial")
    for s in book.sheets():
        low = s.name.lower()
        if any(kw in low for kw in preferred_keywords):
            return s
    return book.sheet_by_index(0)


# ---------------------------------------------------------------------------
# Tier 3 — pywin32 COM conversion to xlsx (Windows only)
# ---------------------------------------------------------------------------


class _ComConvertedXlsx:
    """Context manager that yields a temp .xlsx and deletes it on exit."""

    def __init__(self, src_xls: Path):
        self._src = src_xls
        self._tmp: Path | None = None

    def __enter__(self) -> Path:
        import os
        import tempfile

        import pythoncom
        import win32com.client

        pythoncom.CoInitialize()
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            wb = excel.Workbooks.Open(str(self._src), ReadOnly=True)
            fd, out_path = tempfile.mkstemp(suffix=".xlsx", prefix="finauto_xls_")
            os.close(fd)
            try:
                Path(out_path).unlink()
            except OSError:
                pass
            wb.SaveAs(out_path, FileFormat=51)  # 51 = xlOpenXMLWorkbook
            wb.Close(SaveChanges=False)
            excel.Quit()
            self._tmp = Path(out_path)
            return self._tmp
        finally:
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass

    def __exit__(self, exc_type, exc, tb) -> None:
        if self._tmp and self._tmp.exists():
            try:
                self._tmp.unlink()
            except OSError as exc:
                logger.debug("finance-auto: failed to remove temp xlsx %s: %s", self._tmp, exc)


def _convert_xls_to_xlsx_via_com(src: Path) -> _ComConvertedXlsx:
    return _ComConvertedXlsx(src)


# ---------------------------------------------------------------------------
# Header detection & row extraction (shared across tiers)
# ---------------------------------------------------------------------------


@dataclass
class _HeaderLayout:
    """Resolved column → semantic-name mapping.

    ``data_start`` is the first row index that holds real data (0-based).
    ``columns`` keys are the canonical names used by :class:`ParsedRow`.
    Missing columns are recorded as ``-1`` so callers can skip them safely.
    """

    data_start: int
    code_col: int = -1
    name_col: int = -1
    aux_col: int = -1
    opening_debit_col: int = -1
    opening_credit_col: int = -1
    period_debit_col: int = -1
    period_credit_col: int = -1
    closing_debit_col: int = -1
    closing_credit_col: int = -1


_CODE_PATS = ("科目编码", "科目代码", "编码", "code")
_NAME_PATS = ("科目名称", "科目", "name", "account")
_AUX_PATS = ("辅助核算", "辅助项", "辅助")
_OPENING_PATS = ("期初",)
_PERIOD_PATS = ("本期", "本年发生", "发生额")
_CLOSING_PATS = ("期末",)
_DEBIT_PATS = ("借",)
_CREDIT_PATS = ("贷",)


def _to_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        if v != v:  # NaN
            return ""
        return str(v)
    return str(v).strip()


def _row_has_keyword(row: tuple[Any, ...], keywords: tuple[str, ...]) -> bool:
    joined = " ".join(_to_str(c) for c in row)
    return any(kw in joined for kw in keywords)


def _detect_header_layout(rows: list[tuple[Any, ...]]) -> _HeaderLayout:
    """Find the header band (1 or 2 rows) and map columns to semantic names.

    Algorithm:

    1. Find the first row that contains both "编码" and either "期初" or "期末"
       — this is the main header band.
    2. If the next row contains ``借`` / ``贷`` cells, treat the band as
       two rows (Sample A pattern); merge top/bottom labels per column.
    3. Otherwise treat the band as a single row (Sample B pattern).
    4. Apply column patterns to the merged labels.
    """
    if not rows:
        return _HeaderLayout(data_start=0)

    header_idx = -1
    for idx, row in enumerate(rows):
        text = " ".join(_to_str(c) for c in row)
        if (
            any(kw in text for kw in _CODE_PATS)
            and any(kw in text for kw in (*_OPENING_PATS, *_CLOSING_PATS))
        ):
            header_idx = idx
            break

    if header_idx < 0:
        return _HeaderLayout(data_start=0)

    top_row = rows[header_idx]
    sub_row = rows[header_idx + 1] if header_idx + 1 < len(rows) else tuple()
    sub_is_debit_credit = _row_has_keyword(sub_row, _DEBIT_PATS) and _row_has_keyword(
        sub_row, _CREDIT_PATS
    )

    # Build merged labels per column (forward-fill empty top labels to handle
    # merged-cell headers like "期初余额" spanning two columns above 借/贷).
    width = max(len(top_row), len(sub_row))
    top_labels: list[str] = []
    last = ""
    for i in range(width):
        c = _to_str(top_row[i]) if i < len(top_row) else ""
        if c:
            last = c
        top_labels.append(last)
    sub_labels: list[str] = [
        _to_str(sub_row[i]) if i < len(sub_row) else "" for i in range(width)
    ]

    layout = _HeaderLayout(data_start=header_idx + (2 if sub_is_debit_credit else 1))

    for col in range(width):
        top = top_labels[col]
        sub = sub_labels[col] if sub_is_debit_credit else ""
        merged = (top + sub).strip()
        if not merged:
            continue

        if any(kw in merged for kw in _CODE_PATS):
            if layout.code_col < 0:
                layout.code_col = col
            continue
        if any(kw in merged for kw in _NAME_PATS):
            if layout.name_col < 0:
                layout.name_col = col
            continue
        if any(kw in merged for kw in _AUX_PATS):
            if layout.aux_col < 0:
                layout.aux_col = col
            continue

        has_debit = any(kw in merged for kw in _DEBIT_PATS)
        has_credit = any(kw in merged for kw in _CREDIT_PATS)

        if any(kw in merged for kw in _OPENING_PATS):
            if has_debit and layout.opening_debit_col < 0:
                layout.opening_debit_col = col
            elif has_credit and layout.opening_credit_col < 0:
                layout.opening_credit_col = col
        elif any(kw in merged for kw in _PERIOD_PATS):
            if has_debit and layout.period_debit_col < 0:
                layout.period_debit_col = col
            elif has_credit and layout.period_credit_col < 0:
                layout.period_credit_col = col
        elif any(kw in merged for kw in _CLOSING_PATS):
            if has_debit and layout.closing_debit_col < 0:
                layout.closing_debit_col = col
            elif has_credit and layout.closing_credit_col < 0:
                layout.closing_credit_col = col

    return layout


def _coerce_amount(cell: Any) -> float:
    if cell is None:
        return 0.0
    if isinstance(cell, (int, float)):
        if cell != cell:
            return 0.0
        return float(cell)
    s = str(cell).strip()
    if not s:
        return 0.0
    # Strip thousands separator, currency markers, and parenthesised negatives.
    s = s.replace(",", "").replace("¥", "").replace("￥", "").replace("$", "")
    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg = True
        s = s[1:-1]
    try:
        v = float(s)
    except (ValueError, TypeError):
        return 0.0
    return -v if neg else v


_CODE_LIKE = re.compile(r"^[\d\.\s]+$")


def _looks_like_account_code(s: str) -> bool:
    s = s.strip()
    if not s:
        return False
    # Pure digits (with optional dot) — never a label / summary row.
    return bool(_CODE_LIKE.match(s)) and any(ch.isdigit() for ch in s)


def _build_rows(rows: list[tuple[Any, ...]], layout: _HeaderLayout) -> list[ParsedRow]:
    if layout.code_col < 0:
        return []

    parsed: list[ParsedRow] = []
    for r_idx in range(layout.data_start, len(rows)):
        row = rows[r_idx]
        if not row:
            continue
        raw_code = _to_str(row[layout.code_col]) if layout.code_col < len(row) else ""
        if not raw_code or not _looks_like_account_code(raw_code):
            continue

        parent, child = normalize_account_code(raw_code)
        if not parent:
            continue
        full_code = join_full_code(parent, child)

        def _cell(col: int) -> Any:
            return row[col] if 0 <= col < len(row) else None

        parsed.append(
            ParsedRow(
                row_index=r_idx,
                raw_code=raw_code,
                parent_code=parent,
                child_code=child,
                full_code=full_code,
                account_name=(_to_str(_cell(layout.name_col)).strip() or None),
                aux_text=(_to_str(_cell(layout.aux_col)).strip() or None),
                opening_debit=_coerce_amount(_cell(layout.opening_debit_col)),
                opening_credit=_coerce_amount(_cell(layout.opening_credit_col)),
                period_debit=_coerce_amount(_cell(layout.period_debit_col)),
                period_credit=_coerce_amount(_cell(layout.period_credit_col)),
                closing_debit=_coerce_amount(_cell(layout.closing_debit_col)),
                closing_credit=_coerce_amount(_cell(layout.closing_credit_col)),
            )
        )
    return parsed


__all__ = ["ParseResult", "ParsedRow", "parse_trial_balance"]


# Quiet ruff's "shutil unused" warning — we keep the import close-by for the
# COM tier's tempfile cleanup helpers in case future revisions need rmtree.
_ = shutil
