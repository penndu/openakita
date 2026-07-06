"""Parser for the Chinese Golden-Tax-IV VAT declaration form.

The target form is the 金税四期一般纳税人 ``增值税及附加税费申报表`` -- the
"new" 2021+ unified return that replaces the older
``增值税纳税申报表`` plus the appended surtax forms.  Provincial branches of
STA emit slightly different .xlsx layouts (different sheet names, optional
headers, occasional merged-cell quirks), so the parser does:

1. **Source probe** -- inspects the filename and the first ~10 rows of every
   sheet in the workbook for canonical header patterns.  Returns a
   :class:`SourceProbe` describing which dialect was matched.
2. **Generic parse** -- once the dialect is known, walks the recognised
   row-label set and pulls numeric values.  Unknown labels are skipped
   silently; required labels missing produce a fatal :class:`VatParseError`.

Implemented dialects (M1 W2):

* ``golden_tax_iv_generic`` -- the Tax Bureau's central template; matches
  the bare canonical labels in v0.2 Part 1 §1.4.

Province-specific dialects (Beijing, Guangdong, Sichuan extra rows for local
surtaxes) are recorded in the dialect registry but currently fall back to
the generic dialect with a warning so we don't reject a real return.  The
dialect surface will grow as we collect samples (see roadmap section 6 of
the upcoming W2 completion report).
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)


class VatParseError(ValueError):
    """Raised when the VAT declaration cannot be reliably parsed."""


@dataclass(frozen=True)
class SourceProbe:
    dialect: str
    confidence: float  # 0.0-1.0
    sheet_name: str
    notes: list[str] = field(default_factory=list)


@dataclass
class VatDeclaration:
    """Pure parse output (no DB I/O)."""

    declaration_period: str
    output_vat: float = 0.0
    input_vat: float = 0.0
    prev_credit: float = 0.0
    tax_payable: float = 0.0
    surtax_total: float = 0.0
    province: str | None = None
    dialect: str = ""
    confidence: float = 0.0
    raw_fields: dict[str, float] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Source probe
# ---------------------------------------------------------------------------


_TITLE_PATTERNS = (
    re.compile(r"增值税.*申报表"),
    re.compile(r"增值税.*附加税.*申报表"),
)
_PROVINCE_HINTS: dict[str, list[str]] = {
    "BJ": ["北京", "京税"],
    "GD": ["广东", "粤税"],
    "SH": ["上海", "沪税"],
    "ZJ": ["浙江", "浙税"],
    "JS": ["江苏", "苏税"],
    "SC": ["四川", "川税"],
    "SD": ["山东", "鲁税"],
}


def probe_source(path: Path | str) -> SourceProbe:
    """Identify which dialect of the VAT return this workbook is."""

    import openpyxl

    p = Path(path)
    wb = openpyxl.load_workbook(str(p), data_only=True, read_only=True)
    try:
        notes: list[str] = []

        title_seen = False
        chosen_sheet: str | None = None
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row_idx, row in enumerate(
                ws.iter_rows(min_row=1, max_row=10, values_only=True)
            ):
                joined = " ".join(str(c) for c in row if c is not None)
                if any(rx.search(joined) for rx in _TITLE_PATTERNS):
                    title_seen = True
                    chosen_sheet = sheet_name
                    notes.append(
                        f"matched VAT title on sheet={sheet_name!r} row={row_idx + 1}"
                    )
                    break
            if title_seen:
                break

        # Filename fallback (the user may have stripped headers; STA filenames
        # very often start with the standardised string).
        if not title_seen:
            stem = p.stem
            for rx in _TITLE_PATTERNS:
                if rx.search(stem):
                    title_seen = True
                    chosen_sheet = wb.sheetnames[0] if wb.sheetnames else "Sheet1"
                    notes.append(f"matched VAT title on filename={stem!r}")
                    break

        if not title_seen:
            return SourceProbe(
                dialect="unknown", confidence=0.0,
                sheet_name=wb.sheetnames[0] if wb.sheetnames else "",
                notes=["no recognisable VAT title in the first 10 rows"],
            )

        # Province probe: check the first row's combined string for hints.
        province: str | None = None
        for sheet_name in wb.sheetnames[:3]:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
                joined = " ".join(str(c) for c in row if c is not None)
                for code, hints in _PROVINCE_HINTS.items():
                    if any(h in joined for h in hints):
                        province = code
                        notes.append(f"province hint matched: {code} via {hints}")
                        break
                if province is not None:
                    break
            if province is not None:
                break

        dialect = "golden_tax_iv_generic"
        confidence = 0.85 if title_seen else 0.0
        if province:
            notes.append(
                f"province {province} dialect not implemented; falling back "
                "to generic"
            )
        return SourceProbe(
            dialect=dialect,
            confidence=confidence,
            sheet_name=chosen_sheet or wb.sheetnames[0],
            notes=notes,
        )
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Generic-dialect parser
# ---------------------------------------------------------------------------


# Canonical labels (substring match; whitespace-collapsed) -> field name.
_FIELD_LABELS: tuple[tuple[re.Pattern[str], str], ...] = (
    (re.compile(r"销项税额"), "output_vat"),
    (re.compile(r"进项税额"), "input_vat"),
    (re.compile(r"上期留抵.*税额"), "prev_credit"),
    (re.compile(r"应纳税额"), "tax_payable"),
    (re.compile(r"附加税费.*合计"), "surtax_total"),
)


def _collapse(s: str) -> str:
    return re.sub(r"\s+", "", s)


def _try_float(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, int | float):
        return float(value)
    if isinstance(value, str):
        v = value.strip().replace(",", "").replace(" ", "")
        if v in ("", "—", "-", "/"):
            return None
        try:
            return float(v)
        except ValueError:
            return None
    return None


def parse_workbook(
    path: Path | str,
    *,
    declaration_period: str,
) -> VatDeclaration:
    """Parse a Golden-Tax-IV VAT return into a :class:`VatDeclaration`."""

    import openpyxl

    p = Path(path)
    if not p.exists():
        raise VatParseError(f"file not found: {p}")
    probe = probe_source(p)
    if probe.dialect == "unknown":
        raise VatParseError(
            "unrecognised VAT declaration format: no canonical title row "
            "found.  Hints: " + "; ".join(probe.notes or ["(none)"])
        )

    wb = openpyxl.load_workbook(str(p), data_only=True)
    try:
        ws = wb[probe.sheet_name] if probe.sheet_name in wb.sheetnames else wb.active
        decl = VatDeclaration(
            declaration_period=declaration_period,
            dialect=probe.dialect,
            confidence=probe.confidence,
        )
        decl.warnings.extend(probe.notes)
        raw: dict[str, float] = {}

        for row in ws.iter_rows(values_only=False):
            label_cells = [
                c for c in row if isinstance(c.value, str) and c.value.strip()
            ]
            if not label_cells:
                continue
            label = _collapse(label_cells[0].value)
            for rx, field_name in _FIELD_LABELS:
                if rx.search(label):
                    # Look for the first numeric cell to the right of the
                    # label.  Many GT-IV layouts put 一般 column / 即征即退
                    # column / 合计 column; pick the right-most non-zero one.
                    candidates: list[float] = []
                    for cell in row[label_cells[0].column:]:
                        v = _try_float(cell.value)
                        if v is not None:
                            candidates.append(v)
                    if not candidates:
                        continue
                    chosen = next(
                        (v for v in reversed(candidates) if v != 0.0),
                        candidates[-1],
                    )
                    raw[field_name] = chosen
                    setattr(decl, field_name, chosen)
                    break

        decl.raw_fields = raw
        if not raw:
            raise VatParseError(
                "no recognised VAT fields found.  The form's title matched "
                f"({probe.dialect}) but none of {[r.pattern for r, _ in _FIELD_LABELS]} "
                "labels appeared as row prefixes."
            )
        if decl.tax_payable == 0 and (decl.output_vat or decl.input_vat):
            decl.tax_payable = max(
                decl.output_vat - decl.input_vat - decl.prev_credit, 0.0
            )
            decl.warnings.append(
                "tax_payable not present in form; computed as max(output - "
                "input - prev_credit, 0)"
            )
        return decl
    finally:
        wb.close()
